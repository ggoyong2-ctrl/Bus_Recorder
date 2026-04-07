# ============================================================
# Seoul_Bus_Drive_Recorder v1.15
#
# 【프로그램 설명】
#   서울시 공공데이터 API를 이용하여 특정 버스 노선의
#   운행 출발/종료 시각을 자동으로 기록하고 엑셀로 저장하는 프로그램.
#
# ★ 전체 구조 목차 ★
# ══════════════════════════════════════════════════════════
# 【1】 라이브러리 가져오기
#      ├─ 【1-1】 파이썬 기본 라이브러리 (sys, os, time, threading 등)
#      ├─ 【1-2】 PySide6 GUI 라이브러리 (창·버튼·표·그래픽 등)
#      └─ 【1-3】 외부 선택 라이브러리 (requests, pandas, openpyxl, cryptography)
# 【2】 전역 상수
#      ├─ 【2-1】 글꼴 이름 (FONT_FAMILY, FONT_MONO)
#      ├─ 【2-2】 이미지 데이터 (ICON_B64, CC_IMG_B64, GG_IMG_B64 - Base64)
#      ├─ 【2-3】 API URL 상수 (URL_POS1 ~ URL_SRCH)
#      ├─ 【2-4】 노선 종류 이름표·색상표 (ROUTE_TYPE_LABEL, ROUTE_TYPE_COLOR)
#      └─ 【2-5】 지도 그리기 크기/위치 상수 (CELL_W, CELL_H, PAD_X 등)
# 【3】 유틸리티 함수
#      ├─ 【3-1】  _make_palette()         라이트/다크 색상 팔레트 생성
#      ├─ 【3-2】  darken_color()          색상 어둡게 조정
#      ├─ 【3-3】  truncate_name()         긴 정류소 이름 줄임표 처리
#      ├─ 【3-4】  fmt_bus_no()            버스 번호판 포맷
#      ├─ 【3-5】  format_remain_time()    초 → "X분 Y초" 변환
#      ├─ 【3-6】  format_hhmm()           날짜문자열 → HH:MM 추출
#      ├─ 【3-7】  format_datetm()         날짜문자열 → YYYY-MM-DD HH:MM:SS
#      ├─ 【3-8】  load_pixmap_from_b64()  Base64 → QPixmap 변환
#      ├─ 【3-9】  make_bus_pixmap()       버스 아이콘 생성
#      ├─ 【3-10】 get_app_bg_color()      앱 배경색 반환
#      ├─ 【3-11】 get_text_color()        글자색 반환
#      ├─ 【3-12】 get_base_color()        표 배경색 반환
#      ├─ 【3-13】 get_header_bg_color()   표 헤더 배경색 반환
#      └─ 【3-14】 detect_os_dark_mode()   OS 다크모드 감지
# 【4】 RouteMapPanel 클래스 (노선 지도 패널 위젯)
#      ├─ 【4-1】  __init__()              초기화·위젯 구성·타이머 설정
#      ├─ 【4-2】  set_sect_speeds()       구간 속도 데이터 저장
#      ├─ 【4-3】  eventFilter()           마우스 이동 시 툴팁(말풍선) 처리
#      ├─ 【4-4】  resizeEvent()           창 크기 변경 감지
#      ├─ 【4-5】  _on_resize_done()       크기 변경 완료 후 지도 재그리기
#      ├─ 【4-6】  _calc_layout()          창 크기에 맞는 셀 크기·폰트 계산
#      ├─ 【4-7】  _update_bg_color()      지도 배경색 갱신
#      ├─ 【4-8】  _update_tip_style()     툴팁 스타일 갱신
#      ├─ 【4-9】  refresh_theme()         테마 전체 갱신
#      ├─ 【4-10】 _rebuild_table_colors() 표 색상 재적용
#      ├─ 【4-11】 load_route()            정류소 목록 로드 및 지도 그리기
#      ├─ 【4-12】 update_buses()          버스 위치 갱신
#      ├─ 【4-13】 _speed_color()          구간 속도 → 선 색상 반환
#      ├─ 【4-14】 _draw()           ★     지도 전체 그리기 핵심 함수
#      ├─ 【4-15】 _build_table()          종점 도착 예정 표 생성
#      ├─ 【4-16】 _tick_countdown()       1초마다 남은 시간 카운트다운
#      ├─ 【4-17】 pause_tick()            카운트다운 일시 정지
#      └─ 【4-18】 resume_tick()           카운트다운 재개
# 【5】 RecordTable 클래스 (운행 기록 표 위젯)
#      ├─ 【5-1】  __init__()              초기화·표 구성
#      └─ 【5-2】  add_row()              새 기록 행 추가
# 【6】 SeoulBusRecorder 클래스 (메인 창 - 프로그램 본체)
#      ├─ 【6-1】  __init__()              초기화·시그널·변수 선언
#      ├─ 【6-2】  _make_info_html()       메뉴바 우측 제작자 HTML 생성
#      ├─ 【6-3】  _setup_ui()             메뉴·레이아웃·위젯 전체 구성
#      ├─ 【6-4】  _toggle_log_panel()     로그창 접기/펼치기
#      ├─ 【6-5】  _apply_theme()          라이트/다크 테마 전체 적용
#      ├─ 【6-6】  _slot_log()             로그 메시지 화면 출력 [슬롯]
#      ├─ 【6-7】  _slot_record()          운행 기록 표 추가 [슬롯]
#      ├─ 【6-8】  _slot_update_map()      지도 버스 위치 갱신 [슬롯]
#      ├─ 【6-9】  _slot_clear_map()       지도 버스 제거 [슬롯]
#      ├─ 【6-10】 log()                   로그 시그널 발행
#      ├─ 【6-11】 _make_fernet()          암호화 객체 생성
#      ├─ 【6-12】 _enc_key()              API 키 암호화
#      ├─ 【6-13】 _dec_key()              API 키 복호화
#      ├─ 【6-14】 _load_config()          설정 파일에서 API 키 읽기
#      ├─ 【6-15】 _save_config()          설정 파일에 API 키 저장
#      ├─ 【6-16】 _show_key_input()       인증키 입력/검증 대화상자
#      ├─ 【6-17】 _show_route_search()    노선 검색/즐겨찾기 대화상자
#      ├─ 【6-18】 _load_route_from_search() 노선 상세 정보 API 로드
#      ├─ 【6-19】 _slot_route_loaded()    노선 로드 완료 처리 [슬롯]
#      ├─ 【6-20】 fetch_api()       ★     서울시 버스 API 공통 호출
#      ├─ 【6-21】 _fetch_first_time()     첫차 시각 조회
#      ├─ 【6-22】 _ask_interval()         갱신 주기 입력 대화상자
#      ├─ 【6-23】 _show_api_status()      API 호출 현황 대화상자
#      ├─ 【6-24】 _show_program_info()    프로그램 정보 대화상자
#      ├─ 【6-25】 _on_toggle()            기록 시작/중지 토글
#      ├─ 【6-26】 _start_monitoring()     모니터링 시작
#      ├─ 【6-27】 _stop_monitoring()      모니터링 중지
#      ├─ 【6-28】 _main_loop()            백그라운드 갱신 루프 스레드
#      ├─ 【6-29】 _refresh_data()         한 번의 데이터 갱신 실행
#      ├─ 【6-30】 _process_routes()  ★    버스 위치 분석·운행 판정 핵심
#      ├─ 【6-31】 _record()               운행 이벤트 기록
#      ├─ 【6-32】 _perform_auto_save()    자동 엑셀 저장 (락 보호)
#      ├─ 【6-33】 _core_excel_save()      엑셀 파일 저장 (날짜별 시트)
#      ├─ 【6-34】 _axs()                  엑셀 시트 스타일 적용
#      ├─ 【6-35】 _ask_interval()         갱신 주기 대화상자 (유효 정의)
#      ├─ 【6-36】 _show_api_status()      API 현황 대화상자 (유효 정의)
#      ├─ 【6-37】 _show_program_info()    프로그램 정보 대화상자 (유효 정의)
#      ├─ 【6-38】 closeEvent()            창 닫기 이벤트 처리
#      ├─ 【6-39】 _load_favorites()       즐겨찾기 불러오기
#      ├─ 【6-40】 _save_favorites()       즐겨찾기 저장
#      └─ 【6-41】 _cleanup_search_signal() 검색 시그널 연결 해제
# 【7】 프로그램 진입점 (if __name__ == "__main__")
# ══════════════════════════════════════════════════════════
# ============================================================
# Seoul_Bus_Drive_Recorder_v1.15
# ============================================================

# ══════════════════════════════════════════════════════════
# 【1-1】 파이썬 기본 라이브러리
#   파이썬 설치 시 기본 포함 - 별도 설치 불필요
# ══════════════════════════════════════════════════════════
import sys
# sys        : 파이썬 인터프리터 제어 (sys.exit 강제 종료, sys.platform OS 판별 등)
import os
# os         : 운영체제 인터페이스 (파일 경로 조작, 파일 존재 확인, 디렉터리 등)
import time
# time       : 시간 제어 (time.sleep 대기, time.time 현재 시각 숫자로 반환)
import threading
# threading  : 멀티스레드 지원 - API 호출을 백그라운드에서 실행해
#              UI가 멈추지 않도록 한다. threading.Thread + threading.Lock 사용.
import colorsys
# colorsys   : 색상 공간 변환 (RGB ↔ HSV). 선 색상 어둡게 만들 때 사용.
import base64
# base64     : 이진 데이터↔ASCII 텍스트 인코딩. 이미지를 코드에 내장할 때 사용.
import re
# re         : 정규 표현식. 버스 번호판 "1234사5678" 형식 파악 등에 사용.
import json
# json       : JSON 직렬화/역직렬화. 즐겨찾기 노선을 설정 파일에 저장할 때 사용.
import configparser
# configparser: INI 형식 설정 파일 읽고 쓰기. API 키 보관 파일(*.ini)에 사용.
from collections import defaultdict
# defaultdict: 키가 없어도 기본값을 자동 생성하는 딕셔너리.
#              defaultdict(list) → 없는 키 접근 시 빈 [] 반환.
#              지도 위 버스 레이블 겹침 방지용 그룹핑에 사용.
from datetime import datetime, timedelta
# datetime   : 날짜·시간 클래스.
#   datetime  → 연·월·일·시·분·초를 담는 객체
#   timedelta → 시간 간격 (예: timedelta(minutes=5) = 5분)
import xml.etree.ElementTree as ET
# xml.etree.ElementTree (ET): XML 파싱 라이브러리.
#   서울시 버스 API 응답이 XML 형식이므로
#   ET.fromstring(text)으로 트리 구조로 변환하고
#   root.findtext(".//태그명")으로 원하는 값 추출.
from urllib.parse import unquote
# urllib.parse.unquote: URL 인코딩(%XX)을 원래 문자로 복원.
#   공공데이터 API 키가 URL 인코딩 형태로 올 수 있어 복호화 전 unquote 적용.

# ══════════════════════════════════════════════════════════
# 【1-2】 PySide6 화면(GUI) 라이브러리
#   Qt 프레임워크의 파이썬 공식 바인딩.
#   창·버튼·표·그래픽 등 모든 화면 요소를 만들고 제어한다.
# ══════════════════════════════════════════════════════════
# ─ QtWidgets ─
#   QApplication    : Qt 앱 뼈대. 프로그램 시작 시 제일 먼저 생성해야 함.
#   QMainWindow     : 메뉴바·상태바를 갖춘 메인 창 클래스.
#   QWidget         : 모든 위젯(화면 요소)의 기본 클래스.
#   QVBoxLayout     : 자식 위젯을 위→아래로 순서대로 배치.
#   QHBoxLayout     : 자식 위젯을 왼쪽→오른쪽으로 배치.
#   QSplitter       : 두 영역 사이의 드래그 가능한 구분선 (크기 조절).
#   QLabel          : 텍스트·이미지 표시용 라벨 (클릭 불가).
#   QPushButton     : 클릭 가능한 일반 버튼.
#   QTableWidget    : 행·열로 구성된 스프레드시트형 표.
#   QTableWidgetItem: 표의 각 셀에 들어가는 항목.
#   QHeaderView     : 표 헤더(머리줄) 크기 조절 방식 설정.
#   QPlainTextEdit  : 여러 줄 텍스트 표시·편집 (로그 창에 사용).
#   QDialog         : 팝업 대화 상자 기본 클래스.
#   QDialogButtonBox: OK·Cancel 등 표준 버튼 묶음.
#   QLineEdit       : 한 줄 텍스트 입력창.
#   QMessageBox     : 알림·경고·확인 팝업 메시지 창.
#   QGraphicsView   : 2D 도형·이미지를 표시하는 뷰 (노선 지도).
#   QGraphicsScene  : QGraphicsView에 그릴 객체들을 담는 "무대".
#   QAbstractItemView: 리스트·표·트리 공통 기능 (예: 편집 금지 설정).
#   QStackedWidget  : 여러 위젯을 쌓아두고 하나씩 보여줌 (노선별 탭 대체).
#   QStyleFactory   : 스타일 이름("Fusion")으로 스타일 객체 생성.
#   QTreeWidget     : 계층(트리) 구조 목록 위젯 (노선 검색 결과).
#   QTreeWidgetItem : 트리 위젯의 각 항목.
#   QTabWidget      : 탭 버튼으로 여러 페이지를 전환.
#   QSpinBox        : 위아래 화살표로 숫자를 조절하는 입력창.
# ─ QtCore ─
#   Qt      : 전역 상수 모음 (Qt.AlignCenter, Qt.Horizontal 등).
#   QTimer  : 일정 시간(ms)마다 함수를 자동 호출하는 타이머.
#   Signal  : 이벤트 발신 통로 정의. 백그라운드→UI 스레드 간 안전한 통신.
#   Slot    : 시그널을 받아 처리하는 함수 표시 데코레이터.
# ─ QtGui ─
#   QColor   : 색상 표현 (RGB, 헥사코드 "#RRGGBB", 색상 이름 "red" 등).
#   QPen     : 선 그리기 도구 (색·두께·스타일).
#   QBrush   : 도형 채우기 도구 (색·패턴).
#   QFont    : 글꼴 설정 (이름·크기·굵기).
#   QPixmap  : 픽셀 기반 이미지 (PNG/JPEG 로드·화면 표시).
#   QPainter : 도형·텍스트·이미지를 실제로 그리는 "화가" 클래스.
#   QIcon    : 창 제목줄·메뉴에 표시되는 아이콘.
#   QAction  : 메뉴·툴바에 추가되는 동작(이름+클릭 이벤트).
#   QPalette : 위젯 전체 색상 팔레트 (배경·글자·버튼·강조색 일괄 설정).
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QSplitter, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QPlainTextEdit,
    QDialog, QDialogButtonBox, QLineEdit, QMessageBox,
    QGraphicsView, QGraphicsScene,
    QAbstractItemView, QStackedWidget, QStyleFactory,
    QTreeWidget, QTreeWidgetItem, QTabWidget, QSpinBox
)
from PySide6.QtCore import Qt, QTimer, Signal, Slot
from PySide6.QtGui import (
    QColor, QPen, QBrush, QFont, QPixmap, QPainter,
    QIcon, QAction, QPalette
)

# ══════════════════════════════════════════════════════════
# 【1-3】 외부 선택 라이브러리 (pip install로 따로 설치 필요)
#   try~except 구조: 설치되어 있으면 가져오고,
#   없으면 설치 안내 메시지를 출력한 뒤 sys.exit(1)으로 즉시 종료.
# ══════════════════════════════════════════════════════════
# requests   : HTTP 요청 라이브러리.
#              requests.get(url, params={...}, timeout=10)으로
#              서울시 버스 API를 호출하고 XML 응답을 받아온다.
# pandas(pd) : 표(DataFrame) 형태로 데이터를 다루는 라이브러리.
#              pd.DataFrame(...)으로 운행 기록을 표로 만들고
#              df.to_excel(...)으로 엑셀 파일에 저장.
# openpyxl   : 엑셀(.xlsx) 파일 직접 제어 라이브러리.
#   XlFont    : 엑셀 셀 글꼴 스타일 (Font와 이름 충돌 방지를 위해 별칭 사용).
#   PatternFill: 엑셀 셀 배경을 특정 색으로 채우기.
#   Alignment  : 엑셀 셀 텍스트 정렬 (가운데·왼쪽 등).
try:
    import requests
    import pandas as pd
    from openpyxl.styles import Font as XlFont, PatternFill, Alignment
except ImportError as e:
    print(f"필수 패키지 누락: {e}\npip install requests pandas openpyxl PySide6")
    sys.exit(1)

# ══════════════════════════════════════════════════════════
# 【1-3 계속】 암호화 라이브러리 (선택적 - 없어도 실행 가능)
#   있으면 API 키를 암호화해서 저장, 없으면 평문으로 저장.
# ══════════════════════════════════════════════════════════
# cryptography.Fernet: 대칭키 암호화. 같은 키로 암호화·복호화.
#   API 키를 설정 파일에 그냥 저장하면 다른 사람이 볼 수 있으므로
#   암호화해서 보관.
# hashlib: 해시 함수 모음.
#   pbkdf2_hmac()으로 _SECRET 문장을 32바이트 암호화 키로 변환.
# _CRYPTO_OK: True이면 암호화 사용 가능, False이면 평문 저장.
try:
    from cryptography.fernet import Fernet as _Fernet
    import hashlib as _hl
    _CRYPTO_OK = True
except ImportError:
    _CRYPTO_OK = False
    _Fernet = None
    _hl = None

# ══════════════════════════════════════════════════════════
# 【2-1】 글꼴(폰트) 이름 상수
#   프로그램 전체에서 일관된 글꼴을 유지하기 위해 상수로 정의.
# ══════════════════════════════════════════════════════════
# FONT_FAMILY: 일반 텍스트용 폰트 - "맑은 고딕" (한글이 깔끔한 Windows 기본 폰트)
# FONT_MONO  : 로그창 등 코드형 텍스트용 고정폭(등폭) 폰트 - "Consolas"
FONT_FAMILY = "맑은 고딕"
FONT_MONO = "Consolas"

# ══════════════════════════════════════════════════════════
# 【2-2】 아이콘/이미지 데이터 (Base64 인코딩)
#   외부 이미지 파일 없이 코드 안에 이미지를 텍스트로 내장.
#   load_pixmap_from_b64()로 복원해서 사용.
# ══════════════════════════════════════════════════════════
# ICON_B64   : 프로그램 창 아이콘 (버스 모양 PNG)
# CC_IMG_B64 : 크리에이티브 커먼즈(CC) 라이선스 마크 이미지
# GG_IMG_B64 : 공공누리 마크 이미지 (프로그램 정보 창에 표시)
# ── 이미지 Base64 ──
ICON_B64 = "iVBORw0KGgoAAAANSUhEUgAAAgAAAAIACAYAAAD0eNT6AAAACXBIWXMAAAsTAAALEwEAmpwYAAAgAElEQVR4nOzdeZwk6V3f+c8TEXnVffXdPdMz09NzHxqNNLpnBOgAbA7J2Kwv7PUCZhdsvMbsgs9dL17WYK9tbE6B4SVsgw3YZpERWAId6NZII41mRnP1dM/03V33kUccz/7xZHZlV0VUd1flFdXf9+uVXVkRkZlR2VX5/OL3PM/vMYjIlqy1HlABhoGh5v3220zzNgaMA6PASPPrcPO2FzjW63PPqRPABWAFWAWWm/eX226zzWOqQA1Ya96vtu4bY8Ken7lIjph+n4BIL1lrDe733m/eAqAMlIBpXEO9r+3+BK5x349r7EvNWxkoNu+PAYVe/hxCBCwBdaCBCwLqbbcacBJYBM4Bl5u388Cl5v6w+TVu3hIAY4zt3Y8h0j8KAGRXsta2GukKriE/DtyGa6hbjf1Mc9sDzePk5nIR+CIuKLiIyxw0cFmHF4Gv4zIPNaBmjIn7dJ4iXaEAQHLHWttqxMu4K/RJYKp5Owy8GdewB7ir/OHmMaP9OF/JnRhYAOZwWYIIFwQ8iQsYLuKyCbPN41pZiIYxJunHCYtshwIAGWjW2gCXkt/fvB3H9aXPNLc/iGvcdyXbTEZbLIl131trSZr7bOvrlePcsY4hjCFp7oytJUqund32jKHguY+Ggg/Nu3itjwuz/sHhGTDN740xeDS/N2AwmN39CbMAPIMLBC7jAoNngFO48QlnjDGr/Ts9ka3t7j9PGXjNBn4Y148+hruKP4a7gr8buB93lV/Epe8r5LC/3VrXgEdJqyGGJLFEFhILcWKJm/taXxsxNGJLPbY0Ivc1bH5fi9bvR4l7PLgAoNXIW+vRiCFO3EVpmFjC5NoXqJ4xlH03VKIYQOC5D4qgGQn4nrkSFJR8Q8E3lHxDMWh+9aHY3FbwDUXfEHju8YEx+Mbg+xbfMwTGPZ/vgW/ccS6oyOVHk8UNQqzjMgfLuAGNL+KCgpeAV3FjFxaaX+sacyD9ksu/Msmn5gC8IdyV+5uAR3EN/nRz2z5gDznrj4+tZbmesFxPWKonrDQsq42E5UbCWpiwFkI1dN8vVBMW6wmLtYTV0FKPEqIEwtg2G2jX6IeJa/SjxDXo7sb693EzWLCW7D/jtHblev/k2x+b/RgXGLjG3W818m1ffWMIfCh4hlJgKHouICj6bns58BgteoyUDKNFj4myx3jZo1IwDBc8hgpu+3DRY6RoGCl6jBQ9xkoeBT+XH1/zuEGIZ3BZglVcUPAnuOzBgjGm3r/Tk5tJLv+CZLA1r+ongVuBW4CjuMb+bbjG3mN9BH7ffwettcQWqqFlNXQNeDVMWG24K+1aZLlcjTi3HHN5NWauljBfjVmsJSzVLdUoYbWRXHl8PWqm6688f/PrhvspZ9L82ve3pIvSf8bWd60Lf7PhvmcMpQAqwXpwUCm64GC85DFV8Zga8pkZ8pke8pmpeIyWPEq+YaTkAomhggsehgsuOGFwuigS3DiDGPcGPQd8Fjfm4Czr0yKXlS2QThqMX3/JrWZjH+Aa9ruBx3CN/R5c43+AHqfsrV1Ph7f6yRML9chSixJqkeXiasz5lZgzSzHnV0JeXXSN+0I9YaEWs9pwjXktbL/6doHCdXSjSx+0shG+Z/CNy0IUfRgt+QwXDGNll2HYM+RzcDRg74jPgRGf/SPu+5GSRznwKPsum9Ea39B67j51SyyzPqbgNeDTwFO4boU1INLAQ9kuBQByQ5oj8A/hBuMdx424fz1wV5/Oh3pkWQ0tqw3LUj1hrhZzYSXmtcWQy2sR55aTZoMfcX4l5tJaTKgJXdLGAJMVj30jAftHfPYO+xwYCZiseBwcCzg06rNvxGeCAiNjMFI0lAOD178UQhV4Gvhk8+vzuADhvKYryvVSACCprLVFXDW7YVx1u/ar+73N2yQuld/N86ARuwFsjdg19hdWYk4uhpxbjDm1FHJmKeLUYsxcNaYWJqyFtnlLCHVtJDtggErBMFQwDBc9KiUYD3yOThQ4PB5weMxn73DAkfGAw2MBoyWPgrc+ENL3evIRW8eNK5jHBQGt7oOXWK+guKbAQDZSACBXWGuHcQ3965q3NwD34Qbu9USUuAb+3HLE6aWIi2sxL86GvDIfcrbVB1+Nma8lSsXLwBguGKYqPlNDrkvhyFjA3XsK7B8pcmDEsG/E59BYwHjJ63VXwiVc7YKv4IKCp4ETKpMsoADgptWsbz8GHMQ19o/jRubvwV3xd3wkfmuwXRg30/WNhAsrMS/Phbww2+Cl2ZCzKxGX1xLm1mJWQ3fVH8bucTfHIDnJN/c7aoyh4LkxCeXADUicGfKZrnjcOlHgvr1F7poucGS8wHjJY7RkKAdeN6dBWlw2YBE3A+GTwOeAzwPzxpjlTr+gDD59kt5krLUHgUdwjf2DuP77g916vUZsma/GXF5LeHmuwXOXQl6YDXl10V3Rv7YYsRrqal5uLq0Bi/tGfPaPBBwac10I9+0tcN+eEofHffYMBQwXuz7O4GngBeDLuEzB540x8918QRkcCgB2sWZKfx9uGt7rgLfjGv5RXCndjvTfW+sK0sxX3Tz4cysxz11q8NzFOi/NRzx7scF8Lbky1z1Sv7zIJgZXebHguaJKx6YKHJ8qcHwm4NhUiePTBSYrPhPN2QwdHF9gcVMRa7iKhp9p3j6LW0jpnDEm6tSLyeBQALDLNAfvPYIbnf84bi7+MdyAvo5aaSR85Xyd5y6GfOVinacvNLi0GnN5zfXTq6EX6YyhgmG64jM15HFkLODRQ2Xu3VPkdQeK3DJeoNidokgXcFMQn8FNP/wk8JIGE+4eCgByrm1a3iPANwLfgrvi79TzU48tF1diTi9FnFqIePJcnc+ervHKfMTF1ahZela/SiK9NlI0zAz53LunyCMHSzy8v8QtzRkJM0O+K+Pc2S6El3DBwEdwWYKTyg7klz61c6i51O0M7ir/W3D9+LfiBu/t5HmvDNJbrCecuBzz1MUan3q1ytcvh5xfiZivJtRje2WBGf0KiQwG30A5MEwP+RwY9XnDwTKPHCzxuv1FDowGV8on+50ZZLiGq1D4FeBDwKdwiyFpbYMc0ad3TjQr7j2A68v/Rlx//pFOPPdqI+G1pYiXZkM+d6bGc5dCvnahzgtzDVz5fhHJq6JvuGMq4P49Re6aKfLGw2Xu21vk0GhAKejY3/cF4GPN22eAr2tNg8GnT/cB1TZN71bgHcATuHn5M+xgil5iLct1y+W1mGcuNvjM6SrPXgx5+mKdhZobxBer715kV/INjJU99g/73Le3yOsOFHhgX4XXHSgxVnILLu1w1kENlwl4HvgobqrhV3HrGKj2wIBRADBgmun9e4F34Rr9N7HD9e4TC+dXIr58rs6zlxp87nSNp87VObMcU4uUrRO5eRkqAdwyEfDw/hKPHSrz4P4iD+0rMj3kd6KrIMZlBZ4CPgg8ozEDg0MBwACw1pZxi+a8HfjLwEO4q//idp6vEbvlaU8uhHz1QoM/eKnKl87VOL8SNRe3yVqNTkRuVq3aBMNFjwMjPu84WuFNh8u87kCJQ2PNMQTbH1SYAHO4YOAPgT/GTS9c7dgPIDdMAUCfWWvfDvwPwFtxC+qUtvtcc9WEr10M+e8vr/KFMzWevdjg7HLUXDMe9N8tIjfCM3DreMCd0wXeeKjMu+8Y4sH9RcZ2XtK4VYDo54FPG2PWOnLCckPUIvRYszjPUeA7gG/FTd8rAN6NPlc1TDgxH/Lk2TofPVHl82dqnJhvECVGlfVEpKM84wYU3jYZ8NbDQ7z5lhJvPlLmlvGA4eINf3y1NHC1Bv5z8/Y8sKCZBL2hAKAHrBtKX8Zd5f913Gj+O7iBSnzWWiwQJXB2KeLDL63xhbN1PnWqypnliJWG/l5EpFcslYLH7RMFHjlQ4h0HhvnGe0ocHgvwPdewbCND8CpuauEHgA/j1ijQkOQuUgDQRc0iPceA7wbeievjv9HnYC20PD/b4NOv1vjwS2t89rUas1X9XYjIILAQWGaKAW+9pcJ7jg3x+gMl7ttbYKiw7a6CE8Dv4AKBTxpjGh09ZQEUAHRFM81/D/A/At8E3Hkjj0+spRZZXl2I+PL5Or/97ApPnq1zdjmioSKcIjLASr5b/vj1B0t8y51DvO2WCkfGAyqFbU0xvAB8HDeD4E+ARXUPdI4CgA6y1o4AjwHfC7ybG5y+FyeW00sRnz1d4/dfXOMjL69xZlktvojk18HRgHfcWubb7h7mDYfK3D4ZbCcQCHEFhn4N+LAx5mzHT/QmpABgh6y1PrAXeC+uf/8OXMN/XaNi4sQyV014YbbBf/36Mv/thRqvLkWs1BNN1RORXWO0aNg/GvDE0Qrvv2eYe/eWODjq3+iqhqvAJeAXgd8FXlCBoe1TALAD1tpJ3Lz9PwO87UYeGyeW8ysxv/PcCh96YZUvnKkzX7VYNfsissuNlQzHpor8xYdG+eZjQxybLhDc+PLGJ4DfBD5ojHmu82e5+ykAuEHNEf37gO/CXfEf4zoL9lhrma3GPHUu5INfWeKPXlnjwkpMqPF8InITKniuNPHjt1Z4/70jvP3WMgdHgxvJCkTAAvBvgV/HVRpUv+l1UgBwA5plev8MruF/ExBc72Pn1mL+8MQqH3xqlc+fqXJ5Ta2+iEjLcMFw/94Sf+mhUb79nhEOjd7wDILngf8E/Lwx5kx3znJ3UQBwHay1e4BvA/4WcN/1Pq4WJjx7KeTXv7LM7z6/ysvz6qoSEdmaZc+Qx7cdH+c776vwliNlJivXXTIFYBZXS+BXgJeVEcimAGAL1toibhrfPwAe5DpW4bPWVeh7YTbkF764yEdOrHFyISJKLHq7RUSuj+dZpsoBb7+lzJ9/cJR33THESNHcSPfAM8BvA/8UWNP0wc3UImWw1t4N/CjwftzCPNe0UIv5zMk6H/jyIh89UWWxoTS/iMjOWAq+4dEDZf7aw2N8891D7B/xr3cqYQJ8Avg54HeNMbWunmrOKABo0xzgNwH8BVw//71c4z1KrGWxlvCxk1V+5UtLfPZ0jdm1pDmaX2+viMjOuYv30aLH8eki/9Prx3jXHUMcnbjuAYMLwEeAnwSeVmVBRy1Uk7U2AL4B+BHgCdwCPVuKEsv/99Uq//75RT704hrVUBkmEZFuCzy4Z0+RH3pskm+/e4i9w9c9RuAsLhvwr40xC907w3y46QMAa62HW4b3/8HV6p+4xvGsNCyfPFXll55c4iOvrbK6imbvi4j0WCWAe/YW+asPj/Gn7xrmyPh1VRmsA88C/xj4vZu5kJACAGvfDfxfwBuufSw8db7Ov/ncIv/56yvMaUEeEZG+K/mG1x8s8f2PjvG+e0YYKV1XIdZ54JeAf26MudDdMxxMN20AYK09Anw/8Hdw6f7M9yJKLC/OhnzgyWV+9alF5qsq0ysiMmgCD544WuFvPDbBE7dVGL12IJDgsgF/B/jozZYNuOkCgGbK/1Hgnze/lrY4lloEv/LFFX7lK/M8db5OYm+6t0xEJFdmhny+4+5hfvjNE9w9U7jWQEELnAN+CvjZm2mA4E3VmllrR3FX/f87ML3VsVFi+fCLa/yrzy7y30+uQHJdKSURERkQe4d9fuANY/y1R8Y4Mn7tcd24mQJ/zxjzZPfPrv9uigCgOb3vMG7Qx/uA0axjk8RyajHil7+0xAe+tMTFlVjpfhGRnCr58JYjZX7sbVO89dYyQ8VrXsw9gyse9JvGmHr3z7B/dn0A0Gz83wn838Abtzq2Hll+7/lVfuKT83z53K7+fxcRuamMlTy+56FR/tZbJjg6EVxrnYE1XBDwL4wxi705w97b1QFAc/GebwV+BtgPpIZ+ibWcXIj42c8v8stfWmKhptH9IiK7TeDBg/tK/L23T/Pe42UqhS2zATXgd4C/DVzYjaWEd3sA8MPAjwN7so5JrOWPTlT58Y/O8uVzdSK1/SIiu9rUkOF7Hxnnx98xxdjWMwUS4KPAj+3GcQG7MgCw1u4H/jfcgL/MBXwW6wn/7FPz/PwXl7i0qgWjpD/a/wjbs5Km7Y655rFm03bfuPsGXHEU41JgrQHRvtd8lIEg45PAGPAzUqXGkFl0pfXa2fu699ETJ3bH43YS624AFuvuWzdcPLa4oiCbjnOvfeU8msdb6y403LmtFw2zbSfZOuP2bYltvpZ0hd+cMvj3H5/ibbdU8LPjgAQ4Bfwg8Ae7aXXBXRcAWGtHcFX9/ipbNP6vLYb8xCfm+bdfXqKxa/475fpZ8OyWszs84xorr9nQeZ772r6tdd/3wKN1DBQ8Q8E3FH1DwYei5+4HnqEYQKF5fyhwq5sVfdcYF32uTFkaKhjAEHhQ8Gnu9wi85v3ANM/RUPLdYwq+e22AoLnNtBr45rGtACBoBgDGuPtp3L7s92erACDtKQ1tgUeXxG0N7na1N77WQmzXG+j2RjxO7JUAIOHqACBM1h/TyizGiXUBAdCI7JXXqcfugDCmuXIoXFyN+fK5Oh87WWVNZca75thUgX/8zined+8wxaxfdudV4PuNMR/u0al13a4KAKy1B4B/BnwXEKQdE0aWT75S4wc/fInnLt800z1zpvVh12zUPNYb07b7gee+DzxD0Gz4As9QDgzlgmG06FEODMMFQykwVAL3fSEwroFOAgql2B0fuOcs+e5+sOF1Wq/t7rvGtdi8+QZKgXelMRXpFGstf/RKlR/80CW+fvmmqlHTUwXf8NcfHeMfPD7FzNbrCszhVon91d2QCdg1H1fW2hncSP+/DBTTjgljy29/dY2/+5mLnLiQ+/+73Ak8dwXqe+tXopWCYariMz3kM1V298dLPsMFn0rRXmnIR4oeoyWP0aJhuOhR8t3VdoCH70Hgrz9/4De/eu5q0zfrr+t1Of0s0mmJtfzWsyv8xd+6QKgxSl1T9OH994zwU++Z4eCov9XnxGXgfwF+O+9BwK74JLTWTgL/EvhuMlbxWwsTfuazS/yTz11mabmnp3fT8g0cnSxw/54i+0d9Do8FHBkPODASsG/ENepDBa95de6u0q9zjW+Rm8pCLeZ9v3GeP36l2u9T2fXecqTMT717mjcfKW8VBFwE/rYx5td7eGodl/tPW2ttBTfS/++QUtbXWstizfJTn5rjX35ukdVGwi74sQeeZ+D7Xj/GDz02wcExn0pg3FW/l91vLCLp4sTyjz42x098fF6FybrMAMemC3zg2/fyliPlzPExuEzAnwM+ntdMwG6ob/s3m7fUmv61yPIPPjbL//vZRVYbFjX+vfHOoxV+8l3T3Lu3yETZpxR4BL6u8EW2w/cMRycKlLKma0jHWODF2ZDv+Z0L/JfnVrc6dAb4ZeDNPTmxLkgdKJcH1lof+Dbg7wNDacdcXov5Pz82x89/YVF9Zz00VfH44TdPXGt+7W60hhvgbZtfr+cWtx1P2+Np25e2v31b+/2V5nmAW/e89TqN5nFR80ZzW+s1skpfxs3j0rQ/10at597INvd18y+yCGw5kus6FFj/fPRYXzHUNJ+/9csdtB3nc/X4o3LK9ta5FYCJtudsPZ8H3AFMbTyh4YIbzyK9cXIh4vt+9yJRYnnfPSMU04OvW4F/Ya391jwuKZzbAAB4HW7QX2rjHyWWv//RWX71qWU1/j12aCzgkYOlQR1sdxpYxjVcYfNWx1X9ar8ft+2PcI1WtGFbyHpD18CtLx6y3rDHzeM23m99TZrHt/bTtp3mPtv2fGw4NmI9GGg1xLExRr/xOWCtDVhv/FsBiw/8F+AbNh5fj9enHEpvzNcS/taHL1MKDN95z0jaIQZ4CPhH1tofNcbkaoRZLgMAa+0e4F8Dx9P2L9cTfuITc3zgS0uq7NcHheZUvG2IcMtyzrPe4LYa2qjtdgF4AddQV1lvtGsbtrXva2/c26+c2XA/bVvm/d1YHlR6wxizKXvSXLH0tpTtXFiJqasyUM+dX4n4gd+7RNE3vOfYUNqYgAD4XuCMtfYn0/5fB1XuAoBmff//FXg9KR36cWL5hS8u8rOfV+PfL8v1hAsrMVOV1CxsHfgS8Ce4hn4V11BXcenr08AC2QFAmKc/MJEb9C5SSpfXY3hpLlQGoC8MF1Zifvj3L/PB9+3jscOp2U0f+AHg08Af9fwUtyl3AQDwHuCvk3LuYWz5taeW+Id/PKfKWX302lLEJ09VuXumkPaHUsL1m/07XI3tl9Sgi4C1dhr4IWBTrnmpnvDJk5oC2E8vzYV83+9e5Le/ez93TqeWmjkI/KvmeIBTPT69bcnVKC1r7SHgH+IGz2zYB0+eq/N/fGyeNXX691Utsvzyl5ZYqmf+PxwEfhr4r8BvWGv/qrX2IWvtQWvtmLW2bK0tNJdyFtmVrLVe8/e8ZK0dx606lzqi/GsX6rwwp0qA/fbspQY//pFZ5quZs/7uA77fWptaj2bQ5CYDYK0tAv8z7g3e5MxyxI/+4WVOL0Voql//ff5MnX/z+UV+5K2TFNOHLpdxYziOA+9vbjsHPA2cx3UDLFprz+D6+5dxXQSruFHuK0BDffAyCJrBagGX4Sq23W99LeEGLA8Bk8Cdbd8P40b+v4WU6cyN2PILX1xS+n8AxBZ+69lV7t27wI+/fSprWub34roCfq+3Z3fjchMAAI8Cf4OUP5BamPCPPzbHZ16r9f6sJNPPfG6RY1NF/uz9qaNn0xxo3trVcY19axxAaxR9BJy31p7GjR9YxgUHra9V1gcEtn9dZvNUO7h6+l7a17RtLTbl+43pj+085nps93HXo3162lY8NkfdGx+XdcxVCxy2fe9tcX/juaXdNyn3wTW4Bdan+QXN+z5XN95F3GJirftF1hvy4oZbAZfVmmm+XtrNb7sVcY3+Na9UrLX88StVPq70/0D5uc8vcf/eEt91X+pn2zTwI9baTxhjlnp8ajckF5fKzZGxvwq8L23/h15Y5X2/cU6r+g2gI2M+v/PdB3j0UPnaB/fGHC54aJ9q15qOZzd8TdqObZ/O1z5Vj7b9La1pee0NfLSNx1yPjc/TSR7Xd5EQcHWD37oavtYxrWlwLa0GsnW/1Xi3z8Fvn5sfpDwuaNvnt71G67kmr/NnGghrYcKf+nfnVAJ4AN0+WeAL33eYqaHMkhN/3hjzH3p5TjcqLwHA+4FfJKU4xom5Bt/1H8/zpXNa2W9QHZ0I+LF3TPJn7h1hsuwNan0AkYGyVE/4R388y7/4zKLK/w6oH3xsjH/6rhkqhdRE2ReBP2eMOdHj07puAz8IsDmY4q+Q0vgD/NKTy3z1ghr/QXZyIeJH/mCW7/vdi/y3F9dYbSTYHa7XLrKbhbHlJz8xzy89uaTGf4D9xtMrfDq76/l+4N09PJ0bNvCXYtbatwCfStv3zMU6D//ca5rvnyOegdsmC7z7jiHeeKjEIwdK7Bn2mSh7lAOj7IDctKy1rNbhyXM1fvrT83zoxTUUJw++t91S5r9/z0HKQer19HPAw8aYgbxKHehP2+bV/y/iMgBXqYYJf/P3L/NLTw70GAvZwlDBcGA0YLricc+eInfNFDgy7rN/pMDeIZ+ZYZ/hokdgXODgGYNnwDS/N6CAQQaCtdaN6rRc9RXbHP1pLUmyPmKzEVtWGwlL9YTFWsL8Kry21ODjr67x6dfqnFmKdOWfE5XA8Ovv38d33jOc9nlUB95ujPlCH07tmgZ9MMzdwFvTdnzhbJ3/8txKj09HOmkttLw8F/IybtqgYzBYRkseUxWPibLHVMVlCMab96crHnuGA8bKhrLvUfShUnAZhNat6LffoOgbCr5bklgxw80rsZBYS5xAvOFrlLg1RNzNNdKt+1FiCeP1/WEC8ZohLCbE1lKPLNXIslpPWA0ta2HCSiNhLbSsNSzVKGFlDVbimNVw/ZjlumWlkRCTgG2frCB5UY0sv/jkIo8frTC9eUBgAfhWQAHANryD1LrY8B+eXubSmnL/u4+7klqqJ81CQq3rIPfB6BvwPfCNcV89t0Ja4BkCz30tBYZSWwBQ8g3FoBkEeG5JYh+LaR7rGSj5za9XvvfwPfcY34NC87GF5n2/LYrwPSi21Qc3bc/T4h5D22PceV15DFAMzA0Pygma59QNsXWN4LVsXKQmsVCPk6vS12FsiTceE9mrjmk1rOBeN25eRtdid1wCNCL3gDixV2b9hM1GufU6rca7df5J4n4Wa922xLpj4tbXxJ1/2NbIh83njxL3+OTKY5tX8xYSz5LEzaCidXXffP7kyrHtC0e0vklr5Ad+OJZs4eMnazxzqcE7bq1s3GWAh621Y4M4JXDQA4C/SMo5XlyNrrVOs+waV39YxhbiGNLX7unN2WzMIKRuM5snuF+931y9zcB2r/9MF68a7XW8v62Ud7sEe9V/zcZnSXuMvfohV/bv/H/46iCyb1TYcteqRZbff2EtKwC4D1ffRAHA9bLW3g08uHF7Yi3/8ZkVZtc06V96L63hurLjhp+pU9RbvDU1vNJ9H31ljUY8lVb59DbgCPB8789qa4Ocd/pWUqr+LdYS/vClNVTuX0REBsWZpYgXZlMH+/vAYz0+nesykAGAtXYS+AZSzu+l2VBlMUVEZKDMVxO+dLaetfvxXp7L9RrIAACXLnmIlNzdx1+tstxQylNERAZHNbK8cDkkTl+16U5r7XUvitIrgxoAvBE4tHFjPbL8p69p6p+IiAyepy7UWQ1TA4AxUsa09dugBgDvStv49IU6pxa6te6JiIjI9n31fIPVRuoAtWEUAFybtXYcty72VRJr+eSpGrNVjf4XEZHB89pSxOml1IvUEvBAs7rtwBi4AAB4My5dcpW10PLp0zXV/RcRkYH12dOZAwHvJaVt66eBCgCstQY3XWJ4477lesKnXtXofxERGVxfOJO5OuDdwGgPT+WaBioAwEVHj+DmTV7lqfN1zi0r/S8iIoPrhcsNluupqep9wOEen86WBi0AGCdjoMTHXtHVv4iIDLbZapI1DsAwYAMBBy0AuAs4unHjaiPhk/SMPCgAACAASURBVK9mplVEREQGwqW1mJfnwqzdb+rluVzLoAUAqeUSX5oLOb+s6X8iIjLYlusJry5G2NRFQzg+SDMBBiYAsNb6pJRLtBa+dqHBWfX/i4jIgEssfO1inYyVtGeAW3p7RtkGJgDAvSlHN25MrOXLF+rUr2NdchERkX576nyDOH3K+jhwR2/PJtsgBQD3AFMbN0YJfPqUBgCKiEg+fO1CnTD9olUBwEbN+f/3ARMb9600Er58PnWJRRERkYGz3LC8PJ86ELAA3GWtDXp8SqkGIgAAisD9pJzP0xfq1CKl/0VEJD+evpBZEfBuXJvXd4MSAJRxJYA3+ajm/4uISM588WxmAPAACgCuMgnctnFjPbI8fUHpfxERyZcT8yH19MVr9pEy3q0fBiUAeCOwqU/k7HLEqYXMggoiIiID6exyzPmV1OnrPvBQj08n1aAEAKkFgM4uR5r/LyIiuXNpNebSWmb7dW8vzyVL3wMAa20JeDht34n5kEurCgBERCRfLqxEnM++gH24Ofutr/oeAOAKAO1L2/H503U0/l9ERPKmHsMLs5lj2A7ixr711SAEAIeBvWk7njyXOYpSRERkoH3tYmYAMAMc6OGppBqEAOA2YHrjxlqU8Gz2myciIjLQnjqfeRE7g8sC9NUgVCO6m5RA5KW5aJcWANqNP5PkU9+7IEV2tRNzEfXIUgo2/a2NA4f6cEpXGYQA4I1pG790tkaU7K7GsujDB993gHv3DEQNCLnp2V0RZFugEVvSp1xfrRolpK/Sui62llp47VA9jC1nr2OZ8jB2zwkQJxA2P9eshXpzxRhroRG717RAGNkrq8lFib3qMe0l5pO2H6b954qte544sVjr1lSx1hImYLFEsVu1bjVMWGlYosTdwtgd07ofNx8r21ONEp691OB1B0obd/lkDH7vpb4GAG1rAGzcznOXQnZZ+897jw3znmMVxst+v09FRHrgWsGGTQkz7JV/2r7PeL7kquPsVdutbX6lGQiwHgjEidteDS3VKCFK3DFRs/F3QUfrvgs0whhqsWWtkbBUT1huWJ4+X2elkbDUsMxXY5YbCVEz4IkT9zWMXdBjbfvPYrkZMlCJha9fTg0AYACmAvY7A3AY1xdylbXQ8uJcY1clyysFw3c9MKzGX+QmYq7RxpmsRnBbbWN/GlS7ISNRjy1roWWhFjNXTXhxNuTkfMhSPWG+lnB+JeaZCw0urEashrvpU36zOGGrsWwP9PJc0vQ7AEithjRXTXht8dqptTw5MhbwHXeN9Ps0REQ6qj3I8Y1hyDMMFWBmyF3svPFQGWtdBiBMYC1MmF1LWKjFPD8b8rnTNT5xssrp5YjZtd3V32CBkwuZ4wD2WmunjDFzfTg1oP8BwCNpG2erMacWdlcA8ANvGGekOAiTLkREessYgzFQ8qAU+ExWfKDAo4fK/PkHRlhpWJ48W+e3n1vh155aYrm+ezIDp5dCFmox+0Y2NbcerhvgT3p/Vusn0BfW2iIp/f8AZ5YiLu6iCoDHpgq859hQv09DRGTgGGMYLXk8frTMT797hg/9hYO89ZbyrhkhcHopZr6WmdnoazdAPy9Jp8mYB/nMxd3V//9Nd1S4c7rQ79MQERlYxhhKgeFtt1T41e/Yy4HR3TFe6txyxGJ2AHB3L89lo34GAHuBI2k7tlhHOXcqgeFH3zZJ4O2WeFZEpHuMgTumivybb93DSDH/n5uroeXkfOaqtkettcO9PJ92/Q4ANpVCTKzlmV1UAfDxoxWOTujqX0TkehkDT9xW4U2Hy/0+lY547nJmm7YPmOrhqVylnwHAvcCmijinFiJms5dQzJXhguGH3jS+a/qyRER6ZaLs84OPjff7NDriqxe2XBRo01T4XulnAHBP2sZX5sNdUZ0M4NFDZd54aHdEsCIivfYNtw2xfyT/YwFeuBxeVbWxzR5S1sLplX4GAKkzAF7eRQHAE0crV+bCiojIjRkuGr7/0bF+n8aOzVZjzq+kZrbLwO09Pp0r+hIANKcAHtu4PbGWk/MRjXh3BAB/6aHRfp+CiEhuGeANh8pMlPNdQ6UWWV6eyxwIeH8vz6Vdv97VW3GRz1WqoeXkYrgrpgAeGQ84uEumsYiI9IMxhsdvrXA859Ooa5HlRPZMgL6tCdCvAOA4KQMA3XSJ3VEB8M/eN0LR1/A/EZGdGC4a3n5rpd+nsSP1yPLSXOY4gPuttX2JcPoVANwGbPqBq+HuWAPAAHdOF9DUfxGRnTHG8F3353sdFQucXnRrAqQokVEUr9t6HgBYaz1c9aNNAcBcNeHMdayvPehumQh4w6Ey5lpLgYmIyDXdOh5wdKLfS9fszIn5iJVGagBQAO7q8ekA/ckADJNSAAjgpbnGVetb59WBEZ+7Z/LdZyUiMigmyz7ffGe+11N5bTGkGqaWBC7Qp5kA/QgAxnGDADd5OrtYQq4cmy4yVMj3qFURkUFR9A13TRcJcvyx+tpSxGI9MwC401rb85RxP97OEeBw2o6vZ5dLzJU/fTzfkaqIyCAxBt6R87oqUQKvpA9yN8AhoOcNRz8CgGlSSh/WI8uJXTID4OH9pX6fgojIrnLf3mLu6wFscZF7GOh54Zh+vJv3ApvCuNNLEQvVzCUTc+O2iYDxnP+SiogMmqKf/+mAz1zKDACOcJMEAKnrH59Zilhu5D8AeOdtFQUAIiJd8O13923l3I54aTYkTh/pvh83Pq6n+pUB2OS1pYjl9AESuXL7VIGSCgCJiHTcreNBrrsBZqsxF1dT1wQoAnf0+HR6GwA0qx1tGgBoreXccpT7RYCGi4ZjUwXN/xcR6YIDowGPHc7vCqsrjYSz2bVujvfyXKD3GYCDpPRzhM3Rkflu/mGm4vOWI/n95RQRGWSTFY97clxjZamecGohMwB4oJfnAr0PAA4Bm0ZxRInl1GLmQgm5MVQ0HBjNd7UqEZFB5RnDg/tLFHM6G3C1YTm9FGHT1wQ43qyU2zMDEQCEseXkQv4DgHtmigRaAEBEpGu+6fZKbgutJRZeWQiJ0oe7jeEGA/bMYAQACbyanRbJjffmvFSliMigOzgWUAnye6H18mxImD4TYBi4pZfn0rMAwFrr4wKATTnyi6sxy+mLJOTKPXs2rXAsIiId5BvDW2/N71irl+cjwtSJAAyxWwMAoAwcTdvx/C4oAbx32GOqktOOKRGRHPlTx/NbD+Dl+ZBaeh9ABbitl+fSywCgiMsAbLIb+v8f3l9m37ACABGRbjs4GuR2YaB6ZLOWvfeBQ83p8j3Ry7ewREYG4OuX8h8A3D6lEsAiIr1wbKrAA/vy2+X67MXMNu92XLa8J3rZYo0AezZujBLLq4v5HgBogOmKj68CQCIiXXdgxOfoRH7rAbw8v+WaAD1bTa6XAcAdpAwAnF2Lma2mj4jIi0rBLVKh9l9EpPvKBY8jY/mtufLsxZD0UgDcTspMuW7pZQBwLG3j5bWYxVq+1wAo+YbX7c9vOkpEJG++JccDAc+vRKykL343Qg9rAfQ6A7DJ5bWYhZwHAMXAMKMBgCIiPfPYoZ5lyjtuvppwKX1RIOjhmgC9DADuTNt4cTVhoZbvLoBHDpTwlP8XEemZ8bKX24HXc7XMVQEB7urVefTk3bPWjgB70/adXYpo5Lv95xtuUwVAEZFeMsbwztt61l3eUQvVhIur/V8VsFfh0z5cneNNXpjN/xTAu3K8OpWISF49cTSfAcBquOXst1uttT25quxVADCFK3O4yemlfAcAw0XDWCmfaSgRkTzbO+zj57T39cR8ZgAwSsqU+W7oVcu1B/dDbbLFm5AL984UOTqR3+koIiJ5dddMkVvG8/n5+9ylzFoAE8DBXpxDrwKAGdxKR1dJrOXUYr4zAAfHAvZoBoCISM/dPlngYE7rAWzR/T2G6zbvuq4HANZag4tmNk2Un6smLNfzvQrgRNmjktei1CIiOTZeyu8ibCcXQmphZi2Aw704h160XB4ZSxye2gWLAD24r6QKgCIifWAMvPVIPpcGTqxbGjiFAW631na9fe5VAJAazeR9DQCAJ3I6DUVEZDd4PKczAQBenM0cB3AHbnXArupVAHBr2o4T8/nPANw9rSmAIiL9cnQiIK9J2OcuZ7aBx+lB+9znLoB8ZwCGC4ZKQf3/IiL9UgpMbhcGenmugU1fFego0PWry160XlO4aQ1XqUUJZ5byHQC8+9iQ+v9FRPpoqODxTcfy2Q1wdjlmKX1RoDIZmfNO6kUAcCRt40rd5n4RoEcO5HcxChGR3aDgw20T+eyKnV2LWcpuB2/r9uv3IgA4mrZxsZ4wW833IgCHc5p2EhHZLTxj2DPkk8fZ2BdXY+armQFA1xcF6sVblhrFLNUT5rJ/8IE3XDBMD+XwN05EZJd5cH+JPUP5qwdweS1mPjsD0PVFgfqWAViqJ8yt5TcDcGQ84M7pTbWNRESkx+7eU2AqhwHASsNyfjlzLNxha21Xixx0NQCw1haBA2n7Lq/GrIb5rQK4Z8jn4Ki6AERE+m285DFcyOeI7BfnMqcCTjVvXdPtDMAYKTMAAF7N+QyASsEwUsznL5yIyG7iGcOD+/KZkX156wBgspuv3e0AIPMH2OKHzoVjU0U8zQEUERkI771z03pzufD17GJAM81b13Q7ABglYxngsznPADx+NJ/1p0VEdqPbJ/M5FfDUQkgYp3aHjwN7uvna3Q4AxsnoAsj7MsAP7VcNABGRQTFSNIyV8peVXQstp9MviH3cmgBd04sAYGTjxiSxnF/J7wwAz8BkWVMARUQGxUTZ54F9+bswixK71QVxrgOA/aSsaDRbjanmeAbA7ZMFikH+Ik0Rkd1qtORx7578DQQME8vJ9GWBAe7s5mt3OwBIXQTo7HJMlOQ3AHjT4RJlBQAiIgOj5MNMDmsBhIlbGTdjUaDj1tquNTbdDgAOpW08vxKTPuYhH26fKlLwFACIiAwKYwxHxgOKOYsBrIUzSzGN9F7xSbpYC6DbAUDqQkDnlqNcZwAmyx5q/0VEBsubD5cZLeZvfNbZ5YiV9FUBPeBYt1632+/UwbSNF1Zj4pwuAzBUMOwd9jGqASAiMlBumyzksnv2wkrMWvq4OI8uLgvctQDAWlsADm/cnljLmaUot10Ae4Z9TQEUERlAYyWPoUL+MgCnlyKW0zMAPl1cFKib79Q+YFMoVotsrlcBHCl6HBrLWSeTiMhNwBh4/cH8zQSYq8Ys11PbRYNbFKgrbXU3A4BDac9fDS2Xc7wKYDkwjJXyF2GKiNwMnjg61O9TuGGxdRUBUxjcxXRX6hx3syU7mPb8tcjmehngg6O+1gAQERlQRyfzuUrriexaAHtIKajXCT3PAKyFCZdyHAA8cbTS71MQEZEMYyWPkp+/i7SXshfI24tbWbfjuhkAHCZlDEA1yncXwIM5LDUpInKz2DsccOd0/hYGenkuJL0WEPvIUwBgrfVxUcum51+uJaw2cjoFANg7rAGAIiKDanrI4+hE/gKA8ysR9Sh1IOAYXSoG1K0MwDAZJ/zackRem/+JssdQDotMiIjcLEaLHnuG8/c5vdqwnF3OzI53pRhQt96lIVwJw03OLGYOdBh49+wp5nK5SRGRm4VnDOM5XK21FlnOrWS2j0e78Zo9zwC8mucAYKbImDIAIiIDyxh425EKQc4+qqtRwumlzAxAV4oBdestqgDjaTvOr+Z3AOD0kKdlgEVEBtxDB0r4OZuuXYss55ajrFUBj3bjNbsZAEyk7Ti3nM8MgGfc9BLVABARGWxHxoLcLdgWJW5RoIwy+dPW2tSL6p3oVgAwSkrlojC2XM5pBmCoYHhQawCIiAy8om+YqOSsDwB4bTGiHqVGAEUyVtfdiW69Q/vTnnu+FlPL6SpAQwWP+/fmr8a0iMjNxhh485Fyv0/jhp1Zimikt5ElUhbX26luBQCH0jZeXouzopuBV/QNB0fyWWJSRORm86ZD+QsAXl2MqKcHALnKAOxN2zhfTQjz2QNAOTCUNABQRCQXDo3l74Lt/ErmRXKRjAvrnehWAJCaqri8FmdFNwPvDYdKaPyfiEg+jJY8CjkbBlCPbdZUeQ84ZK3t6EC0br09+9I2zlWTrP6NgffGHKaTRERuVgdGfQ6M5i8L8Mp85qJAB4CONkQdDwCstYaUVIW1bgZAXgOAPC4uISJyszoyFnA4h90AL81lTpU/hJti3zHdyACM4EoBXyWxMFuNSXLY/rdqAIiISD6Ml30mclgS+MR8I2vXAXIQAOwBNl0uh0l+lwHeP+Lnsra0iMjNquQbhgr5G7j12mJElJ4pn8HV2OmY3gUAMVxcTV3qcODdNVNk/4iWARYRyQtj4O6Z/NVuWaglXEq/WPaB2zv5Wt0IAKaBTR0vYWKZreY3AzBWUgAgIpIneSwGtNLIDAAAbuvka3UjANhLagbAcmklnwHAWMlD7b+ISL7cm8PqrUt1y8XstvJYJ1+rGwHAFCkZgCghlxkAgysoYVQEQEQkV6Yq+btyW20kzFZj0hcF5JZOvlZHA4DmFMB9pGQA1kLLfDV/YwAKPrzj1o4OvBQRkR4oeLAvZ+O3qpHl/EoMpEYAh621HUtrdDoDUAQm03ZcWInSf5wB5xnDcdUAEBHJHd8zvOFg/sYBnF6KsqbMV3AD7TuiGwHAdNqOcyuZxQ0Gmmdg73C+IkgREQHfwH1783cBd2ohMwAYImOtne3oYQYgf/3/4Pr/fU/9/yIieeMZ2J/DVVxfW8xnAFAAJtJ2XMxpEaB33Jq/9JGIiAAYhoseebuEe3UxJEkfBTjEgHcBTKXtuLCczwDgkQMdXXxJRER6xBg4NJq/Sq7nV2Kq6csCl4GDzQH3O9azDMAWhQ0G2sxw/tJHIiLi3DpRYHooX+O4Egtnl1PHzbVm2nXkB+p0ADAMjG/cGCduJcC8GSsZpiv5ihxFRGTdwVGfiRwu5vZy9qqAt5JSa2c7Ov2upKb/VxpJVjpjoB0ZL+RyOUkREXFGih6lIG+jAOC1pTBr1wEGNADYn7ZxsW5zGQAcGPE5OKoAQEQkrwq+YaSYvwDgRHYG4BYGNABInZ6wVIuphfmrAjha9BjO4XKSIiKy7qH9+RvM7YoBpV447wc68gN1OgBInZ6w3LDU8pgBGAvwVANARCTX8jiba74as9pIbTcDMrLtN6o3GYB6kssA4HHVABARyb07pvJXDXCxnrBYy8ycd2RRoJ4FAHkcA5DHtJGIiFwtj6sCLtUSFuuZAcDhTrxGxwKAtpUAN25nsRbnLgNggAMaACgikntFHyo5mwkwX0tYqGVOn7+tE6/RyQxABVemcJOFWpJV13hgTVU8CvmbOioiIhuUAo9bx/N1QbdcT1jKzgAc6sRrdLKJm8CVKbxKYuFCDosAPXFbhUADAEVEcm8oMNyzt9jv07gh1chyKbvt3G+t3XEfdScDgHHcWgBXsZDLKoD37iniKwMgIpJ7xcDksqbLa0uZbecIGWX3b0RPMgBbRDEDa/+In7sVpEREZLPAI5dl3U8tZFYDHAMmd/r83c8A2PwtBFQJDNNDPsYoBBARyTsDlHNY1O21xcxqgAOXARgjpTqRxTKbswBg34jPXTP56i8SEZF0xhj2DAWU/HwFAaeXMgOAUQYsAJggJQBILFzMWRfAzJDPbRP56y8SEZF0d80UGC3lKwA4sxyRpE+hGyZj8b0b0ekugE2tZj2yrKSXMxxY5cAwmsPlI0VEJN3hsYChnM3troWWhfRqgAEZpfdvREfeDWutD0zD5nFzl9Zi8tX8w90zRTz1/4uI7Bp7h/3cdQEk1nJ6ObMb4MhOn79T4ZBHRn/EXDV/qwC+/Wil36cgIiIdVA4MXr4SAK6OzkpmF/qBnT5/p94On4wpCXmsAfDg3vwtHCEiItl8z1DOWTngxMLZ7AzAjtcD6GQGIDUAmK3mKwAIPJjM4cIRIiKytaM5G9xtgXPLmW3ojssBdz0DMLuWry6AO6cKuRsoIiIi13bvnnyt8NrKAFibOpLuUHMRvm3regbgcs5qADx6qMxwMV9pIhERubbbJ/PXvTtbjWnEqQFAmR1WA+zBIMB8BQBHxvNXLEJERK5t/0j+uncXqglrYeZcuoM7ee5OBQAFUgKAxNpcZQA8A5NlDy0CKCKy+0zmcD2AuWrMapjZlb6jqYCdejfGceMArlKP7FaRy8AZK3ncs6eoNQBERHahUs5mAQAs1RNq2e3ovp08d6cCgNSKRGuhZbWRn0GA42WPh/bna5CIiIhcn5JvGMtZldeF2pZdADuqBdCpdyK1JnE1SqhG+ckAVALDvuH89RGJiMi1DRU89gznKwBYqifUstvRgcgAzKRtXGvkqwvgtskCBQ0AFBHZlYYKhr05u8hbbViWszPp+6y1227HuxsARJa17MELA+e9x4b6fQoiItIl5cDkrtCbBc5nlwMewa0MuC2dCgCm0zauhQnVHGUAHlT/v4jIrlX0DWPFfHUBwJbrAYwB275y7WoAUA3z0wUwVDBMlvP3iyEiIten4BuGcxgAnMteD2CMQc0AVENLNSddAPfuKTI9lK/UkIiIXD/fQCmHH/MXsxfVGwG2vXztjgMAa20BVwdgk4VaTE7afx7YV2SPAgARkV3LM+RyoPfppcwMwAQuCNiWTmQAhoDUzvO5ak5af2B6yKeo9l9EZNcyxhDksNTrxdU4a0GgEfo8BiAzBZGXMsDlwPDmI2VVABQR2eWGi/kr975YT2ikN6dl3DiAbelUBqCctiMvGYBKYHjrkW13o4iISE4MBYactf9EsWWhlnlBnToN/3p0NQCYz8lKgJUcFocQEZEbN1nJXwYgSuxWF9R7t/u8nQgAKmSMAZiv5SMD8LoDJZT9FxHZ/UaLHl7OPvCjxK0KmGEwA4DFnHQB3Lun2O9TEBGRHij6JncXfFFiWci+oO57F0D6LIDsPouBMq4CQCIiN4XAz+EYgATmB7QLoExKAGCBxZx0ASzV8lGtUEREdiZvjT+0MgCZF9SphfiuRycCgFFSAoA4sSzV8xEA/MHLqygEEBHZ/VbDhCRnH/hRApfXkqxaAPu3+7ydCABS5yAu1fPzJp9ciHh5rtHv0xARkS5brCUk6Q3pwHIZ9Zg4/bTL1tptrQfQiQBgMm3jYi3JzVX1Ui3hJz4+z0r2mssiIrILnF/JbEgH2mI9IUw/cY9tdgN0LQBYyFGUZYHf/NoKH3hykbW8LF4gIiI3xFrLYi3OTXa63WItyVpbx8etCXDDOhEApL7wQi0mJ+0/ANXI8nc/OsePf2SWs9lLL4qISE7VY8tcTganb7TcsETpkYtHHwOA3HcBtKyFlp/53CJP/MoZfvrT83z9YoNGHnNFIiJyhbWQWMvXL4V8/JVqv09nW5bq8VZdANsKAIIdnZGTuhTwUj3JVQagJbHw4lzIj/7hLL/y5BKPHizzPa8b5bGDFYZL+SsgISJyM2g18omF2FqqoWW+mvDyfMhT5+q8OBfy7KUGpxbzmeFdaVii9OSFxzYXBOpEADCVtnG+mr8MQDtr4bnLIc9dDvngV5cJipa3HKzwTbcNcXQy4M6pIodGC0wPG0q+wc9bcWkRkQ5qXfBZLFhIaDXKru89sRBZN0U8Stzc9kbs7tcjS5hYVhoJq81Udz122+qRpRFBI7bE1hLGsBYlrDXcVPOlesLZ5YjTSxGL9YS5akI9ynPrk26xFhOmdwH4ZGTir6UTAUBq6mExpxmALFHD8ImTNT5xskbBg30jAXtGDFPlgIMjAffuLXJ8usCtEwEzQz7DBY+Cb/A9CDzwjMEz4Jn1QhRafljk5nL1Z+LmD8i0j8z1hvXq+812FovF2vX97V/bj2HD/lbDHDfvx1caZ0vYapRj10jXY0stars1Euox1OOEWuyuthdrCfXm8WEMYeIeX4/XnydKuHKFHifumDiBsNnYV0N3fJS4c4mtCxSS5jle/U7dXJ+fy1tnAPrWBZCeAcjhGIDrFSZweini9BJAuGl/pWAYKRj2DAfsG/HZP+IzVfGZrHiMlzyGCoZS4DIHU0M+k2Wf8XJzu28oNm8F3+Ab8M3VwYLiBrnZbHUxYdvutTcScduHZWLtlePihLb7649pNUruvtvXut8afFWP3POEbY1ZI7ZYa6nH689Ri9y2WmybDa25UsSlNTsqSZpXyay/VpS4nzW2ltZFbD2yJImlkTSvgptXxbFdv3KOEqiGbpR4NUwIY65cQUcxRLZ5VR2tN/itdHmSuPvrgYAlSdh6qpxp/af068Po5vsQXK4nNNIzG/0JAKy1BdxaAFdJrGU5J1UAu6Eaukj20lqdZy9B65fV0MwANDMBvjGUC4ay7wKCwDMEHs2v7r5vwPNcIFAODEMFj+GCoVIwDBU9xkoeB0Z9ip57jqLnAoeibwh8KDSfp+C5DETguaxE6/Xb/47aR4Qas/F7c9USml7bY03z+Ov9k+x2b8mVq58bfEzWb6xtXiml7gOSjDlFlo1XLVfb6i/ENQ7pD259aKdpb8Q27yNrFDGJdY1a+17X2Fx9TKsBbD+mPS1pLczXYi6trZctjZsNV7tGtKGxTnn9MLZXGiHX4K43kldes+3/2r3f68+QWK7sTNqOu/p1bVvDvL7PNdrN+7j/f8t6UNFqPG3zyrmV4m69TpSsH2+bl+L2yl+I3RTQXPkZbPa27V1QtR7V4T8624XnlC0llqzqugYYsdYGxpgbGuCw0wzABCm/BY3Yaj49sPGtsTSjaru+pTpAfVXtZ+sZKPjr3/vGBQ4tLqBYf0TBv86Pg2YQ0s0gIE6yG8Hsx2Rf8cQ2c/qN69dMIDWda122KOvqdXD+56U3+vE/rkZ6N9liSeBRoAj0NABIHXhQj2C1oY+3vGn/H4stxFHW3rTvRUSkm2azA4ARXACwdiPPt9M6AKn9DsoAiIiIdNYWK+wOA4Ubfb6dBgCpGYBGbFkNdYUoIiLSKQvZAcAQ28jo7zQAT6DbyQAAIABJREFUSC0+ECZuqoiIiIh0xnJ21/oQfcgApFYBVBeAiIhIZ81njwHoSxfAaNrGMIaaugBEREQ6JmMaIPSpCyA1AxDGgzW9TUREJO8Wtw4ABiQDkCgAEBER6aSlWmaJ/b4EAKmDABuxpaYxACIiIh2zFiZZRckKQOVGn2/bAYC11scVH9ikGrrFJERERKQzwgRWG5mN6w2vB7CTDEABKKXtWFbrLyIi0lFRYlnL7l5PzchvZScBQAkop+3YolqRiIiIbEMUuwx7hp4GANkZgLoGAIqIiHRSZLcssjd8o8+30wxA6qCDhexiBSIiIrINUWypZnexp87K28pOAoAiGV0Ay9mDFERERGQbYgu1rHXLe5wBKJLVBaAAQEREpKOirdfZSZ2Vt5WuZAC2KFcoIiIi2xAlW5bZH4wAYFmzAERERDoqzkMGYIt6xSIiIrINYQLVaHAGAWaMAdA0QBERkU6KE0s9ytw9dKPP1/FpgNbCYk3TAEVERDopSqAaJmRcYve0ENBw2uPrUUJD7b+IiEhHWXAr7aYvCViw1gY38nw7CQDG0zauNCxJxnqFIiIisn3V0GYtCeyTMS4vy04CgNQBB6thQqwxgCIiIh1XixIymliPjHF5WXYSAKROOViLlAEQERHphnqcmQHwcIPzr9tOxwBsUm1YErX/IiIiHVcNbdYgQJ8eZgBSA4BapABARESkG66RAehzF0CYkL1WgYiIiGyXywCkNrIePRwEmFp0oBZZrMYAiIiIdFw9GuAxAGuRVQZARESkC6rZAUBPxwCkdgFUw0RjAERERLqgHtuBmAaYPgsg1DRAERGRbqgNchdAVbMAREREuqI2INMA0wcBhgoAREREuqGRPQ3QAIUbea6OjgGw1k0DVAAgIiLSefV4y2mA3e8CsNaWcOmGqyRYapoCICIi0hVhbMnoA/DoUQYgdSngOHGDAEVERKTzXAYgVc8CgAquv+EqibXUIgUAIiIi3RBmBwA9GwMwREoA4DIAWgtYRESkG7YYBNizDMBQ2mNjizIAIiIiXdKIs4YA9DYASMkAWKoKAERERLpiEKYBVtIemygDICIi0jWW5kyAzQxQtNZuujjP0tFBgLF1KxWJiIhId9Szp9sXSGmbs+xkGmBqF8CaBgGKiIh0zRaZ9iI9CABSuwCUARAREemuKLvcbkAPAoBy2osk1lKPt/mMIiIick1RdqLdpwcBQCntRRILjVhdACIiIt2yRQagvwHAFoMTREREZIe2uM7etEbPVjrbBZBAI9rmM4qIiMg1hX0eA5CeAcDS0FrAIiIiXRNnt7MefQsAEs0CEBER6aYtmtnudgE0qwylzjUME4sSACIiIt0TZY+163oXgI8LADbROgAiIiLddY0MQFcDAI+MAKAWKgAQERHppjg7A9D1AMAnY8UhLQQkIiLSXX0bA8AWGYDqFuWJREREZOei7Iq7PckApAYAmgEgIiLSXZHtXxfAFhkABQAiIiLd1M/FgDQGQEREpE+2KAV8Q216Z6cBahaAiIhIV/UzA5A9DVAZABERka7q53LAWxQC0iwAERGRbrrGcsDXLdjGa3tkjAHQLAAREdmkNWrdNm9bMW1fzXVfzN5UOpUB2E4AkJkBUBfANVgLsYWkeb/9j6H1y+4BvnH39bsvInmWWAgtWEtQ8pgoeUwUPcaKhlLBo1Iy+J6h3kiohZa10DJfT1hsJKzUErfCXOA1mzV9ILZ0ahpgRwMADQJsY7nSyA8HhsnAMFbxmRkvcHQyYLLsUwoMRd9gLdSihGpkObccc3I+ZHU5Yq5hWYgSF+15AAoKRGSANT/3PGB/yWNmvMAbDpe5f2+R43uLHBkLmBnyKfrGXe80P9ISu35bbiScW454eS7k+UsNPnu6zqmLDc5VY9Zi2/wMvLk/C7dYDlhdAH3TuqoHAt/wxoNlHj1Q4uHDZe6ZKbJ/xGe05DFSNBR8g9cW0VogSSzV0LLcSFiqJ7wyH/LVCw2eOl3j02dqnFqIXPaglS24if8ARGSAtC54DDy4r8R7jg3xp+8Z5uhEwL7hgGJw/R9We4Z9bp8s8NZbKlhrma8lnF2O+fSrVX7/+VU++kqV5XriXtO7OT8E+9kF4GU9rp69QMHulriGf6Lk8fDhEn/q+DDfeMcQx6YKVAJz5XfUbJHCMoDvGUZKhpGSx4FROD5d4N3HhogTmK/FPHm2zu89v8ofn6jy0uUGYcx6CC0i0mvNz769wz7vPDbE9zw8ypsOlxkree6jaYdpe2MMUxWfybLHvXsKfM/DY5xaDPnNp1f4nWdWeOZ8ndDiuk1vIlssB9z1AMCQkWYIs9MSu5O1EEPZh286Psxfef0Ybz1SZu+If9XV/XYZYzCA58Pe4YBvvjPgG28f4sXZkN96Zpn/9PQKz52rkwQaMyAiPdTq3jTwxPFhfvitEzx6sMRE+YYy0Net9VlYCuD4dJEfe8ck3/3ACP/+q8v82pNLnJwLsa3PwZvAQGYAwuwFCnafxOID33nvMH/tjeO8987hnrxs0Tfct7fIfXun+aE3TfALX1jk159c4tnLoQbKiEhPFCzcv6/EP3nPNE/cVqEcbGdG+fYF3v/f3p3H2VHVeR//nKq79b5nIwuQQCCELeyLCDgI6CiCj4A6bjgjOo7OM6KiPqOOM+rgOCiOjuOGPm6jo/jgOCICsi9hNwl7gCQsIXt30tvtvtt5/qgbDElVr9X33qr7fb9esX317e7Xud1FnV/9zu/8juGgrhSfPaOLdx7VylX39PHjPw7QN1qqi2WBatcABAQAdZABsBYDHNSV5K+ObeM9K1rpbKjsxb9bZ4PLZad08PqDm7jijj6uWztEf64cGioQEJGwWUtjwvC+E9r5wPFtHNztWw9eUQd0JPniWd2cuLCBL9zay+M78lhrY30PHKPjjkO1lgBycV8CsJaUMZxzcCP//NouDu1OTXuNa7pSruGIOWm+d/4srnlskM/d2su63ryWA0QkXCXLfq0JPn1GJ+85upVUDa27N6UcLlrezIq5aT7y++3csHaIYoyDgOBdgJO780+1FbBv4DBGd6Los5bmhMPfndLO9940uyYm/z01Jh3edkQLP3vLHE5dmMEt2TGvEhGRiTIlyyFdSb77pllcUmOT/26OMSztTvGtN8zifce1eXvVY3oPDOtdTTUA8M8AxLUGwEKDa/jMmZ189oxOeprcmpr8d0s4huPnZ/je+bM556BGjDozi8h0lSxLOpJ8/81zOHtJI8kanPz3tKAtwZde281fHN1KIqap0JrLAFgb0xoACx0NDl8+p5v/fVI7DcnqrPdPxsFdSX7w5tmcu7RpSms8IiIAWDi4O8nVb57NifPToexuqoSWtHfPfv8Jrd6MGLOpqRT8hmY8APDdZlCyltjN/9bbXnr5qR1csqK15iPf3Ywx9DQl+Mq53Zy4MOPt1RURmQwLmYThS6/t5qQFmZrMeo6ls8HlM6d38er9G6o9lNCNkwGY0dMAfcs+CyW8yssYSRr40AltXP6qjkg8+e9taXeKH10wmyWdydiuhYnIzMi48A9ndPCmQ5tJRHRrXU+Ty7fe0MORs1PxugfawL/HjGcAfNsAF0s2Xg+a1nLSogwfOaW92iOZlkXtCT53Zhed6egFMCJSJdbymsWNXLKirdojmbaDu5JcfloHScfEJgiwIa1phJoBiFMA0JZ2uOrcbha0+cY7keEYw0WHN/Pe49pic/GLyMxqSTn8y9nd9DTNTGe/SjLGcN4hTbzzyJbI1DCMZ5w+ABMWWgagYG1s5hfXwKXHtXHE7HS1hxIK1zF88Pg2Du1Jxa4YRkTC5Rh47zGtHNwV7YefPTUkHD5wQhvzWuJRFh3WXBviEsCYlYmRsqQzyXtXtOJGdN3Lz/zWBJce20Y6Hte/iMyQ/dsSvGdFa2TX/f0YA8fMy3D+sqZYZELHeAczXgTonwEoxSQDYC1/cWQLS2IU/YKXBXjrES0s6ax+604RqVHWctaSRg7ried94p1HttDR4EY/E1rFPgD+GQAbjxqAzgaXC5c3x2ataE+zmlwuPrw5FhGwiIQv4RguOTpe2c89HdyV4vT9G4h6BDBGEeCMBwABRYDx2AXw5mXNLGyLb578gmXNzI3JOpiIhMjCsftlYpf93FNL2vDaCHQzHM8Yc211MgBx6AOQSRhOmJ8hHfGLYyxzm11WzE0rCyAir2TgpAVpWlLx3TJsjOG0/Rsi2ddlT9U8C8A3AxCHPgBdDQ6nLIpex6vJaM+4nLQwE9sUn4hMTWPScMy8TOSfjsdzaE+K/Vqjvb2xmp0AAzMAUT97pqfJZXFHfNNf4FXDrpibJp2I93/kIjI5zSmHA2N+/wNvdlw+Kx5bvH1UrxNg1LPKy2elYx/9AizrSdXkcZ4iUj1NScOBHfVRH3T47GjvcigFT7YVOQxoH8UYHD+/tDv+0S94PQGaIr4GJiLhSrmGjky0U+MTNac52u9znD4AEzaVWcA3RCxZG/lGQF0N0b4oJsp1DG0ZBQAi8ieZpKmLDChAa8QLHavZCdA3ACiWop8BqKdJsb2O3quIjC+TMMS4/vkVor4LoJqNgHwfk0s26q0VwI34NTEZiTp6ryIyvuF89Ou4JmooH+2S9Wq2AvbPAMTgMKBdI9G+KCZjZx29VxEZ30jBMlKI+E18gnqzxWoPYVrGuHtXMQMQ8Wunf7Q+JkVrbd28VxGZmFzBsm24UO1hVMTG/mgHAOP0AZiwEDMA0V8CeHpHvtpDqIhtw8W6ynaIyPgGcyXW9dZHALBm82i1hzAtNRcAlGLQB+CRrbmx9lfGxlPb8+SK8X+fIjJxgznL2h25yLd0H0++aHlsW67aw5iWarYCDuwDEPVtgJv6C2weiHZqaCJWbx5VACAirzBasDywcYTRmN8bHt40yvahaN/nxzkNsPJFgKUYFAHuyBa558WRag9jRg3lSqx8YYSIF8GKyAy447mRWC8PlqzlD+uGyUa82LHmlgDiUAMwmLM8uHGEQtRPNRrDlqEiD74U7fUvEZkBBp7pzbMq4uvjY+nNlrhtfZaYJzkmLMQagOjvArAW/uvRQTYPRjs9NJZbnh1m7fZor3+JyMwoleDr9+5ipBDPLMADG0e5fUO22sOoGeoDsCcDG/ry3L4hG8tCmNGC5UerB5hklkhE6oWBhzaOcu8Lo7G7BxZKlu8/vIu8Hv9fFl4RYGnMwoToMIZ/v28XWyNeJLI3ay2/eWqQ1ZtGNf+LSKDNQwX+9e6+WNUJWWv57VND3PTMMHXT73gC1AjIx6rNo/zq8aFYbQnszZb4j/v76c/H5z2JyEww3PTsMLeuz8bmHrgjW+I/HuhnVy4e7ycsUwkAfM/MLVo7VnvCSMnmLVfc2cfG/vg0xfjlY4Pcul5rXyIyDgO5Enzqpu2s74v+PbBkLd99qJ8bnxmu9lBqjjIAfgy8sCvPZ27pZddItJcCrLXc+VyWL9/RW+2hiEiErN40yudv28FwhNcCrLX8/ulh/n3lzmoPpSaFeBxwDIoA92QM1zw6yA9XDVR7JNPSN1Li0zfvYN2ugtb+RWTCihh+snqQbz0Q3V0BT/fm+cj129g4qPufHx0HPIbBfIm/v7mXa58YjGRvgN7hIh+7YTt3aNuLiEyWgYK1/OOtvfzgjwOR6x762NYcl/5mG0/VyRkvUxFiJ8AYLQHsZgwDoyUuv3EH97wQvQ6BV969k5+tHsAao8pXEZk8Y9g1WuJzN+/g548MRGYL3cb+Ah/87Vbu3JD17n26//kKcQkgJtsA92bg6e05LvzZJq5/eigSmYD+kRKXXb+NK27rJRvNzJ2I1Apj2DJc5P3XbuWqlTvZWeN1UXc9l+W8n77E7euy1PZIqy/Ew4AsEZgbp8YxbBkq8qH/2cavHh8kV7Q12STDWsuWwQKfvGk737pvFyXfv5SIyCQZQ7Zk+fwtvXziph1sGyrW3D1wtGD57doh3vfrLTz80ujUZrc64/s0P5XviWMNwCs4hmf78vzVr7eyZnOOj5zcTldj7cywxZLl4U2jfPyG7dyxPktJaS8RCZNj6C+U+Pb9u3h2R57LTmnnrMWNuE517zPWWjb2F/nWg7v45n076cuWoMpjiorQAoBYtAIej2MYyJX4yt19rN48yidO7eD4BRlSbvUuNmstvdkSP1zVz3fu28XavgJ2UgdCiohMUPmh4uZ1WVa/NMrfnNTO245sYVFbgmQV7oM7R4rc8/wIX7ytlwc2lY8514PPhE0lAPDfBRCDw4AmxBhGSnDdk0M8uCHLh0/r4KLlzSzuTFVlODc8O8zV9+/i/z0+RMmhfPHrPwARmSHGy/ZuKxcHXrOqn3cd38aFy5tZ0ObbJy502XyJ+zeOctVdfdyybpj+nAVXWc/JCiUDYG08jgOeFMewJVficzfu4CcP93PpCe2cfVAj+7cnSLsGM0MXYsladmZLPLI1x49X9fPL1YMM5kt7TP4iIpVRcuCR3jz/5/fb+cFD/fzZkkbefmQLizuTtGccnJDuSdZackXLxoEit2/I8rPVA6x6cYRtIyVvrb+KWdgom1QAYG1QctnGpmf0hBnvf3IOPLE9z8eu38a3H0jx54c0cdHyZpZ2pWhKhRcIFEuWbcNF7nwuy3+uGeTO9cP0Dpe8dL8ufhGpBmPAwKj19t0/vjXHNY8OsGJ+htcd1MRpixqY35qgOWVwprAunytYdmSLrN48yu+eHubu57I8viXHSL68zq9737RMNgPgEpBfjmijqOkzgAN54IntOZ68K8fXV+7kwI4krz+4iRMWZFg2K0Vng0Nr2iGTGD8oKJQs2bylN1tk82CRh18a4bZ1WW5dn6UvW6KwO9eip34RqQUGcAwWy8bBIhufHOJ3Tw6RcAzz2hIcMyfNUXPTzG9LsF9bgq4Gl5RrSLrgYMiXLIWSZShveam/wMaBAmu35FizNcfqLaMM5ywFu8dGc6X7QxFKAGDxlgDq1h4XogVGivD4thxPbM3RmjTMbXCZ1ZlgaU+aRe0J5jS7tKQdGpMODQlDwXoT/nDesmO4yMb+Iht25Xlm0yjbBotsHimSK7HHhK8LX0RqjHn5fwAo4mUu1/fmWb8jzzWPDdLgGloShqaEIWkg5RgcIG8hX7KMliz9BctgwVIs8fIDlpr5zIzQMgDF2DYBmAIDGIMFdpVg12CRJ/sL3PFs1ouUdk/mbjmFtTuCKu7xZO8aSJg/Xfi1s+NQRGRiyksEu2WBbMFCYZz5wjHax18B4WQArNcJUALsXqffc71q75oJtxzqKsgVEZEKmGwA4BBUA1BvRYDTpXSWiIhU0WSTLME1AMoAiIiIREYoAQCoBkBERCRKwgkAbJ3vAhAREYmYyQYAvjUA3hKAIgAREZGoCG0JYLxdHSIiIlI7phIA7ENFgCIiItEylQDA93u0BCAiIhIdU6kB2Id3GqACABERkagIcRvg9AcjIiIilRFSIyDvJCcRERGJhvAyAJr/RUREIiOcDICFgpYAREREIiOURkCgXQAiIiJREt5hQNoFICIiEhnaBSAiIlKHwjsMSAGAiIhIZIR3GJCWAERERCJDhwGJiIjUofCKALUEICIiEhmhnAUA2gYoIiISJZMNABKoEZCIiEjkhdYIqKQiQBERkcgIrQZAGQAREZHoCLERkDIAIiIiURFOHwBrdRqgiIhIhITXB0AZABERkcjQWQAiIiJ1SKcBioiI1CEdBiQiIlKHdBiQiIhIHQqxCHD6gxEREZHKmEoAsA8dBiQiIhIt4e0C0BKAiIhIZITUCEgZABERkSgJrwZAGQAREZHICK8PgDIAIiIikaFWwCIiInUotBoAzf8iIiLREdoSgAIAERGR6AilD0BRs7+IiEikTDgAsNYaAgIArf+LiIhEy2QyAGMEAOEMRkRERCojpCWAEEYiIiIiFRNSBkBLACIiIlESSgCgcwBERESiJZQAIK8lABERkUjRNkAREZE6FM4SgDIAIiIikRLSEoAyACIiIlGibYAiIiJ1KJxtgNoFICIiEimqARAREalDoSwBqBGQiIhItOgsABERkTqkVsAiIiJ1SLsARERE6pAyACIiInVIAYCIiEgdCuk0wHAGIyIiIpURzjbAYggjERERkYpRJ0AREZE6FFInQAUAIiIiURJSJ8AQRiIiIiIVo06AIiIidSikXQBaAhAREYkSLQGIiIjUITUCEhERqUPaBSAiIlKHwskAqBGQiIhIpIRTA6AiQBERkUjRNkAREZE6pBoAERGROjTZJQDfr9dpgCIiItEy2QxAwu+FvCIAERGRSAmpE2A4gxEREZHKCKkToCIAERGRKPFN6QdQHwAREZEqMLs/GsjmAx+6M0ziwX4yAYADJP1emN3scvKCzCR+VHX0j5Z4cntO2xb3knDgmHkZXDP+14pI9K3vy7NpUE9u1bF78ja4BtIJQ0PCkEkYMknH++gaMklD2vU+35p2aG9w6Mi4dGQcjpqbDvrha4HsREcymQBgCFgDnLL3C586rYPh4IikZqzaPMq7r93C9mFFAHtqSTn84i1zcCe7ICQikfSlu/r4+n27qj2MyDGA64BjwDUGxwG3/P9dB1zH4BjIuIa2jENLyqGj0aWrwaEt7dCSdmhOeZN8OmFoShqaUw6t5c83pxxSCUPSgaRrSDiQcEz5Hy9/NAQ+ra3Bm6snZDIBwCDwMD4BQEeDS0fDJH5Slazvy1d7CDXJGOho8C5OEYm/hP5TH1PKhcNmpWhKepNyU8rQknZoS7t0N7q0pvecuL3X2jMOTUmHTMIh6Xq/Y9cxJB1DygVjZjzFOgrcY4wJPwNgjMlba38HXATMmsLgREREat7CtiTXXjyXxqTz8kS++wncNeX1+N0fMcz83D4hTwLXT+YbJpMBALgR+DFw2SS/T0REJBI6GhzmNCdIJ2pjZp+AXuDTxpj1k/mmSSWCjDEl4MvA1XhLAiIiIrHSnnFr5al+PFngfuB84LrJfvNkMwAYY7ZYa/8OWAm8FTgJaJzszxEREalFLakxyuxqwxrgPuBm4D5jzIap/JBJBwAAxpgBvCzA1dbaDLAc6JzKz5oh5wF/Xe1BxMQVwK3VHoSITFoL8GngyGoPJGraM07YGYA+4EVgBK9YLw8UfD5uAVYBu8r/snjZ9mG86v4hYNgYUwhjUFMKAPZkjBkBHgxhLKGx1p5d7THEyFPGmBurPQgRmRxr7eJqjyGqWtJOUAagAPQDu/eSW+AFvIz45vJrWSBX/lcof9wObMWb/HNAsfyvsOfHsCb2iZp2ACAiIhInbenAGoCngLPwiu5KeJN2ZBvLKAAQERHZQ1smsAagAOw0xoxWdEAzRO0gREREygzQnAqcGkfx1vFjQQGAiIhIWdKFxqQT1Lmv3xhT+33vJ0gBgIiISFnSMTQkA7cAxOoABQUAIiIiZd4hPYFTY18lxzLTFACIiIiUJd0xMwD9lRzLTFMAICIiUpZyDI3BAcDOSo5lpikAEBERKUu5KAMgIiJSb1KuoVE1ACIiIvVlnBoA7QIQERGJo5RraEwoABAREakr4ywBqAhQREQkjhqShoTrmwHI4530FxtxDQB8WzUaA4FHPIiIxEuJPx1b+woJR/fBIB0NgUcBD+AdBhQbcQ0Ain6fdAxo/heROlEkIABI+j/hCtAafBDQMAoAIiHv98mEY2L7hqcr4OxrCLiBiEjNKxLwMJRyKzySCGnPBP5yBlEAEAm+fyTHjDnR1S0DjPFA4HsDEZGaN0YAENdb//R1NAb+brQEEBG+fyTXUQVAEDd4TVABgEg0FQi4F2aCt7nVva6GwAxAPwHZ5aiqrwBAGQBfxigDIBJDWQKq1jsb4nrrn77uxsAAYBcKACLBNwBIKAMQyAmOjBQAiETTCAoAJm2M342WACLCt1tTY9KMlequW64xQZmRwG1EIlLbjDFFoNfvteaUo2WAALOaAjMAfUCugkOZcXENAHy7NbWkHRJxfcfT4AT/TiwBPRVEJBJ8A4BMwtAU3O++brnGmyd8WGCgHFTFRlynQ98Tm1pSjva/+sgE/04Ci4hEJBI2+H2yNe3QEVzsVrfaMw5J/yxxgZidBAjxDQCG8UnVJBxFvX4CIl7wfoejFRyKiIRrg98nW9MOXaoD2EdPkxv0kJgHtlV4ODMurlfAKDC09yeNgU5FvftoywReBiMoABCJsqf9PtmecekOXuuuW/NaEkFNknLAS5UdzcyLawCQJaAQcHZzosJDqX2t6cAbwSgKAESibCM+OwGaUoY5uhfuY15rgpR/BkABQIQMAjv8XljUpot+b63pwGURBQAi0ZYDntz7k44xHD47VYXh1LZ5La4CgBgYALb6vbC4K1nhodS+luDDL1QDIBJtReAhvxeO3y9d4aHUNtfAnOZEUFO0ncaY/goPacbFNQDoB7b4vXBwZ1LdAPfSE7wWOIRXUCkiEVTetva432uHz06P1QG07mQShjlNLsZ/gthQ4eFURCwDgPJFvxqfLnazmlzmt2gZYE8HdARmRXYS0FNBRCLjMXyWRJtTDsfMy1RhOLWpJe2wtCdwWeTBSo6lUmIZAJQ9jM8e9o4Gl4O6tQywm2tgbvOYh18MVnA4IhK+dQRsYXvNgQ0VHkrtak45HNDu+3BogUcrPJyKiHMA8CA+WwE7GhwOn5XSmQBl3Y1jbgd6whijVsAi0baBgBT2qQsztAb3AakrS7uTQc2RtgHPVng4FRHbv7wxZhhYtffnE47htEUNpNUHG4A5Le5YDUEeq+RYRCR8xpgCcKffa8t6UkFPvXXn1IWByyHPEFBUHnWxDQDK7vL75JkHNtDdGPe3PjEHdiSZG1wTcW8lxyIiM+Z2v08uaEtwwnzVAWQShtMWNfq9ZIGniGEXQIh/AHAfPssA7RmXCw5trsJwas/x+6Vp8D8h6WlgU4WHIyIzYy3wyN6fdB3DJSta6/6QtKPnpDkKWySGAAAKJklEQVSgw/dBqAg8YIzJV3hIFRH3P/uTBLTCPH9ZM13KAvC6g5qCtkU+gA4CEomLXcDdfi8s60lxzNz67QngGHjVokzQdugCcE+Fh1QxsZ4BjTHrCEh9Hb9fmvOWNVV4RLXliNmpoG5gBeCe8tqhiEScMSYH/Bqf44GbUoZ3H91Kpk7rolpTDm89ooWE/ymA64A1FR5SxcQ6ACj7ObBPJXtj0uGyEzuCDn6IPdfAX65oDWp6sZ2YbnsRqWN34bMbwDGGCw5t4pDu+mwNfMaBGQ6f5fveLXCdMcZWeEgVUw8BwEPAjX4vLOtJ8anTOknXYTusExakOf+wwAzIswS0DxWRaDLGDAE/9HttVnOCy05uq7tagLktLh8/tRPX/+l/K/DfFR5SRcX+z10u3rga6PN7/X3HtPLqxfVVBZty4UPHt7Nfc2BDpN8aY9QASCR+bsCrat/H2UuaOHP/xrrqkfKuI1s5ak5g5uMhYp4JjX0AUHYTcIvfC3NbEvzL6T3sXyd7YV0D7zqqlTcdGlj89xLwn5UdlYhUyLPAT/Ap8O1pcvnsGR10+jfDiZ2j56b5+KntZILTHv9mjPE9Vj4u6iIAKP8Rvwhs9nv98HlJrjq3m4V1cFTwaw5s5DOnBy57FID/i7b/icRSubD3B8BGv9dPmJ/hQye2xn5Z9LCeFP92bndQ578ScD1wR2VHVXl1EQCUPQJcBWT3fsExhjcubeILr+mkPRPfX8mx89J85Zwu5rcmgor/VgPfjOueVxEBY8xG4Mv43Atdx/A3x7fz5mVNsV0K6G5yuPKcbk5dFHgOwjbgSmPMPr+fuInvbLeX8qT2NbytMH6vc+HyFr7x+m7mBB+OE0kJB85e0sjP3jKHZcGnXe0EvlC+OYhIvP0C7yl3H50NDl89p4eTFmRiFgRYFnck+N4bZvNnwYcgFYArgdsqNqwqqpsAAMAYMwJ8iYAimJRruGh5C/90ZudY/fEjxQAXL2/mm6/vYUlnMujJv4BXKPn7ig5ORKrCGLMN+DQw4PMas5pdvnpON0fHqEHQovYkXzu3hz9f2hhU9Q/emQk/KB8pH3vxCvAmyFp7EnANMM//dbjr+SzvvnYL6/ryRPXXNK/F5cMntnP5qR3jfekfgAuMMfvcDEQkvqy1fw1cAbT4vf78rjxv/eVm7n1xlFJEd8MnHHj1ogzfeP0sDgnOgIK3BHqhMWZthYZWdfF4zJ0kY8xK4PPADv/X4VWLGvjOG2dxzLx0efqPytVvcQ0cNSfF1efN4m9PbBvvG+4HPqrJX6Qu/Rj4L3yapQEsaE3w7TfO4vxDm0hGcLZoSBguPbaNb79xzliTv8UrfP5EPU3+ENVH2xBYa1PAO/DqAgI74ry4q8DfXr+N69YOM1qs/SCgLe3wjqNauPyUDvZrdYNS/rutwXvyj+VZ1yIyPmvtbLyM6CkEzAn9oyW+tnInX7t3JzuyvrFCzTm0O8nHTu3gHUcGtvndbSvwYeBX9db+vG4DAHg5CPgw8FFgdsDXsH2oxE/XDHDlyp1s7C/UZC4g5cLx+2X46EkdnLm4gZb0mOF6Hi/t/0ljzOrKjFBEapW19ljgu8BRQV+TzZe4dX2Wr67cyd3Pj5AtlKi9KcTS1eDyxkOa+NjJ7SzpTpEce/LfAPwj8ON6m/yh9v56FWetTQMXAf8K9AR9XbFkWfnCCH9/yw7ufG6kptbD5rW6XHJ0Kx84ro15LeP2MhgGvg983hizZeZHJyJRYK09Du/esHysr+vNFvn+Q/1c/ccBntyRq5nVUcfA8p40nzmjg9cd1EjD+GsWTwPvBVbW4+QPCgAAsNa6wGvwgoDDx/g6Ng0WuebxQb5+3y7W9eUpVSkb5hivc9fblzfzrqNaWNqTJj32aV4lvHWufwJ+WN4RISICgLXW4E3+3wWOAQKfJvJFy0sDRX6+pp9rnxrikS05hvPViQQc453r8vYjmrl4eSsL212csZc+c3jdYT9pjHmkMqOsTQoA9mCtPRH4BrCCcX43978wwvcfHuBHa/rJFip54Vta0y4XHtbMW5Y3c/r+DaQm1rVrDXAZcFu9RrsiMj5r7TK8tPj5jFMoXrKWF/sLXP/0ML9bO8Qt67MM5nY/Fc389DKn2eXC5c28/9g2DupKjrfWD17zo58AnzXG1H3HUwUAe7HWzgP+GfhfQOMYX0exBI9syfGj1f3891NDvNRfnLFCwaak4eDuFOcsaeJthzWzdFaShMN4RX7gnf/9U+AKY8xLMzI4EYkVa20SrzbqUmDROF+LxcsKbNhZ4Lq1w9zz/AiPbB1l82CRwVwptCVTA7SmHQ6fneLcgxp551GtzG5ySY7/EFQEnsC7t/9CD0EeBQA+rLXdwF8Cf48XBIz5exotlHhiW55b1g9z7RNDrNkyylCuRNFO/ddrgIQLPY0ux81Lc8GyZk5emGFRW3IiFzt4F/x64LPANcaY3JQHIyJ1pxwEHA18EziSMZYEXvl9MJQrsWFnnvU7C9z5XJZHt+Z4tjfPpoECI06JUs5gLWMGBgZvS7ZjLI1Jh8WdKU5ZmOGsAxtYMTfD3BZ3rIY+Lw8Hr+j5f4BPAs8YY2qkaqH6FACMwVp7FPAPwHkT/Z5CyfLo1jyrNue4dd0Qz/bl2TxYZNtQkYFcCWvhlVUz3p+gMWloTTvMbnbZryXBEbPTnLwgw9Hz0sxvnfQhRS/gdfa7Usf6ish0WGvbgIvxHojmT/XnDOVKPNOb5/lNJTYX82waLPBMb57ebJGRgi3fG71JvzFpmN+aYElnigM6kizuSLC0OzVendPeisAfgY8Dd+qpf18KAMZR3iP7AeCdwEJgwgcFDIyWGBgt0jtS4rmdBbYOFRkYLTGUsy9HvkkX2jIOPY0us5td5rUkaE07tKQdkhNL8e+pH7gbL2hZbYwZncw3i4j4KRdKrwDeB7wJ6GSajeQKJUs2bymULIUS2PKDkQESjiGdMGQSZryCPj9FoA/4HnC1MeaZ6YwzzhQATJC19hC8NbG3AYEnSfh/L+WL3LvQc8U/BQCuA2nXkHAMSQec8VNafvLAjcB3gOt1mp+IzIRyIHAY8EG8QGBWdUe0jx3Az/G2M/5R6f6xKQCYBGttE3A6XmHMyUBXVQfkRbkr8S72PwD9uuBFZKZZazPAEuBMvEBgKV4wMOn1yhDs3uK8Aa/F++31cJRvGBQATIG1thGvb8DFwOuA9goPYSNeUcvvgRt1sYtItVhrm/GyAmcAJwLHEXDQ2gx4BLgB73jjJ1TzNDkKAKah3EVwPvAG4EK8i34OEPYZmkPA88CjwG+Am4HtSvWLSK2w1jp4975mvIZCp+PVDSzE67Law/QyBDlgC/ASXsbzDrwMaFYFflOjACAk1tou4AC8bTNn4UXD3dP8sY8C9+Jd7KuA59TBT0SioryVcD7ew9FSvMBgf7zeAocyfj3VDrx74FPAw3jtezcA27TcOX3/H0xDIluO1ABWAAAAAElFTkSuQmCC"
CC_IMG_B64 = "iVBORw0KGgoAAAANSUhEUgAAAMgAAABGCAYAAACJ4ts2AAAQAElEQVR4AeydB3xVtf7Af2mBFsooQxBZZYgIDhwo4gLnc2Fdz62ogOyNyBDK3hsUEJShPBQQcMtQXIAgCsgWRNmbsgRk/fNNm0vu6bm3LbTwPv9XJTc7JyfJb/9yGiFJ/1XR0TQddutwICtI1hrI/+wa7NHnH1gAJgQAqawLFuoQf12V6wrectsteapWq5oVstbgf+4McPY1DBQAFnQAJioDIH2zZc8unbp1kiEjh0qnHgnSuWeXrJC1Bv9zZ4Czb2BAw0J2DRMaSPoCINdXuqKS3H3fPbJv3z45eOCgHD58OCtkrcH/3Bng7AMDwEJFDRMaQK4HQKLz5MktR/4+IseOHpMTJ05khXSswfHjx0UpJdE5oyVXTIzkzZdX8hfILwULFZR8sbGSW69tzlw5BYx08uRJof2FWmOezRxy5MghOXPlkjx58khs/lgzV+bM3GP0O0TnzGneifYXaq72uSpCSU49n5jcMWZtCxQsIAUKumubS7LnyC68V0bMFxgAFvLofdMAEh2hf44zmVOnTplF0fkM+uc/DAclJndu87JsSqGLCkmRi4vIJcWLSdFiRaVwkcJmw9i4PHnz6o3MKREREXL69Gn/AS9AKXPhPWLz55cSJUuYDdy5Y6esX7dOfpr/k3w24zOZ+uFUmTvnG/lt6W+yeeNmOXLkiFxc9GIT2Gyl1Hl5J+aqlJLces2LXlLUrPWhQ4dk018bZcmvS+TrmV/L1A+mmDkvmLdA1q5dK9u3bZOo6GgpWaqkAfbsOXKcl7mylcwXAOZsFC9RXEhv27pN1q5eI/N/mC+fTP9EPpqctLbLliwz73HkcNLa8n7nurZKKQEWgAk9n+MAiI4z919UVJSG+gLCC0drTLtn927ZsP4P+XnhzzLzi5ky6b1J8uag4TL6zdEyfcp0+fbrb2W5OVib5O/Df0uhiwpKMQ1AYLhs2bKdt83yrgoLFxkZKWxErphcsmr5Chk1fKS0b91O6tV6VRrVbijtW7WTfj37ypB+g6XrG12kZaMWUu+lV6VZ/abSrWNX+eLTL2TPrj3moOaLzWcewaEwiQz8YUwCzwAwd+3aJZ/O+FS66jk0ebWx1H+5nrRu0kq6deoqQ/oPMXPu0Lq9NK7TSHiXdq3aynC9JwB4tN6/S4pdItmyZzOHJwOnGRiKubK3rG2OqByy9Nel+vnDpF3LtsL6Na7bWN5o00H69+wndm1bNW5p3qNp/SbmvT77+DPZs3uPQUK8N2MSAg85i0SmAIidFOwFCwsFWKyBYWDvAdJeb0KDV+pLg1caSNsWr0uvLj3lrSFvysTx78u4MWNlUN+B0qVDZ2nhHKyE9gky/p3x8ucff0pefaiSFjHqvAEK7wNwQN0uKnyRfD/3O3mtaWtpqg/92NFjZfGixXL06NGQ84H8b92yVb76/Cvzvo3qNJDB/QbJ7p27pViJYoZKMv5Z7J9vF8YCEZUoWVL27tkrg/sOkoZ6vft06y2zNEKC2jEn38668J9//pEli3+V98ZOkBYNmwsHcc7M2ZJfU8yiRYvqFhLyXU1lOn5YWwJrC9WYM3OOfl4rHVrK++PeN4DCfEINyXtAYWbqte3dtZc0rJ20tgBKCU0Bc2r2jPUI1T+18gwHECYTo/lFMP6+vftkzMgxZpHbNG8jkydNll/0YTp08FCaFti+/NzZcw2mBqjatXxdY+HPJbumJABfZmI1Fo/Ng2qUKlVK1qxaI681a22AHAxH/dmEA1oRMvWDqWYzR7/1tkRGRGoqeVGGYGfWv1ChQhIdFS3vjBojDV6uL1M0CwVbdTZzpc8KTSkT2iWYQ7v458UCJwDrw7OoP9vA2kI1ihcvLss0u9RKU7TOGhmu1M872zERtKdMmiINNBIe/dYozaJlF5DaqZOnzmrIDAMQXlYppTFicTlx/ISMGDpCGtVpKO+Oekd+X/v7WU3O2+nUqZOycP5C6dm5pwa6Zppv/tRgNbDP6VOn0wR03jHD5TkA0ZoXBxN9PP1jaVKvsSz4cUGKLnFxcdKgQQN56623ZOrUqfLjjz/K4sWLZebMmTJhwgTp37+/3HPPPZJXy1Ru5/379+v1eVfatGgjuzULBFJhHQluu7Sm6VesWDE5sP+AtNUs0mgNfImJiUHdkUXuuOMO6devn5nbrFmzZOnSpbJgwQKZNm2ajBgxQpo0aSJlypQJ6kcGSglFmTh+osC2wSGwRtSlN9AP7A43MHXyVGneoJnAZXjHKVu2rDRq1EhGjhxp5sc8lyxZIsybte3bt6/ceeedRsZy++5P3C/vjHxHQMx7NRUtVrKYOR+skdsutXSGAAgvGxUdpQXtSwQSyUGa8O54gYKkNoGzqted1qxeK700SYXVWf/7eikZV9LwyACKrj7nf7wTckYBrTV5c9AwDZQ95Pg/x4PGfeKJJ+Sbb76RDRs2yPDhw6VevXry6KOPSrVq1eTaa6+Vu+++W5577jlp0aKFfPXVV7JOC/EcwHLlygWNg7zVuG4j+VWzNZaFCWqQhgzvzWFbuWKlNNBsBpTa7QYFHDp0qJnDnDlzpGXLlmZud911l1x11VVy4403Snx8vLz66qsyePBgWb9+vXz//femjTsO6WEDh0qf7n20ciKXoEhhrShPa6A9a5snbx4ZouWfAb36p6CezzzzjHk+a8a869ata+bHPK+++mph3qxtq1atZPbs2ea9hg0bJiArdx5Qetiu3zSFguNw69KSPmcA4WXz5csnBQsU1GzQKElo21E2/rkx7LNLlCghLECnTp1k0KBBBpOBbT/77DMZPXq0wW5t27aV++67z1CIcIMt+mmRcLimajbi4osv1qrWnCkWO1x/vzqwDIqFiy66SAuv/bRsNC6oGQAAYHz44YdSvXr1oLpwGcbjAC5fvlx69OgR1HTXzl1GoP950WJhI1nXoAZhMrRFA7jitxVGLkIL5Tbv3LmzrFixwmDiIkWKuFVh07fccovZm3nz5hkK6DaeNvkjgR3KrdnpGK0aZs3c+lBp2oFM88fml97demkFzX+CmkLdoMDvv/++8PygyjAZ3qthw4bmPbt27RrUEpmrab2msuSXJYLMx3oFNQiTOScAQUaAv4uIjJAuWmMzTgusoZ512WWXSYcOHQzbAcZlARISEqRp06YGS4Ft77//fnnllVcMduMAff755waTTZo0yWDnPFpv7zf+US0g99Pajf4aE+XRwArApmcRgsY8LUatzELyPjOmTg+qrlWrlmGhqlevHlTuZni3bt26yW+//eYWB9IAHwjg559/lpJakLYV6PE7vv6G/LnhT6PuTss70AZVOYcALc9RrU624yGL/PDDD9KxY0eJ0YfYlqc3vummmwwFhBK6fb/75ltDATgDyBIcfrfeLx0RESEgsjFaPvry0y+DmnDAoW4goKCKdGRyafsO5+ynn34S2E3blbVtr7WNW7TKHbMC62brwsUR4SrD1QEc8P5oC+BLZ30507c5LMOQIUMCkA0gREZG+rb1K0Rz8uSTTxr+fvXq1dKsWTNhM/zafqRtD2C17Dmyy9kCySk5JfDXs76cJSOGjQh6DJj43XffDSpzMyz6Qw89ZAD+jTfekGuuuUa++OILt0lQ+rrrrhM28vrrrw+UHzp4UFPhToIiA3kh3KGjjgPBXiS066TlGHxNk4aqUKGCAIA333xzUkGIX1gY5KN//etfsmfPnhCtkor7a1kKOSspl/QL5f7wPx8Khy6pJPQv8+XMoOoe+3bwOvbp00dgkUL3Tl/NDTfcIAsXLjR7YHsim3V4rYMcPXLUyCzMR1L5LyKVet9qDgKW4sOHDhu2ANLu17BNmzYGMBo3bizpAQq/sSi75JJLZODAgQKL8uKLL1KUInwz62vhsMDjpleIZMFi88UKLEr3Tt2CxgZ7gomDCj0ZBPRPP/00UMrBhUKyXoFCTwJsCrsWFxcXqFm/br1RcqDLV0oFylMmlDGqjhkxRpYvWx6oLlCggDAmckegMEQC1woEXmQkjJkhmgWKkbPYg0CBTqC+RxGDvMYa6qIU/ygHaW3etEWzVr2D6tu3by+tW7cOKvPL/PXXXwboAfw//vjDuML4tbNlnBfWAS2ZLWOeID7Or1Lh1japR7oBhBfNpck1bAL2ik0bNyWN5PyC1cCcvXr1SlWGcLqlOQm7NnbsWIEy+XX64dsfpKe2r8TGxgrzZM5+7bxlSimJLZBf698nGgOlrUcWAnvafKh406aUa7Fz505J7flQii+//NJgNTv25598plWfvxkDq19/yvLrua5ZtVqmTJpsuxm2ggME4AUKwyQ4RLbaZUlsmV8MFUeWsnXMZfyYccK+w0LZcjdmbRHKJ457T45pltjWPfLIIwI7avPeGPYZ6gKbBxKpUqWKENBuVaxYUZ5//nn55JNPvN0CeYCStY3W2khb+PFHMwSkHg6gbdt0AQgLAXsDzzmwzwD5eeHPdpxAfOmllxoVJyQ7UJhJCSgTgOgnm3ypLdYTtHERgVdSRxTmEHPgsBx/Mu3jwIzBxtOnTw/kwyXQYLE+bhtklrRQT4D+zTffdLsa4ykHjhBUoTMqQkm01hyicnUpVMGCBdOkpFi2bJm8/vrr8tJLL+nRkv49/PDD8tprr4VlC5NailEHc0ht/pvZ38jCBQt9AZpzE5s/1hj9YK9sH4AYRYfNe+Ovv/5aKlWqJHAiqHe99Rs3bpT33ntPatasKZy3zZs3e5uYPGO47BvzwQgK8oyMCM/upwtAeBr8+QfvTRJcQsi74corr5T58+dLBc3/uuWZmWZhFi1aJGiIvM8ZOXykzPthnlystVvuIfK2Ix8ZEaEPXLTBxrBGlBHAXhjFSKcWwHBr166Vp556SuD96TtIa+lS62frwYbILTb/43c/GM0LlNCW2Ti/1gKhlZn91WxbZGIOPmpkFB4ff3wG0E1l8k+7du0EVWnv3r2NPSG52GBi7Ar0vfzyy2X8+PG2yjdGHexWfKSNn5FaviS45VAP1hBK5+4Dz/IiFNsPFTP2DVgpWxYuhkVE7li5cqVvM5Q/UBxbOXfOXFmqtVq4L9kyvzjNAMKLQTnQs+Mm4R0MjQm8LBjMW5fZebCvy/u7z+vXo6/8/fffRosD5nDr3HSUJsG4LMzTh9KWs6AsrM2Hi8Fe8MhopTBqIY/AVyP4ssnh3CXccWFL3fyiBYskR3RKZ0FcScDYbls3DWWFIsCOjBo1yrjCUM/4PXv2JBk2oBBh3uEaPfDAA8bmY9tgSNyiZQwwsy0jzhWTS7Zu3mooDHkCiAA7Bmlv2Lt3r7F5eMvJYyKAIvixg9u2bTOUJJTXAAiLMWxg/WD7wp2LNAMImiGlsexwbTTzGxA+EF20ffj5jsEekFvvc7ds3iITx71vhFlvXSCvVbss1CptZDt48FCguH79+oF0uAQq2/LlywuYG3mCTYSKkkZAhBWBuqJcCDcOdWiUYFNJE5YtXWbkochsZ1gBWK7Dhw7Jqti5PAAAEABJREFU0sVLaGICVnqeZzLOD/II8sIVV1whzNMLHMwZoyFaNzC27QogJSQk2GzIuE6dOoE6ENHKFauMBjFQqBPIJqtWrjKaOZ01/8Kt7YABAwQgMQ2Tf0CCqP1///13o6RZs2aNQCG9gIIciPEwuVtQBEC77Zdq4yHAhGd2UEMnkyYAASAKFy4sc2bOliXOpthxxo0bJ1WrVrXZCxY/++yzxobincAH738gf6z7w2wc7+Kth5+H1C9euDioCuE8qMAng+YHSy8xLtIIlQcOHJDExESjZSFPt7Vr10paZRmAhD6E39esNXN3Dz82jQ3rN8ia1WtoYkLt2rWN3aWZVoMjN5lC5wfLOIeeudliAJED169fP+nSpYuxSGOVxisAvt+2CxejtgdgbZslv/wqJ0+dFLcsIjIyyI2EA0k/28eNobRjtQLGLYNFRhvFfljqxBqgUseI6WVBoeB//vmnO0QgzRg2g8/Xhj82GOWCLfPGaQIQ+Mdjx/6RyfqgeQe499575YUXXvAWX7A8m82CuhNg0blHAL8JP+zWkc6WPZvgF7VeAxF5Aq4iYH7S4QLyCpqScG1sHUBo0+FidxP/0eu+ZdNmicoRFegSo63XGBN5L1t46623CjIQKlj4cIAhtfmD8efOnWuHMDFUBL8yk0nDDxgZDZNtukGvIVjZAgjvfFibAwBo2wZqz1xt3o2RobZs2eIWCYgPe1pQYXIGlhbVc3I2EGGsDWSchLu2uA5t/muTcL6dJkHJVAEEjIt2Z973P2iMtTaoMxnIIfF/U/Dymszti08+F+QEyD15N0RGRBrjEUY6W165cmWbDBsDcJEaQ4ZtlFyZ1nZe+8WBAwfl1OlTyaOI8EyXFeQwuvOF1YUCIEfAdoay+nMQa9SoYag/Lj5ctZaz+A/2x3Y7qA2dJ0+cNHOkLAlADhlqSp4Au0fsDThLgnC95R988IFAKbzlNo9PnE3bGCHfpt3Y64R5QFP7iMjQYBC6JnnUiIgIgf9F6k8uCkQY6xBkAwX/JQm0QfD87nRYiIXzFkq+2HxGpevW8Y7Hjh0zVMSW+7Epti6zY+QJgn0O1M2mAQ5YuQP799siwzr6qbo5nGBf2BMC2jXeNdAxOYE1H1mCNYPVApEkV6UpwtvBNgTIuLIakS3paHF2uPSGZ4Btg0LHpt0YSoT2zy0j/dhjjwlzI+0X4BhcFpQ269atM1fHSbsBdsylGKxtRGSk2yQonfQWQUXBGTDuxg0bBW2KWxOpBw1n4HHbAs3o+HHVQDWIQL9161a3Scg0vkS4N9AXlemMGTMEQSxkh+QKMGhyMhDN/3GegN04OIFCneBdYAsOaUyts+ZfKJJuKjP5B02ge4hcr2gLIHv37gvMgrZ+ABJooBNQkf/85z/GtR05hWfo4qB/+MjhQArSwy/ql19+CaoPlXHXCmBA/sqeLbtpzlqzthxEU6B/3PY6G/jHQUfwfvnllwNlCQkJwr6Hez/kY6hmoJNO4CHAPHQy6B/tABJbuG9vomTTZ9nmvXHqAKJVdCtXrBBIp9sZhzI0NG6ZNw15R5132223CQvOyzZr1swYdlDVwTtyFdTbjzybiRwAbw1PTN/mzZtLfHy8MR5h4AoHZPCaLqZgzOVLlwuqQAxs5N2gRMlppdyi/5q037SUOjPXU6dOCWr4tEwY9gY5BY1aKDmFvQah4Sv24IMPGvtIuLEBgkC9npZS+scWaA2hUkqUUrYkBQUPVCQnxowZI7BNnBkANrk4ZARVVOrM+DQ8fvy4LwVBZCDQhqBUcD/K3JAqgHDIVi5f5fYxaayXJhHiBxsArM6SJUt8WwDdaBsAAlSRbiO8MXGH//XXX93iQJoNRNPBBmKYDFQ4CVgkL7kGiyHc5oiKdlqKIGhDovM63sLhgM/tzGLDVrhlodKuUB2qDeV7tR0A+wlpgsvC8DwOZIEC+akyYfv27QLGNJk0/oCtobK4wbOWoVzLuYLAXnMPI5Scwrrax6JdyqfZWA4oZbCDKBXy5ctH1oS0rC0Wdtf6bTqG+OH9CW41zyO4ZaRx/XHni3wNV0GdXwgLIGwE5BFVmNtZKSVYW90yN43fEtokt4w0h7ZUqVIkAwF+FxUxFmgKuVDUvXt3kkHBry+LAiUL5VaOGjBoEJ35Y/0fEhWVQ6fO/AP78qGAvLF5A4XuAQ0U+iSUUsbfjLVSyh8bUUfgBp3PECmKQADITLaCA2fTFkDyOgcOACXYNumJo6KiBFkSNhj2BnbNrz+escgpXK7ivgX7ZttBlW0aAOE97RVXkA9sekyeGNtEQnENNOC92Ve7Bjt27EiVOtKec0p/G+BumIfN2xjgcBEVsh5qaVvvjVMHEG0425+YGNQP/Tl8alBhcgb5wGuoYaIcfA4ypB0rs4sVGYuNh6rgNZs8lIn8+rqYAVYNdaVp7PlBnegpkgOJB4TD6pazQDyH23G2PBTls/U2ph/OcLT3UizbBrsC9RxEWxYuxiLPetg2zEupM1sFQOOpbOuJWVfitAT2yK8dCCWUfcK2x7qOVzOsGuwyKmWMdrYegyvra+cPBUF+yB2T2zYxhr5AxpPgbLCnGFpxd0E4h3J5mgVlJ0+eHJQnA9Il9gbsQW4ZyAcgdsvc9JlVd0uT0xGREcZFgYvwyUUmAjpNwucHtwZvMTwv1lw8R2FlcOp75513BPlkwoQJgu4bVeE4bXDE4Ob253B5+8Kj4kuEjMPBgPy7fWwabBihtXA2T4wq97ScFqXOYHvu0IORy11aliYmcKDdjTeFPj+MjzWaTYXK+TQxV1qp9xOM/doDcLY8KjpK+EjCP1rLZsuwK5QuU0ZTwmhbJFiZA5kwCSgBVAD1LpoetymYHY2WWwabjHbJLSMNJkbhwnu5Ktgy5coKwAvSoR0AklPLsWWdtWUO3oNKWwK2G1hM2DACcwrnEwalB/nS1w3cIXLzNo3Plk2ztnzXzKUots7GYQEE7c6xo8eEiya2AzEHj9gvQKrdcjQMCNRuGWmEbbQkrj8OVlzqbMBlww/rovbjAKPCtG39Yg4sJN+tO6D19LwTB9uWg+2OH/9Hrr+hii0ycVoPnWmsf0ItNJRRV6f5H9ePbeNLy18qZcqWFpeFgGLGlSkll1eqYJuJC1SBQk8CfzWutDIfDIQACoZBWCcUIch0UAjbjbUbP368sUGAvFCY2Do3RkVu81dfc7VBPtyRt2UnT5yQ62643mYNy8TNwUCBk4AKQ52cInODE/WzW0YaSovdhPchbwPAH4qCuHtasVJFiSsdJ6ilbV9vnCqAwBN6N75o8reRvIMd15oDlzelvmzZsmEtlbSxAfdlmyaGxLoHmbL0BDC2y8rR90DifrMg3nExvFWoWEFi88fSzATUyyZxHn9w+MTAZx955dVXCp8JddkAAJrDW/m6a2wzo/rGoBYo8ElwG9KVVaDWuJTDwvCuXtYLKm+HwVviu+++Ew52KHsKTomXV7w8BUIFoFnbfLFnBPVQAjhI2U9zRRkKHa44oPJnDgD04sWL7RRNzL6GGhvzgns+WVvmzLk1nX1+wgIIDzuoLY3efqEoCFjO3QD6eQ8oZX6BA2D9lmw9vKtNn03M/F2dN2MwRwAe9pG8DceOHpXCRYrILbfdYosEP6VQix1olMEJNEvukFWq3iBgaKXOsIQACPaGG2+6USKzZQs0535HIOOTgAr06NEj6B68TzNzPQCg+fe//52iGgqECh4g5rKT2+C6KtcLny5ivm45GJovrlS7pVqgGHkUVjlQ4CQef/xxc23ZKTJJtJrsByp/kAHslalwfihHpnWKAkk0q4GMTrC2IEalzqytLg76FxZAEAZzxcQEdSAD/0nsDQis0dHRQcVQoKCCEBkOMxoVt9oLbG5dWtMAhNuWOeIs57IA1POuR478LY89+bjm7c/4PSGQhnpf+mVkAGNzCOyYt9a4VSpfW1n2a6pny2zMnCpdWUnuf+g+WyQ46AEAgQJPAvkPj15kvilTphjHTlgs2wy+HSpDfWpu/ihqvAf0cb12J06eMGpzOyYxAH3kyJEUa4s637s/tCcgm6bGQtPOBpA2RmSAy5a5MXYdV6a84+475OprKsuBA2c8Etz2Nh0WQMDqaIzQStgOxAhPxN4AcGCpdMuxznL43DK/tFJKkFfcOq8Q6dalJQ1weu0D+TSZB0h4N3cMpZQ+iIlSQbNZDz8eH6iiP/JSoCCTErAKrgu4Ukqefu4Zc9j81o9Dd/jw3/Lks08ZFsxOi/vdXlnO1tmYPUWOQxUPpbDlCMO1atXyvXxm29gYuQWWy+bvvOcOLWdcJ1j9lQrGyEopQXvI2tZ8pKbtIqhnmUegwJNACYPSB1nJUxXIIqei+eTjd9hrAhVOgtuIfAzPFnFt45kXnjGU2aqjbZ03DgsgdOZiTt58eYP67d595usZQRU6A5+oo8A/HOJQ3QUKkhOQYWwYXPG0QpZXsAJA0M0ndwlE9EVzBUuBxiNQ4UkwTy+GQu8dlTPKCIqe5trCK2aDn37+6aD7Iwi0HAhve2+eQ+stS0ueNeJyE6yfbf+QPkiVrrrCzEep4ANHG6WU4I9VKq6UPPPC0xQFAmPBAgUKwiSwM9hqN23L/GIEZhewMCY//9KL5mKaHzAzxmmtOYTqAdAFCxWiSPhBIYFcQdovsO4cfjRlqHOhOtxfwUmWs0Eddjc0pH79EeRRX7sI8bEnHpXLr6go+/btNQoFv362LCyAMChUAd227UDsGobIu8FP61S/fn1zuYXxaAupg3/FCs61SzwsUS/Wq1eP6qAAqYd02r7IBZBRVIVcGUUJQH1Qp+QMgOcVwJBrYOf8DrNSyrjUIDe90bWT0C55KPNBOz6BafN+cShgBaD92lOGHQG+HiAhT6ioN69uw7qyf1+iBtrTFPkGpZSABJ554Vm52ZGdEIphnVgj345OIVeVMfoCVDE+7LTT1CRh0RCYTSb5p3W71lKmXBnxox7JTcxBPHjgoBQoVEDaJbQLWlvkimbNmtmmvjGqZvYdIyWUD7cjDj575dtBF6IlrV69ulkjnTX/rqp8ldTRa8vnqrxstmng+UkVQJBBMFS5/bB6ezVOth7MjsbD5onhVdkASCVqPGwYXAmljoBwDg+JfcWLTTgA8fHxxpsTdwj6oq6kHwEgiIuLI5kiwEt7C3PnzSMnT5z0FgfySinZuX2n3HjTDdK6/WuBchLYZLiVFgoQMJCiYcHHicDtPTYUYxf9vYH3gGqynraucJHC0rFbJ8Gyz0FXKiX1sG2Jj2k1PJSHQ1dOq4QpI8AG4wOHQE0+VAA54U7CZS7U4qHaQYmRUXgvt81LdV6SB+MfEj5c55b7pUE4fEGyarWq0qJNi6Am2FRAmlCZoIqzzCCsc16Qy+wQfMCjc8/OBjhRciilbFXIOCyAgH3z6gPFH1JxRwAjhtO7Q37ZHLcPabAlpJL+5AmoYmFhoATkWSg+xEDaDatWrTL6cIQ9Www29AYAABAASURBVA67RF+Axpa5MQfQzZPmXThQpEOF05od2Lplqzz86MPyyquvBDVDj87zEGaDKnQGrQ9CPZooAiwgLAGHUFcH/qFOtRgQlwpbwfsAHEUuvlj27t5rNtLWhYo5dAjxUPoEvfkcAtuWdcanDQoeyjBn24aLkQUqV64s+Ee57eIff0RqaQDZsX2HQOGVSv3AKaWEu/81H42XF2sHf9sMIMV4jCzkPic9ad6T933qqafEPSsFChaUrr26Sv4CBQTqwbqlZdywAMIAJ06elEraoELaDX6Hz9ajJUJfDkYNR7bBxrABrkuIUsp8dgZh06uiteMT8/FiWLTbb7+dbIrA4niFVYTz0mVLC1g3RQenQCllPEERIl+q+7K0atsy6LCiS8clGww1ceJECcdyOsOaL/0BQAAYLtxuXbny5WTQW4MFzdTOHTuCnue280uz2WDmohqwBg4fKNc49hHaY2nmgAO0ICjKUgu8E598ZX2xpnPw3D51GtSV5q81l8R9iQIgKpU6cNj+ANMeLce+VOcVada6maiIM31R6sCmg2DTs7a8F+/H2vK+9lnEGASHjBgiZbQ1f/u27cJ6UZ6WkCqAQIouv7JikOqTgTEwoSUi7RfQfCFfoO9GfQlUQ6K5Ow3rweEGyLzY1Y7FXRNYJLBJw4YNhb7IIzgyoj3BoBZK380YtPGSa7QoYFhYOtqEC0op8zV3sCOq376D+5lvPrl9+MgyqkhYR/hh5gz1xCCFSwMHjPvqaFBQXsB+wnLt27fPHUaq31ld+g3tL6VKlxKep9SZAxPUMEyGTccto0ChgtJzQC+5v+YDQa1hkbhpCYsLsLCmzI05wg2wF9glUBNbdvjpp58W1tEdCHm0c88u8rJGHPxZARCRUumbr1JKoOIAyeNPPyG9BvQWL3uHR4ZdW+bDvjM/u7awjnw4kPcAKHgv3i/ZNBCY8t3/ulv6DuknRYsVNWvLOkk6/ksDgByWMmXLyLVVrg0alokgJAcV+mRKly4tuJqgh2Yz3n77bYH1gPf2aR5UhAoP7IUQR18OH990CuXy4HZGkHPzpG/QhjU8S2EdyacWlEqiJJs2bpbrb7xeho4aJjUfezjFZ1SRkzhgfBkErQvqRthEDhjAwUF07Rv2uSVKlZDmmhdv16m95MgRJVCB9G6gHYuYvnt27zHUr1Xb1tK2UzspXaY0VUEBzQ/7wdyYI3dnAHCQF5Qb7RDv5HZSGssDdENGDpUad9WQzZs2m0POM912aU0rlbS2fA7oRr0vg0YMFlTAkZHBt/uYB/OBVWV+dm1hHfmsK+8BIvU+N65MnLzWvo20eaOtURCc7dqmCiCQw4jICLmtRkpWBvUabIh3chc6D3vnZa9g+xAO8StTKu0YT6mktvDNBQoWkFavt5ZBbw2R6ndUT0FV0/reOMjVqV9HOGyPapUj2B2VbYTHsTKt47ntGOPwwcNao7RX7nvwPvOM+k0aSMm44GsGksb/kG1uv7O6DHpzsLTp0EZQpyKfoQVUKmlt0jhUimZKJfWH7eGbay3btpIBmkXkrEV5riSk6ByiAITA+7K2Dz3ykLFtwemwLiG6hC1OFUCUUkZgvP2O2wWM544G74nazy37b0hjOPLOA6x36WWXBjn9eduEy7PAABcftsZJsH2XDgJf27RVM7np5puEL06iefKOAebFu7V8hcu0Jfkx6dyrizkEz730vKFEfLfrxIkTwvjevmebVxrbg9g4yHzs4annnpKBwwZK1z7d5ImnnhBYTWxbXmzN83LmjBbY0Gq33myoGwetQ+cOWja6wrAoaA0zcq48k/FYW9hLZLAOXTrIYC0zNGFtb7lJcFPBHkdbN8DG54vNJ5drGZl37N6vp/QfNkBIaz2LsLasA+O7/dKTThVAGAyhFusri0veDWPHjjVfEnfLLmQadsaP5MY/Fm9c90MZstIyZ6WSMB5sDAellCbjaHI4ePC5CMjIEm907Wiwbfe+PQzm7T90gPQZ1EeatGxqfL2ioqIErMln+M9l81KbM2OzdzwrW45sgi9UoxaNpffAPtJfyzwD3xwkPfv3NKxYQo/OGnAH6AM20PDsXTQgw/KUjCtpBPE9WqjmeUolrQHpjAxKJY27R7OIAEtcmdLyyOPx0qVXNzOfARrAOfyduidoO0p7I7cM0PPnPXoP7C1QjarVbjRIh/dFzuT9z3WOaQIQMNKuHbs0yb5fyl9WPsUz0V+7+uYUDc5TAVZZ+Grv42A1rrrmas127DP8qLc+vXmllGBkYiN3ao0TBjK8gMteWk4qX1NZELrvuf9eqXpzVcG7FdUyLAkbxwFg85RKOhDpffbZtFdKCYDCs5kDSKJYieJy2eWXSZWqN8pd994lt1a/Va68+iojb6Ju5p2wbfCOzF2p8zNfpZTxcuC5PD9RKzRAzsjBGPluq3Gb3HnPnUYmvKxCeU3tihmZi/fi/eBqlMq4uaYEkBA7cPTYUcmeI7s0atEoRQu0RVhj0UykqDxPBbhWxMfHp3hanjy55WVty+CzMxyMFA3OsUApJRwgPFbZVIyIbJQNUBpsHSgGlMq4jTvbaSulhLmwHthP9u7ZY+wCzBeggF8/cuSIeSelLux8lUpeWz0f5sX8mCcBDRrzR347oVlUpTJnrgZAIEU5cuTQmpSUAaBQSgltgOhrq1wvtevXFu9/GPIAEhbfW5fZeagXmhg21vusxpqtgYcFiKlD4RDqXbPKU+7//+qacN7NeeEHQQ4o9IbjJ46LdVgE+4Ipd+3YKRjP7rnvXroGBbRHt2vDHW4OQRWZmOGDaFycAUi8j6n5aE1BkwH5pS5btmwSHR0tvJf3XbPyJwyrkrUOSesATIj+z1AQ/s5ErWdelAav1Dd/MbZx3UYmblK3sTSp10T4M8tFil5syC48XuLeRGnzRhu5svJV4v0PAyCf3MfF2FuX0XlsKjj6wdZ4x77hphsEDdPuXbuNGwQYAT8n/hIS72XfMStuZPY6ax2S1gEYABaACc6UARC0KVs2bRHUguj73bBFG4SGDhwqJ44fN1/BVkoJPDUUBc0H7AsDuQEKgvclls3MYLlwR6hXr57wt7Pd59o096K79Oyq53xC+MABlI/vQC2cv1A+nPihuO+Xld6WtR5bz6wBMLBFwwIwwXkyAEIiXFi/dp1MeGeCoOtXShl5ZM+uPZI/f6xWCw6Q8lqbID7/Wd8Y+6VtnybpKoJSYEnnUj8fnfPrXOOuO7TqspeZI0KdUkpwjzhy9IgM6jvQr0tWWdYKhFyBNAEIvd8f+578NG+BVqtdYtRwCLsYdgCaQW8OMqpN2nkDwvtzzz0nUBQc9PxcLrx93Dw8MepbAAPfK/y4cAN329j08y+/ILgzK20oQ8tBjAWdCzrDBw2XTRtT/pFN2zcrzloBvxVIM4CcPHVSEtonGDYMXh4WC0ssdyciIiIloWcXefbFZ/2eYcqQSXDxxmkPl2Yu0HOdEh8rBG0+gYmPEE5+EyZMEHxs8JjFDZ5PuwAYoW68YXhr17m9NGza0Ng69ifuFwBYKSV8RIC/MDVj6nQzj6yfrBVIzwqkGUAYFP1z+9bthAtH+CUBJBzExMRE4/PSQB/QdgntpUjRIjQPGfimFc6EfBAOZzkEbS4VVa5cWXDy48IVXprcuQh1McsOjvFo8IjB8kDNB40DnWuEwzt27pxvZEj/wbb5OcV4jOIghzGSG3AAPBe8ateuLZUqVQqMzb0Q3F38ZCQ0afThqjFq8UCnDE7gis9cWWPv0HgVU8cH/Lx1IDDqQFzcAyGNoc7bLjPyyIk4snJFe/r06cLNRc5EqGc9+eSTwvzC3WsP1Tet5ekCEAb9fc3v0qFNe4FqFCxUMIndiogwf4AGgff+h+6XoSOHyaP/fixdznwI0oyf1sDXBpu2bpYsA1WQzX9tFPxu6B+h54ND4E/zfpLO7TtTlCEBr2TYRC518bVI7kXjco02DfcWPGF5UGysls369xfkJACdMhvYdPrgCR2tVc62PKNjgJa5cjcCgHTHx0OaOu/tQObPxyOo4/Dhbk6aKwtoJt0xMjrN+CBOOAUAFxf3hIQEgSX3u+ocFxcnADHz4wstxYoVy+gpmfHSDSD04uC1aNTcAEWRi4sYIFFKmRg36Dx58giXaYa9PVwDyqNS6KJCdMuQcFmFy6RR80YaCIcaxzu0VDt37BC+VMEDwNAAxzezv5aWjVpIKHmFtukNuF7TJ1FTTGQiqAAfMIA9BCi5D8KfBIN1xEWbtgCCvevAptKPcljIjz76iGSmBO6G2IGh1u4XY7gCTR0Xv4gJfFmR+ZOGzQWoYIO5OMW1A9hg7q9Tn9EhMjJSuOtRpEgR4cYmtrTq1aub7xjwLPzruPNB2gaA16aJcX8nzuhwVgDCJJb9ulSaN2hmVIQl40pqihJh7CQcFNwCtm/dLqXiSmlAaWHuUbzWoY1Uv6uGXFI8fZCOJf+Kq68U3MJ7D+oj/YcPMJ+6yZYju/HWxL2FZ8LuAZjFS5aQaVOmSbtW7YxLBXPN6MDhA9NxIQyKAFton8HdetKwVygYYmJihMs+lHFQmSvaODR8lJ2PgDUcKmefpZSySRPDQnF1gQy3+Hgf7uAw32rVqhnEx3sgE9ImowNIBRabcQFSLml9++23AhXhDyhxRRkgpd6GWrVqmaS9zw9bbgoy+OesAYR5rPt9nTSs3UBmfTHLaLdwcoNVUkrJaf0/N+dgu1Cz3l/zfunco7Og8SLgQt24RWN5QWueaj76sNTQ6llc0v/1wH0GGF5tXE+69Ooqb45+S/DW5K4AX+7j0KGrxvdJKf2c06eNAyIu2gBJry49pXfXXkwv0wJKAQ4dD+BwoaXjvbkFh6KBchQKXKAizcGCklheGRkGHyLqMjvAHuFmw0UjlB08T6lgAOFbtpbd8wIuHgpoIMHQ3js2jJURAYqWqKkyY3H7FNnOUgwuxwFA3HqknkBZuXLlBOoGu8jlPVT/sGnUZ2Q4JwBhItgaOrXtKL30oTx+/B/Bc5UDxIFRKmkjuLa7Y9sOwZcrZ86cUvHKSlLj7juEq6y169cxVvmEHgmSoAGoY7eO0qJtK3nm+Wek2q3VBHYJKoFvP451/NVXpZIAg+fzB1B45uJFi6WptvpPm5x5bAvPI7BheBTwjmws1OS4NqTWqVMnIAfRDh6fu9wAE+wYZXgaZJRdiPFSC7AsVlDn06O0B3iJmT8x6nNiviHlsl2UEfhuAFdckU/IZ3SA47BzrFChgiDbIY9gEGYNYb1AjPa5yIKkARrmxtVn8txlJ87IcM4AYifz8UczNDVpaNibbNmzCerVqOgoU81GKJV0qHEo5KI/GjEAZvu27YLlkjTXIpFhtm7eYi7nYMtg8fhzvWYg/cNYEdrOgSBcvGRx07dH5+6G3Vu/br1ukfn/0JTBk2OfmTt3rsAyAQR82KxMmTJBE2DjmTOFvDtARPp8BYRXPoEDZeOgcYj4K0vu8y05a4yGAAAEUUlEQVT14L3c8vOZ5ospyGgI3fxJC55NHoqGp7YFYtjoJ7X2ijVFCUI7FBHEaESxe5HOqJBhAMKEMMT16dbbCMeff/K5YX24d4C2i4krlURRaOsNvDAskrfc5uHdoT5Fi10iFxUuLFz1HdhnoLxaq67MmDrDNsuEOOWQuNKgmoZlgT2B3KPFgo/u2LFjUAccOBctWmTKuLeOQG8y5+kHRMKjYEWI+bYXGi7SSiXth+XjAW5YRurcgBYJhQKCvFueUWmUGFBlKBjsFesIQCCPwD7xDmjYeF58fHzA5Yl76vajHtQxDqwk6YwKGQogdlKrV66WHgndpVGdRjK470BZsWy5UfkiJxDYBLAWWBfAQfMUERlhboNBfSjn+iofWAC4oBSFChcSMPCXn34hXd7oInVeqC2TJ35oPhxgn3u+Yi8gI2uxuTyfO9vEbsB3jTyyAPH5DHauyCJg3BitNPD+FSmoIXNiL7BDkLaBPUCLxKU4MLotz8iYsVk/1+aFBwYyHKwdzyqV/Kf7LHtFGX9gFMDnK5LkCbVq1SLKsJApAGJnt/HPv4xzYNP6TaVl4xYySGP8KZOmyJrVawQeXmlWCWCIyR1jPvvCrTyoBGo/xuDgzf9hvrwzcoz06NRDs3ANDODN+Wq2IIvQ5kIElBF80wttD5uEqpevgjAXtC/EbuDgkUc2I75QAaMa7KD3+Xz2FL6fcgAEOw3fHeMdYSEpZ7+w/ZDO6GD/iCuaKthBKAiAzMHn+1g8DxYMpArFJs9npMqXLy8YEqHglnKzJ0VD/P0a+qU3RKS3w9m2X7VilUyeNFlwGGymAaah1n61btJK2rVqa4x5aJ+6d+ounV7vKK+3eF2a128mDV6uLx1eay+j3xotM7/4ynwW52yfnxH9rB0AmwLsEp9P5TtNYDrGBxNjRCTtBtgH8hm5cYwXLti5IoPYdhxyVwbiQNo6bB6oeMljA+G7Y7xjlSpVzF1+VK5ojajP6IB8BPAyLl4IsKto+dCaAQQAEEBrARSKjNzBd5rpiyIEz3H6w41YYCF/ruG8AYg7Ucg+Qvna1WtlqbanzPv+Rw0AM+XrWXNk4YKFsuK35cJf1mUh3H4XOs2B4dtco0aNEjRC8L9jx44V7AVYf5FL2FjvPLEpYAizB9Bbnxl5ABeVKcEdH1mCA0Q5xkpbx548++yz5u+TI6egRcJ/DoAHo2M8tG0zIwYAcI95++23hUPPn4LjHZB/AFKeif2JdUTxQd4NAD/qc+pRTbt155K+IAByLhO+kH2nTZsmYGA2CF4XjRA8MdiXulBz69+/vyAY41sW1CYTM/gz8SVKFwjs4/i6I3UWI9tyYtw2OJQ4lGL/wOeMP0NBXWYH1LUYWGGbcFJF4AZY7XMBbNbRyiW23MbIStRjc7Jl5xoDIEmqjHMdKat/1gr8/1sBBYDk/v/3XllvlLUCGbICuQGQ2Xqog1lBstZAstbAAwez/w8AAP//SOl2egAAAAZJREFUAwD2BtwtebCIwQAAAABJRU5ErkJggg=="
GG_IMG_B64 = "iVBORw0KGgoAAAANSUhEUgAAAMgAAABNCAYAAADjJSv1AAA//UlEQVR42u2dd5hU5fXHP+8t02c7SG8iTSk20AhRLNhLNBYEewuWICKKBVAUNWqwxBjz06ixd2PBWGiRoKgoqChFIID0sruz02fuve/vj1t2ZllgQVFMuM+zD8vu7C3vfU/7nu85R0hLSokEAUi8wzRNAFRVRQjB7mP38V95WBLTMkEIVEUBIZDSFgQhBMKyLCkKfmhZFoqqILCFwjAMLEuyW0T+Cw/nnfM/qgClZaHpOoqqOP+XWNJCUZR6IZH2gRACy7SFwzQMJk16h+nTp7N69WpMw0ApNjC7j1/A0XDbS6h/8QiEKH6nosFn/6vXxjEKus9H+7ZtOXLQIA4/8oh6I6EoWJZVb0HcHy5bspTrr7+ejz76iHw+Tz6fB0eAdh+/QAFx3puUElVR0DUdKWzPwDRNCr0HUSAVUvz3r437uD6fD93n4+ijj+b2O+6g+R7NPXkQlpRSWhaKEKz8fiVnDx7M4u++w+fzoSgKbdq0QdU0pNztZv0i3QjnX1VRSKVSVG/chBRQWlZGaUkJedOwrcn/4LooQpDP51m5ciUA6XSagw4+mL8/9RQlZaW2BTFMUwrnw8N+9zteeeVVIqEQe3Xpwqhrr2Wf3r3QdV+jJnv38csQDtMw8Qf8TJ0yhWuvuYZ8Ps/oG25g6LnnkslmbCDmf8GvarA2Ashms3wy62Pu/sPdrF+/nng8zuVXXskt42/Fsiw0y7LQNY2vvvyS99//gGAwSKtWrXj8ySdo07bt7l32X3SUVZQjJRimSTgapbSslBJK/+cV329++1u6d+/BkLPPJp/P88Ybb3DxpZfQpk0bNPdDn8z6hFwuRy6b5cKLL6ZN27bkc3kb0VKU3dbjl6glndjCNE00TcMwDBC2q2yaJlJKckYeXdP/Jz0ES0qQYBkG3fbuwamnnspf/vIX1q9fx7yvvi4WkJraGnK5HOFIhO7duyOl9IQDAWJnR22C3TDZTkFrwELYmL7jW6iKgirsn6lCQXEBmP+19RfSzn/oGpa06NW7N4qikDMMNmzYAFAvIAIB0rLhXmzUSiJRhLDRDdlQxUj7Ap7gbFuA3I8KKTf/vOV8Rqm/lNgtMD/s/SsCrHoEUggH75f1srCj6KRlWZ6FQsqfLpciaESYJUIo9vPKJj6TAAXFsSQWqlDRfT4P0FAUpaGAONcuQKsEwtvMUjSEeh2pERRjgo2A6ZuJgwvAu//IxgOonW7z5RYWf2deYwtrtNVrb+sc21JKBTrMfmtbQCRFYy9hc5fN9S52zeBb2nu1CQskGu7TQoF3Dq0pomYJiYVEk6JAvSveeyv8sSKLX7olbOMgALXhwjuKxxIgHN9YSFB/Eun4iZxusR2CsCvddyNC6m4gRVH44vMvmDplCvF43P61ZaGoav0m25YiaiiAhVao6PstP7ubv/P7fOy7334cd8Lx3s+E8uMsktaUlVGQ9Ru/wYt2n9VyTyaKF8ATDOcP7VMI52NOEGnLPTqOO/dfkJTc3Gpu4/c/QCDkzpAbsblmNk0TXdd56cWXuG7UKLLpdNFmtixrm+6Nu1cENs/P09zOV+Hfm1Ju9faEIyRCCPKWxSUXX8wt429F07QfbRmafiZhu1NS4gV7rmukCHuNBJZnwC0kQoIi1KIVtwSecFiWhVDtm7DFUIDysynIH/W64gf+flczKFJKdF1nw/oN3DFhAqZpUlZWhmVZ242/KKqKZVlk0mnyjtAF/X6P3lFIFtwSMldoRWyXXfLXv/6VgQMHctTRgzBN0xPAH11AZMPHc2Jxy3GnNG9HWSBkfULKMY8qquMmQY2U1OZypLAwJPiEoExVqdJ0dPcBLPsc3rNu5aUXkirtxdnaTheeP1kfqBbEP1sg60kpC373w2OPwmsXndsFRwp+19SgufA8UkqbFlFAK9nWeXY0OF/83Xds3LiRUDBo05CaIryWRCiKHQxrGvFYDIBOe+5JtKyUeKyOJUuWIIBIJIJhGLawmGbRuynEhIQsXmZV19BUlblz5nDU0YN26PmKaDdieyyIBOHF/K7MWFgYCBTHmghUVKQCSzD4eM0yPt+wkmXVG1lTXU3WMrGEgqoqlPn8dCyvontlS/q26sC+kXKiAIaFqdpn1oS2WaBVaMK3qB1EIzin+/emhWEaqJpahHZY0kIRCtKSHvrzg7lnjQS4Hr+noUBacode5pbucXsEZUfQK9GIJt/qoQgsKdE0jVhdHT179uT60ddz0MG/IhyNkEwmmfXRx/zhrrv4cu5cSktLkZa1mXBsS2EIKX9071xr4tvwACs3KLdccbYEUgoUVWFxPsubC77hlSVzmZOvISNNEDoIFVVTMIVlR/FGkmmr16Ou+Ja284L8qnkbzuixH4dXtCUCmGYeFLURzW7/KJfLUV1djSKUJpl1v89HJBJB8+lF6Itl2sGcWngeIYjVxsik0z8sFpI2sKEqKlXNqrwNbVkWGzdsrPe7hWCPPfbANE2PZt2UTR2LxUglUzYcKaG8ogyf319snQo04o8dnjQ1beUyiFVVJRavo3efPjz/wguUV1bYXodpEQ6HOeKoI9l33305e8gQvpw7l5JIxKtJ2iI651qVIqMtfnoBkcL+Upw7sZykhbDsApO0qvDWyiXcNW8GX8fWo+o+pC+ILjRUB/2w6nFkECr4Qmh5yUorz3MbFvHmv75jSJt9GNX7UPb0Be0alM20sIUQGo899hiPPvooJSUlWIa5TXMZ8PspiUZp3rw5Xbt3Z8CAAey3/34oquqZc3sDQyIe5/Jhw1i0aBGhYBCzCT52o8bDufl8Ps/tE27nyKOOAuCdtycxbtw4QsEgCEEymeSy313GpZddtk2/2RWemk01XHTRhaxZswa/309dbYyLLrmYK666EsuyUB006Ue1IHLzkLTJBsRxlzTdx5ixYymvrCCXy6HrOkIRmFJi5fNUVFUyYcIETj7pJPKmaScwXTdSNH7NwgyD5MfPdWrbtTZO0kJ4dyyoUQR//PojHvn2Y2p10MIhhFQRUmCaObK5vINbSRQhwbLFy/L5sVQfPqHjU3Ryislfl3zOV+tXc/fBJ9G/ogpTOonKBjfy/fffs3TpUsrLypGGsU3f2rIs74s33uCBBx6gX79+DB8+nF/1P6ReSBSFTCbDku++Y9X33+Pz+bwgdHu0sXcPikIikfDYogDr1q1j8ZIllJWUgJTkTZM/3PUH+vbrR58+fbYuJE7slE6nWLhwIXW1tfh8Pmpqali2fPnm8dOugugJQTKTplev3vQ76CCklGi6Dt67lai6hrQkvXr1pHfv3syePZvSkhI7DmlEGLZmrX5yAXEhXg+isyQKgloFbvn0fR5ZMgclXIJPSCwpMWQeK5un0h9k76o27B0op0UkgoYkaRosqavmq0Qt/4nHyAkLEQqAAaFIMz5Ox7hk+iv85denMqCqORksdMtCtRRvhfxBH/6Aj6A/CLqB1UiQWhQEi/oNbjqCMuPDD/n0008ZP34855x3LoZpoGCfIxgM4vP5CAQCNoAgHC3oJlG3pZUdze3SO3w+v/crXdcJBQIE/H4s0yIYVKmLx7l13C08/8Lz+AMBL9YSws6Ee9cT9Ro5HAyRTabw+f34/H6CwcDmYMSPuWE2Iz40/byaEGBKWrdogaZrXixW+I6k47PpPh8tWrSAJkDGDQ8VUMXP4GK5wmFn08GSgrSicM9XM/jrok+xKiqxTNAQ5PMZWksfZ+95ICd07E6H8mY0AwKOhGeBPLA8l+GzdWt4ZuHH/Lt6Gfgj5A0FNRRhQSbOiBmv8X9HnsH+0XJyUqIqAuEEs9K0kJZEGiaWtJGORCJhb37H6riMAInNXnVjkWAwiJCSaCRCLp/nxhtuoFnz5hxz7DH1OL1lB9TSlEhpkTcN0pmMrfGa4LZIKZFCoAhBKpUq1oJSYjlEQSkt8lmTSDjMRzNn8te//h/Drx6OYRhoTg0OojjXKBwg3TIN2yqaJpZlYZoWu9rh1lxg2Ro2FovZiFZjn3UoMYZhUFNT49BGfn5L2PQYBIklLBQJmqLy3OqlPDz/U4zyChRTIBSBkazj0LLW3Nz3KA4va2ajXpaJFBbSsjCkhU/VCJqCnr4Ae7ftyMC27Xliwcc8+uUnrA1qKJZA9weZm45x6yfv8egRZ7CH0BCApThaQoJu2prUNG2N/fvhw+nQqRO5XM4uvnc2smGaJBJJVq1ayZdz5zJnzhx8uo40Tfw+H/l8nnvvvpuDDz6Y0rJS56U44Z4qSCUy9O7Tm3POOw9LSqSD4mwL1JAF22TgwIGNuk5S2oiZZVmEQyEe/vOfGXj4QHr16uUF7VsFTn6yLMiOGx1p2fFnKBzmq3nzWLRoEV26dd3MirhJyBXLv2fevHkEA8FdwlXUmvygWI7Z01icy/LIZ9OI+XyoloYuJUY8wUmt9uTOQ06km6pjmRY5xURRLFQpsBQFVfjsja6CtEykNGmPwphuh9CppCU3/vsd1gdAtSSESnhv0zKe/O4LRnc5ENOwvL1gCVAKHFLLsjj1tNPo3GWvrT5HJpPhlZde5tZx4xyta2vv+d9+yycffcSg4461KdAFsYRhGrRr354zB5/1gxY6l8ttJiBCCFtzKrbrFYvFuG38bTzz7DNoms0wVRW12M3aQsJslxUSIbCkha7rbKquZuy4cTz9zNObZbs1TSOfzzNu3Fiqq6upqqggn8//7KXeTWOceXkOOzB/7rs5fJWoQdOCWAjMXIoBpc2471cn0E3VMSwLqQqE0BDogI6GjmoKhANnqVJFQwNFoFoW57TqxC19jySSSCIUUAyBGQjy3LezmZ9JITQF4QTMeVWSc4hdQlGwHNjTNE1yuRymaWIYhl13nTe87wOBAEPPPYfzzz+fZDLpoT2maTJnzpxGXSVFUcg2PKdpbvMrbxjkTYNsPu91ymjMobc7zdiuRUlJCdOnTeOxRx+tR6K2BrWK+vhqV2XnSCSqrpPP59F1nX79+m3VKh5w4AEEAgHy+by3Zj+nkChNtpVCgKKyysjyxvKvSIV02+0iT7kiGf2rY+ig+ZCWiYaCKgW6FKhSQQpRAPOCJSwsxfHTpWq7YNLgrA7dOKtTT7LpNKYQ+JUgS1IJXlu1oOhGhaVgCjsm8gI0VbW/NPtfTdPQNA1V11A1DUVVyefzWJbFgEMPJRwOezi7oiisXbeu0ZfhCol3/qZ+aSqKoqDrWiNBvSxyswzD0ZSWRSQc5k8P/on58+ejqVpx/EJB5t9F1aSLXO1CQlEQpymKgrQs4skkE+64kxHXjEDTtM3WWQiBpqoMHz6cO++8k3gyadci7aSE548rIIApLSSCf61ewZJUDOlXkYqFL5bg7I59OLS0OdIyEApIRWIK7C9pZ6otLAyRQ5JFkXksYYIiPApLjjwRTIb1OoQO/jCWZSAsSdqvMWXZApJSeneryfriH6QsFp5GfHLpwMwuzz8SjeL3+z0BkVJ6gXxjEPFmAXgTvgQCxSFfFiazcJBACZiWSSgSpmPHTpgOB8mn69TW1HDruFvI5XL1VuQXyt8UisKmmhoGDx7MBRddYFc1biVLb+QNzjnvXM4991yqq6vRCgiNP4eQNElApJCoCDLAR7HVxPJ5VEXHxKRlIMRvO/VBsyxMoWChem/T5uabKNJClRJN6CiKHyn81FqwOptlo7RAaIREEGFKOoWi/LpDV2Q2jdRUhK6zLFHDglg1mqI5G0xxqPWySZkhgbBFxEGnamtrSafTXkwgAL8DxUohPfq+a5kK1XN9UnHrXxQkOr2seYEroXi0bIXRN95Am9atyWazSCmJRqP8a/p0nnz8CRvNsqxfVLWfywhQhCCXy9GiRQtGXXedZ423iHg5CkxakmtGjqRVq1ZkstkfH7LeKTGIopC2JIs3rUPVgmCCsCxymmSllbLr1qVNUZZSokgT1bLQpEBRNHJC5fNsigeWzuPy2dO4cOorDJ38PIM/eJHfzZ7MO7VryKsaISQDKtqiZTJkclnQ/GzKZlkYr8FyzbZjOUxRTy7wkoGmVZQY9KBQ0/TyGf+aPp1UOu1UTYJhWbTv0L4I0i5ysTyqTROThbKegN4oZ0raeQRF00gkk3YHmeuuqw9KpSTg9/PA/fez4Nv5KJraJNbsruJeeXUjqkq8ro6TTjyJNm3bYDl9uBqzBB51RYBlmrRq3YqTTz6ZRCJhK6kiZHAXhHkFkDbyrInHEJqKZUkURWNDPsO4f7+NNuAUji5rgd/VBqgYAtYCX29YwzuL5/HB6kUsVvKY2TxKJkNpIEi56ue7TWv4+MtPGd57AJceMIBDKpszuE13vq2rZVU8zoa6TSzIbCJFZyKugCAwhQ2TSgsi0aingbZ2vPDc8zz//HNEIxGbJKiAPxBg/wP23yyEllISDAb57LPPOPnEE4sLerYB84LdnO2OO+6g9759GiQvnSpNKdE1jbp4nJN+cwpvvfUWb77xBmXl5fgVhdrqam6/7Tb+/vTTXvksqv1C5C5uPYQQmKaJz+fj6GOP3ox93JiQKI4QSCe+OuGEE3ji8cftBnfuuv7EVmTbAiKoT74hyZPDwkBT/ZiGhfAFWJTLMmzKy5ywR0f2bdGWiD9IIpVm0aaNfJmo4ZvYOmIyi6WBblgcWNGKkzv1YO+K1rTyBVCl5LtENflMCgPoEi7h74efxkZgTmw9by+eTYnlLLAQWE7Qr6GQtwykEEydOpVly5eRzWbtPAh4xVfZbJYVK1bw0cyZfDJrFkIIdNUO3OvidfQ7+GD2329/x6QKt/TFw+pjsRjrZ8/2YFllW76wAEUoJFJJVq1cubmAOLw21RES1zrceNNNfDFnDhvWryfo91MSiTLlg8k8+eSTXHTxRfVQsbScjYSXSNzVQhQhBJlMhtZt2rD33vvYDSIcaLfRWELWu8OuG9t5ry60aduWVStX2ry4BvT3XUNAJGgSME328Pk5e8/9mfjFNGKGxBcIgylRNB/VGDyxfhF/X70AVQhMITA1AYqCCCj4ExZtAyWc2/tALmi/D601HavAx+sTLSUPaEiMXJrHn3qWPUqrOHjAAI7a/zgMgFwefDqWYjm7307aKcDtt94KQnjkwsJNU9hzIhwKOcGFQiabQff5GH711fgdqkbDF+AKSTQa9Sggcts6xeaQqYrXCKAxh8LtAeDmBDru2YlR113HiOHDCQYCSAnBYJCJE/9I/wH96dq16y7vYnk0GSf306lTJ5o1b7ZZYnBri+e6suWV5XTs2JH/LF1KKBRyYkl2PQtiqyoLgWR0j4PYq2wPHv5qJp/VrSWHQFVUDL+K4gshNGHXisi8jdcYOVqYGse168HFvQ9h30gpGhJp5RHCsre4VJCoNs1DgWQmw9/++ghLFy+lffv2HHhAX35/5VV032cf+6YtiSJt0qNdtWhzk4Sq1JMLG/SZLeztZVkW8XicYDDIHXfeyaGHHerROxpqQdWBhxOJRNNdLMfVSKZSZLLZRgJMWY+uFZAgc7kcZ5xxBtOmTuWNf/yD0miJnWDbuInbbh3P408+ga7rvxD4ynaxWrZsWRSbbBGJEg28FtNmJbfv0AGzgDa0S8YgSAs0jUwqwXvvfMCg447jpGOGMqV6DZMWz2PO+u/ZZGRJmTacqwtJiarRPlDKgXu0Y2CH7uxfVkUUsCwTUxF2gtBwqBaKaguKYaEqCptWb0TkDKKlYWritTz51BPs2WUvevTp6WwsOzsvvMAZDNMAEw8WFY28LMs0yeZyBMNh+vbty3XXX8+v+h+CaZmNB46Om7DXXntx1NFHe8KnNKFaT1oWUgh69erpCUyDxFKjb1xRFW666SY+/eQTaqur0X0+SqJRPvjgA55/9jnOu+B8m4ayC2fOC9egsqrKU0pu8nN74NpoSdSrLFScllS7nIuFYg8/CPoDvPT8s9x//31cfuElHHn8cRzZ9yg2AqtTNcTTGQxsU1oVLaG1FqbSvYhlIoWBUFRUVJu/pqpgec4GCraWmffNfFauWUM0GkZIhVaVzenVs6d3S3lVOIiW3XFFCAgFQ/b5GhEQRQgikQgtWrakZ8+eDPj1rzn0sMPqmaVCoShn7agrRSjksjm6du3K9TeM/sHBa/GiyvoErPsyNA3TNGnXoT2jb7yRa66+Gt3nw5KSUCjEvffcwxFHHEE4FPr5UK1tbPCGpcMunO2CFlurhCzKuzlKSxHKTiv8+vFQLASm44L0PbgvY8aN5dqbF1H1wL306N6NG0eNpt/++0KoeA+4WtTCaWKGhl28azcxM7Brp5BOZkO1F++f771NFpOIopJLpKmoaEbXLgX+t1TsXIiw+3UZpsWf7ruPfXr1JJfLeRlY19VSgNLyckKhEJqubRZfbGsz5HK57d+QbtC5neW7iqJgGAann3E677/3HpPefpvS0lIURWHdunXcfttt3HnnnUVDXn7KwLWp29Q0TYSisHrVqiahiw0Pn2LHbqtXr7bdY1FPid/1XKwCb+DAfv1otUdLFGlSV1fDW++8iWXmee7FlzGkBYpAsWxXSYjC+mtRpNntro+Oy4Kdq1A1P3O/+JSpk98jGo5imZA1Tbr22of2nTpg5Aw0n4YqBarl8LqcvEXrNm1o3bp1k16cdEphm/zShNjuF/xjwKXjxo1j9uzZ1NXWomkapSUlvPfuu3Tq2JFoNMqGDRuKWor+JC5UA0uxpSAdIBqNMm3aNP704IN07rSnzVNzlVJjFsFNCFoWPr+fpUuXMnnyZCLhsJ3H2lUtiHAo7lJK+h7Yj26duzJ/zpcEomFatmzDjH9/xJuv/4PfnHYa+XwOTbMrCqUUCKVYwAT11VeqsyjSsstwM5k0E+64i2QqSyRUgi4VMhb89tRTPFaofT82vcRFixQhyDu0DK+2u7DBmHTwD2ejyx1sTPdD+k+JRk4ktySMTilwuw7tGTVqFKOuvZZSv9+rfHz88cdRFMVm/DrNDX7KgTdiGwrDLkOwcyC5XI4J429DVRSnc83W3SXXSEgnbgxHIvasGpxivS38reLMF/x5LIjz4nKWzYg9c8hQbvxsLj6pYhgWqurjjlvvpEePfejavSvZfBZVV51OKMLrmCiwKesWYClOlxTLQlE1VAS33jyOmdM+IlxWgSkN4uk6eu3fh6OOPcau5lPqgVtTWJiKBMPWOG4z5kbNudNgQm7NHdlKr6GfSnMVtlx1Ebmh5wxl2pQpTJo0ibKyMgzD8NrtKM6AVdv9++n0q5TSLlBrhL/WmADoul70Mzch6HXhkG5fNdewOKXWuk46lSJT4O7KRhAt973vjLhM257Xpzk38dvTT+PlF1/gi8/nEIgE0fw+1m1Yw2WXXMKfH36YvXvtg+VUzlmKkxNw4OJ6Qp+JVBRURSGfz3PbreP522OPUVJagmHkUVRQFMHVVw8n4nS42Gzjy4KXsm1wnR0qMSqKIbajvZyUHgXB5qRtLg2iQb6m6HuhIKUdw908dgxzvviCmtpau9jLSZpaTmMDwc6nhMuC5gmmZXHu+edTWVmBUdiNxclfuARDrwlcIZdK1JdBe3mlgmSp7VxIh16i2A0GpbSDdemUNzjnVdT6rL1A8Le//Y1c3vi5LIhdpGSaJsFQkLG3jmPwmWdh5HL4fDqBYIClS5cw5OzBjBgxgsFDhuAL+BtNAqnCbUcKn382m3vvuYepU6dSUhIlb+Xx+3ys37CBYcMu55hjj7U3QgMN4W4ma2doeVmAxMp6n9qFlLcpI7JeKF0B3oxGL9hKrYdzPeeZO+65JyOvvZZrR43C7/PZUKdowFzeyQbE6yGGzay4/PLLadOu7RayAvJnaXA9adIkG+7/OQTE1RCqqmIaJgcceCDjx49n5MiR3iz1YCBIXayOG264gZdeeomTTjmFAw88gI4dOxEtLQEJpmmwZu1avvl6Hu+99x7v/vOfxONxykpLkVLi9/nZsGEDxxx7LKNvusHzsTdPtu38w7Qsgv4gX86Zw0UXXOi1U91WotALnIUgm8/ToUN7Ro++gXAkvN172Z4+bHLWkLOZOn26jWqVlGy1Z9TODtKFENTU1tKqTWuvlNit0FyxfDkvPP8CsViMgw46iOOOPx6f34emaTzyl7/w/fffM2rUdWSzGZ566imuGTmSt958k9WrVpPL5vD5fQghOOLII+i8114kEgkmT5nMkUccSSQS4ZGH/0L//v3Zp+c+nvVy49OtUel3Mswri1p2KorAyBsMHjqEVDrNrePGoSgKQX8AXdfRdI2vv/6a2bNnU1lRQWWzZoQjETRNI51KURuLsW7tWox8nkgkQjQS8TbOho0bOeqoo3j4kUcIRyKNNkQWBb761jTxjkfRol4hKAobNmzgnUmTvORXU90ZRVHI5XO0bt2GK6+8inAk7CUyZXFGZOsCAqiaxk0338xnn31GIhbzBqv+lHGHdz+WhaaqnmslHbrMihUr+Ntjf2P27NnU1NSwevVqLMvijLPOZMoHk1myeAmmafL2W2/Rp08fPp31Caqqkkqm+NOf/sTVI0Yw6+OPWb9+PSeddBKKopCIx7nvj/dxcL+DKCkpYdasWXTu3JlefXrbhRXeICC1yNL9tChWQ5q3EF6dwkWXXEzLli0ZO3Ys3y9fTkkkiqqqhEMhwuEwuXye5cuWFVENVFUlFAigRCJeE4RkKoUpJRdefDFjxowhEo0U5SkKfdnCCj9rK2WpOywnqoKiqiiKAKGioNotgLbDigmnIUM2l6W8vLyofFRVVTtXoyheoL3VcykKhmnQac9OjB49mmtGjKCstNTLTqtNOMeP5WIV+vwNj3Xr1rFmzRoOO+ww9urShXcmTWLRokUYubw3Am7ZsmXsu+++LFy4kNKyMgAGDhzISy+9xGW/u4xIOMyKFSto3bYN0pL4dB/hUAhNtRVCxEG1iqz1Tnx2bXsXqOF2Mk2T4044nh777M2DE+/j7bfeIhavQ9d1/H4/uq4X8YfcohnTNMlmMuTzeUzLYp+ePRkxYgTHnXB80eeKru9szlQ6TU1trX0eKcnn8z/YvLov3rIskvE48XgcI5+3k53bQGm2CIUKQTqXI1pbi+W05clms8RiMe9a2WwWw0GltnReexMoGKbJmWedxfvvv88br71G2LG8tbW1ZDKZoufYmVCbO33MPVRVJZvNMnXqVCKO1f/E0fTLli1j1sezOHTgYXzzzTcsXryY444/nquuvBIpJflcngcffJBYLMbFF15ELBYjnU4zZfIUjjjyCObNm8cnn3zCN998w2F7DMRw9stWwZGfB8Xa8kYwDZMOHTow8cEHGHruubz88st8NHMmK1asoMbp5K0XdPQwDAPN56NZVTN69tyHE086iRNOOtFDq7aEs7ub8/AjjqAuHvc6jEfCETp26viDn0MiKS0r5azBg1m+bNlmvW6bCmQVDprJZDJ07tyZsooyJ9Hal4svvcTWqPk80WiU1m3bbHVD2iiYrSBUTWX8+PFUVVU67VkFmXSaI448cqe4GA2fqeH3rqD7fD7OOOMM3nj9HyxcsICl//kPPXr0YL/99qP3fn2Y9PYkpkyZSlVVFX9+6CHGjh3LhAkT0H06o667jltvuYXxt93G5MmTidXW2sLx9Tzuf+ABbh47hon3TaSyWRXhSGTL4xHEz9RZcZsBqSq8jb3fAfuz3wH7U72pmgXz57Ng4UK+//574rEYliXRdY099tiDPTt3pnuPHnTp2qU+KC4IuraYDJIw6OhBDDp60BatwPZuaE9ALInP7+fa60bttE3Wp08fHnzwwc1+Z5n2hKbGM2fSGzgjLUm7Du25d+LERq+h7qQabg/JU0RRd3dZ0EUyFAqxYMECmlVVsVeXLsyePZtWrVoRjUaZPn06I68dSb+D+vHwnx9m7bp1REuifPvtt7zy8st8OONDnnz8cb6YM4eamhpC0TDVm6o55ZSTOf+CC9hrr73YtGmT40lYjT6jDR8ru4aAuDGFoigO7Ci8KjIpJRWVFfyq/yH8qv8hWz2PJSXS6bLesPZYbKEXlFtb7uVWkEW14jvsYkk8NMbjOjUCEGwPUuy2K1UV1cP+6+FqO+WlqG52324voWyWL3HyRw5YYhqmN2fQDfIVRXHyTWLrCuMHHi6qWMgFc/dCLBZj2bJlBAIBUqkUzZo1o7KyEsuyuOfeezzv4arfX8XChQuRgM/no0PHjowffxt1dTFOOvlkLMsiEAxw+m9P98oNTjrpJABee+VVdN1X1CCjMFn4Y9tPTTQG4zU5h1bvi7o3V6jBChs/F27+hiZa1dRG0ZLGLIj70r3v7f6WPxgGFk6Zn8se9e53a2DXtlIhuNQIWUQDL9xcrouiKKIwsVw0A0ggnEe0BUXVVE94xWbXKhaMH83lcE4TdHoH53I5D3Bwjx49evDKK6+QyWRswXU4VYXusltz07VrV+68807Kysro3LnzVi9dGMNOmDDBbh/rAEWFRzKZRNPUnWhBZNPjnPrkWeObpuHibT3Y37HfN5ww1ejfiu2Sks2v8QNRYyEaX6HCa9QjXFtHErd5j4387ocKRLGnJ+nWrTvt2rXj1ltvZeJ993kQvXtdXdcpKSkpsjjVG6uhsKeXtMc2a5rKpo2bvKE8jbrUbsLVoaDouk4imcSqq6fXCEXwf3/9P5YvX87AI47YvmcUOzkG2X38lx4N41/VRh6jJVEm3n8fw4YNo3///kTDYdtNLkAnm6r0dsQF3CzucLpSptNpxo4bR7+D+tl9A34k9vVuAdl9bJcbalkWhxxyCFOnTmXGhzPIpNMet0oRise4/qnvq1fv3nTv0d3uIfYjRiK7BWT30eTg3NXcpmlSVVXFb079za51j6Zp1yApO1lAfs5OdruPnXy4dPGCgL5JgEwDgMQFYGQhDf1n6p/rDUfdSopga3GI2EozjoZtPLymXa6gbA1V2n3s6rIgN/tXOIVLO6L8CpG4XWk/7Gi8UThSfEvn0DbTLk6dgdvG3wPldxuTX6yQuLkdN0fl5mdMbzqV+b+5OM54PXcNrKZYEIBwJOKR4HYf/z1HOBJ2UFZJKBjc/Y4bPHo4HN7MsmoNtY1QFN577z2+/34F6UzGniHegJz2v6N++cWOHSgkFZqmSSgUYu7cuQihoGoas2bNIhyNkkglvVatRbryF/y6m1qfr2D32XLHTsydMxdd18k6zf6KBMSljStC8NADD5Jzhrq4AZ1bLrnZBvJW1OXC1NcN14+ctHeaS6MRjRTCFTaM9nJiFAd/RUJaEFy5PrWiCG+gTFFfLFWtb0kqixNW7oK6iTir4HqFwazDa6Gwjs+7Nzcb7jy7dKc+SZc8IupjQFF/r9IqiPXc9iwNAkb3nIVUHkVRkdIqKIopZiZIac9CMZ17ct+t3+cjHAyh+v289uqrPPvMMwhVRTo+uDfFt0GZbOE9bBaYF65l4b079+ue2ywory1a38LguAFNRmlQSi0K3rUovGahInN+ZBW26RebKz1RsNbOdHN8Pp1oNErKYQJsEcWKRMKb3VyTAznw+lKJBpJgWvWls0UTiAqCSEWxad3uS3E5VltKQBUW8FAIKmATAC0pyZsGuqZ5C9ywLWnhvRZ236gP5OonOLn3Xljf4VXVORbYNAyvsKohrcRdC0taXiWe95wFm004/bEUxW58YVh2h/NCQqNoBFBxSaN24zthr7kjBBQ0cQsGg0QiEW8Tu6O0PUSowfupX39nLaQsKtoqrPdxzyOd+SdF1KCCdXXvxb2me133XFaDNXTrX2igoApJru7z7ygC6z3P1gTEsix2NN2TTCY9Ql5h9wpN11FVFb/frlN3xyO7Wssq2PCarhPw+1GEIBGPYzmFMg0funDBa2qq7VHIzv81TSMUChIMhvAFA6RTqXqtjChuRC3qk2DV8XjRInsaWVUJhULouo5lWZt19XCzyaqmEQwG8TudPNw6FVGwOV1uUfWmTURLSuzKOed8qqp6wXQkEsEwDLLZLNFoFCEE2XSaZDLpXbsQWfL7/QQDgfpWSg4HLOeMnpMFFsmlAsXjccLhsFfQpCgKyWSSZDLpnVdVVUqc+yxk79bU1GAYBlFn9IRwOunnnUpR6YA+8XicTCbjdTcxDINgMEg0GvU2t2ma1NTUeByvSCSC3yk3qKur89Y+FosVcfxUVSUSiRCPxzENg7Ar9DuI0jVEs36URGGhFD/wwAPs1aULuWwWTdeLkJKPP/qIBx54AE3TuOuuu+jVqxfZbBZd18kbhoeszJ49m/snTiSTy3HBhRfSunVr7rzzzqJKMle7uf2Xhg49h1A4hGnZ7kQqmeTll1/mzMFnc8nFF3PiCSeQcaZKFWH2DlXK1U6nnHIKwVDIbiXk9KdSVJVEIsE38+axbt06VFXl2GOPpbSszPtc3hGqbDbLt99+y8oVK/D5fJSVldmlxum0NzcxFovh8/u56OKLefGFF6irq2P//fcnGAoRdzaDpmnMnj2bbt2707t3b155+WUAqqqqOGrQICLhcJF7kjcMFixYwJLFi+0ZgA75UQKRSIRAMEgun/eG82SzWTKZDJdffjnvvfceq1atIhgMUldXR9euXTn8iMNpVtWMXD7PV199xdTJkzEMA7/f7zX/7tevHy1atmTa1Klks1ksy6J3795069aNF198kUAgQCKR4LDDDmPQoEH4fHa9eSqV4s033+Tzzz8nFArZhVP5PMceeyxt27VFWpKZM2eyePFifD4fl156KR9//DFLlizhqKOOIhAIIIQgbxisXbuWOXPmcFC/fnTv0YOXXnzxR+2P9aNl0t0Nl0gmqamtJeFojdLSUnK5HMcce6w3Kdbn85FMpdhUU0MqmSSVyVBSUkI+l2PQ0YMIhkLcc/fdZHM5+g8YQO/evbnrrrsaHaNsGAa6rnP+BRfQsnUrUskULVq2AASv/eN1SkpL6Nipk6fB3WlFyIJWPLLevevbrx9VzapIJZIYpkkkEiGVSnH8CSdwx4QJPPH445SXl9OrVy867rkn6VSKXD5PNBIhkUwy6Oijef6557hl7Fi6dOnChAkTqKiooLKyklqnCvLBBx9kzpdfMnz4cN6ZNImamhrOOPNMOnTsSI/u3VmyeDF1sRjffPMNXbt25YwzzuDB++8nHIlw1llnMeyKK3j26acJOyMZEvE4hw4cSF0sxnnnnENFRYVnjWKxGKeffjpnnnkmQlWJhELU1tby3Xffcecdd3De+ecz98svWbFiBclEgrOHDGHkyJFMmz6d1StXEggGuer3V3HB+edz2WWXkc1mScTjXHHllZx++uks+34F5517LldccQVLly6ld+/enHbaaTzzzDPkcjmuuuoqTjz5ZJ7++99Zu3YtiqLQsVMn7r//fh588EFeeuklSkpKSKfTlJSU0KZtO7As/E6jPE1VufCii1i3fj3z58/n14cdRmlZGYl4nI6dOlFWVsbBBx1Erz59OPXUU7l34kTatWlTHzftCgLimV1VZcyYMaTSabp06cJB/frx5ptvsn79euZ98y21tbXkDQPd72fChAkkEgk6duzIYQMH8o9//IO1a9cya9Ys4rEYhmkiVNUWpOpq1q5f72kbD6VTFCrKy+0NfOyxKJrGuvXruf766xl942h7KKRhkMmknTZV9TGLUBwhcVw8xRknPWbsWNLO/VdVVjJr1iyyuRwfzZxJMpVCURR8fj8TJ04knU7Tpk0bOu25JzM/+ojq6mref/990uk0uq6TSqUYNWoUiqLw1ltv8cgjj/DeBx+QSCSoqKhg48aNtisrBNdeey3+QIBPP/2UO+66i3fffZdoNEomm6WispJrRo7k2WefRdM06urqeGvSJMrLyxFCsHHTJlq0bEnnzp29YaSu8ohEIrw9aRLPPPccw4cP59ABAzjzzDPteRtCkEylPGvSokULbhk3josvvpiXXnuVaDhiW11NY87cuZw5eDAT7riDfn37ctmwYZwzZAiTp05h+rTpXHrZZVxx1ZUYhuGVFBuGQf/+/dmwcSMPPfywp/lNy+KU3/yGvgcdxP899hh7de3Ko7ffTnV1NRs3bkRoGjfefDOLFi1izM03E08kMJzR2k899RRlZWUkEgkOPfRQBg8ejM/nIx6PE4lGOP+88/j4o49IO+/KjW92VFi0H2IxCgNEIQSbqqtBCGrrYgwYMIB7J/6RyVOnomzaRC6XZdYnn1BbV4fq+LuxRJzeffpw3wP3M2XKFNavWw9CMH3Gh1RXV1NZUYHP72efnj154oknPBdLOjHN6lWruP+Pf0RRFHr16UMwFKKmpoYOHTrYaJOsr+eWyKL68oaBsyUlG9evx+fzEYvXceKJJ3L44YfT/9cDaNu6DfFEghkzZmBZFps2bkRKSXVdjFP7nca1o0bRq1cvwuEwhmEwbdo0b0JtTU0NgUCAQDhMbV0d0WiUP/zhD1RWVhJwpiYZpkl1TQ03jxlDWXk5w4cPZ87cuaz8/nt7DFwoRMvWrdH9fmbMnEnP3r254oorvFhPVVV8Ph8vvPACirr5PMNMOk11dbVHR6+trcWyLCKRCJqqUlZWRmlpKZlMhrlz53L++eez0bHuQgj6HXQQVRUVLFq0CCktDjzwQJYuXcrXX39NeUkpb735JmcPGYJKffxnOSOtL7/8cu665x6+/OpLvv3mWxRFYd/99mXmv2cyZswYysvL2bRpE9dffz3du3dn4OGHUxeLcffdd7NixQriiSRS2i2YAsEgTz/9NCtWrGD9+vVEo1HeeOMN4skEyWSSZlXNOODAA/l89mwSdXUojkv3s1iQhsJhGAYXXnABbdu1IxaL0b9/fwBGXH01NbW1zJs3j86dO3PuOecQCoXo0qULdXV17LfffgD8fvhwampr+O67RbRr147f/e4ynnzy7/x7xgxCoRCtWrXyXCzLstB0nUQ8Tl0iQVWzZjz51FM0a97Mrk5UFdauXoOu616jZG8CrZQU8j3d+CkUCjFkyBBKSkupra1l4MCBtGvXjhuuH02zZs1YMH8+5557LtOnTWOvzp1p27YtdfE4B/btS/PmzRk9ejShUIgVK1ZwzDHHUFlezmeffYZpmnTr1o1mzZtx8MEHew0IotEoPXv29DbqTTfdxJlnnslxxx7L0HPO4d1332XEiBEEAgGWLlnC768ezvWjruOQAf1ZvdLumO73+9E0jWQySTqd5ogjjmDvvffm/vvu8/pXucFzIBjkgAMOoFefPrRq04ZNmzYRdQLSBx54gEceeYTx48cz5JxzGD16NGNuvtkeZ6eqpNIpLrzoIiZPnoyu6YRCIbvjvbQHkaYzGfx+v93oTkp7kCtQ1bw5mzZt4o3XX+fmMWN49NFHCYfDHHnkkbz66quoqkrLli1ZsWIFe+21FyOvvZbJH3xA+w4dGDZsGC+99BK3T7idTnt2IpfNeiO8H374Yb766iuaNWtGeXk5Q88eQp8+ffj888+56qqraNOqFb6CMd+7hItlWRaDBg2iS7du1MViJBIJpk6eSs9evWzoVEoqm1V5aMuvDz2UmpoakokE06ZOY5999rE78QlBs6oqSqMlzJ79Od988y1z5swpKlV1TabP56Nb9+6sXbuWYCDA/z3yCPdNvI+qqiqy2SyrV65EmvWkuoY9qAohR0VRqKis9Bou/+tf/0LTNFq0aEFdXR2pVIrWrVsTCoUoLS2lrLycdDbLV199xdfz5lFRUUEikWD16tW0b9eOSCRCIpFASskNN9zAG6+9zoknnsg777zD7XdMoHvXbgwePJh8Pk/fvn3p3r07/fv3Z+F3i5g2fRpDhw4llUyycuVKUqkUzauaMWPGDFavXk3Lli2prq5m6NChdOvWjTFjxhAMBtm0aRNr1qzxLKW7TqtWrWLwkCG0aNGCKVOmMP622zjj9NMpKS1FAlePGMHs2bMZes45+Hw+3n77bT7++GNat25NMplk2bJlhEIhLr30UqZPn87y5cspr6iwvYVYjOZ77MHGTZswrPoZgplslnPPP5/999uPRCLBggULOOOMM1AUhZkzZ3LJJZdQUlLCzJkzueWWW7j55pt5/fXXGTN2DOWlZUx65x0GDBjA3//+d/bddz9UZ3bKl3PnMnToUM/1k5bFkqVLMfJ5fLpO82bN6sdQF7aM+qkFpLCnqqIo6LrOpZdeSjKVolevXvz60F/Ttk1b8vk8y5Yt48MPP2TRokVEnAq0eDzOPnvvzaEDB9KuXVu7K9+KFUyf/i8WLlyAruk8+eST9O/f35tp7g8EEOAhJn6/n48//phLL70UoSjkc3mSiQShYJCS0lKOOeYY2rdrh2EY9jgGitsHuQunqSqpVIqbb76ZaDTKZZddRteuXdF1HcMwCIVCPPvss8yYMYPWrVrx/vvvIxSFSy69hH377Ot9zjX5t40fT7OqKiorK7njzjspKS3l+BNO4KKLLuKll17iyiuuYPHixV6jg/fee48PP/yQSy65hJYtW5LP56mpqaF79x50aN+ezz77DJ/Px8KFC/H7/Rx55JFkMhlWrVrFhg0b6N2nN/vvfwBfzZ3LNddcQ+uWLb2EZzab5cQTT+Tee+9l1LXX8vbbbzPj3//mmWef5Z4//AHTsli3bh2ZdJqjjjqKUChEnz59kMC///1vWrVsyYABA5g6dSqqqlJTU8O0adOYMGECl112Ga+99hoXXnQR995zD0LU54Z8Ph+P/OUvJJNJDjjgAA455BCvPv2LL77gX//6F/PnzycUChEIBKhzumv6dR/BYJCSaJTPP/+c115/nbFjxxHw+0klk4wYPpwBAwZ4a24YBnXxOHvvvTcfzZxpN0N3J3z9nChWwwSVO5PvvPPO47bbb6e6ehOLFn2Hoiic9tvfcs21I7nm6hF88MEHKMCQwYO54647icfjLFywECEEp5xyCiOvuYaRI65h0qRJ3Dlhgh13CEEyHueP999POBTisksvJRKJoKgqNbW1+P1+8rkcV1x1JUOGnoOqqTbka5rMmzePXC7nNXl2M7SFeQkX1WrdujUvvvgi3333HU8//bQ3Eal79+488+yzXH/ddUyZPJmqykoee+wxTGnx8EN/9uKNjh078ueHHuJPXbvy54ceYvjw4XTo2IGzzjyT0pIS/v7kk/h0nUN//WsWzJ/vJe8URaGsrMzOw1RXe3BqLBajvLyctm3bcv/993u5ADdfUFNTY2vrTIaA30+rVq0I+v1erXo2m6VZVRXjx4/n3nvu4fXXXiMcDvPbU0/l1tvGs1fnzqSTSaKRCLlslutHjWLVurU8eN/9+P1+Lrt8GAcf2JennnmG4b//PWvXrKGyshIzn+eKYcMYM3YsZ511Fs8+8wwvv/yy1xLVNE00TWP9+vVceeWVnHLyybz15pv8Z8kSFEWhZcuWTPzjH3noT3/i1VdfJeT3c+eECdz/wAO8/MrLNG/enPnz5/PPf/6Tdm3aelPDLCmRikLX7t0Jh8PeddLpNH327UMmncFwkp4/FOh1k8k/Wh7ERUxuuPFGvpk3j8MOO8xuCmeaNG/enDlz53L1iBG8++676H4/140ezfffr+Tggw8mmUgAUFpWxmezZ3Ptddfxzj//ycKFCz1hjMfj1FZXk89mmTt3LuXl5XZCTQii0ShXX301kWgUy7RIp1N23uLbbzn/vPO4Ytgwp9KsPmHX0OzmDYNQIEDHTp149NFHmT59Oj6f3T1j0aJFjBw5klatW5PJZAj6/XTac0+ee/ZZ3n//fQ+dmTt3LldedRWtWrVC0zSeeeYZ/vrIXzEt02ui98STT5LJZOjRvTuhYBAhJT5NI5VM8sB993kvV9d1NtRUc9F55/O7YcPQFIVkIsHB/fox/OqruW7kSDtf43z+xeefty20A/267UBrYzFOOeUUYrEYFRUVgN1o7rxzzqVFixZUVFZ6ycyysjKSySQlJfbw0KBu53Ei4TDlZWXkczl0Xcfn8zF79mxOP/10IuEwGzZsoKysjHQyia6qhINBpMOG+O1pp7Fo0SJun3A7fp/f4fsJjjzqKI4//nheeOEFysrK+GbePM464wz6HdSPulgds2bNspkGpknA70cVAr+uk0kmGX/LLV62Xdd11ldvYvgVV3LJJZcgnZnq8keaqa7Vz6fecV6em2lOp9N89sknHDnoKMbcfDNLliwBYN/99qX5Hs15/ZVXEIBhmnzyySecfMrJ3HjTjSxauAiEoFevXrRr344nH38CyzQpLS/3rJOUkpKSEkpKSohEIl6TOQAjl2PS228XDXv0+/1UV1ejKArhkpLNhrc0pD34NI26mhouvegiRl57LYMHDyaVSKAoCnu0aMGLL7zAyy++SFVlJZlMhssuvZQbb7qJyVOmEKutRQhBsz2aM2PGDP780ENUVlaSTqcRqoJPU73sfGlJCX6fD8MwmD9/PvlczmvZ07x586Ist6qqaLrOqlWr7NjN52Pp0qV8/umnHHjAAR7txLIswuEwrVq1Ys6cOUWN66SUpNNpQgVzDX0+H5WVlWiaxvfLl5N35q+7uZPly5Z5AEcmk2HhwoWeW+u61dFoFMMwSKZSXgtRVVWpra3lP//5D5qmoes6lw8bxk033cT0f31I9caNADRrsQdr1qzhtvHjvfcYiUSoi9fx5htvemO3XXRx6ZIlJBJxD3SoqKiwn61gUpiUkuXLl9fnyn5Ih5tCECefy0tN17jnrj9w7z332KbeMLd7YpE7yy8ajXLVVVdx5KBBVFZWIi2LjRs38tqrr/K3xx7DcHxDXdf5/fDhHHvsMZSXlQN24+q33nyTRx55BMuZiehSQlKpFKNvuIFQMMi4ceMIBoNOHsOZP6IID+J0F7K2tpbTTjuN35x6KldcfjmZTMbuKbyVKUV1dXX4fD46depEKGS7aatWrWLt2rXerEAhBMlkEiklnTp1otTpk7tmzRpWr17tUTca4/cUoWeG6XCQ3ClbxS2AQKKqGkJVvPakLljhUnYKXcXC7wvx/4a9sgoVhM/n89wi93P5fL5oEJFlmo1ywBSnY4vrKtrXNz1iqhAK6XQa0zS8pB7Axo0bvf5ZfgdtUhw43iXIGh6nTHpER8UhhhZywtxch6apCASGaTpEzh2rLFRUew/8ceJEzj5nKCKfz0tN07j3D3dz9913U76DAuJusFwuRzKZJBIOE4lGkZZFPJEglUpRWlrq8Z/yhkEykSAcDhOORFAUhVhtLel0mmg0il4gHO4LdTlKXp8kZ0CnKWU9qbMBh0paFnnDIOD3FzF1GwqJR1Z0NGk6nfY+4/P5vBfpaifFoazYG8C+L7/f77llNMJCpQETVcr6UQmygDBczxJ2muS54+MKXnrh2lgFc//qKTSiKabfmxnYlM4j9Y2rC1i4sr79kvfcTuNuaUkPmUynUuQdrpemaZ6AF3Xvd8m5VkHPLymL2OJuvy1XIN1ntqR0SJ6iiLS6vQKiaiq1sRh/nDiRIecMRfPKDgtwc6nueJDjDwQIBoMYhkEimUQAPr+PkDuI0fGZAz4foaoqm4yXToOUBAMBe2BjAcO1iEjW0Hy645oVUcRqLiTwoSioum5v6oY9tbZA1HTJcg1h7MLN565bKBwuor9vqTa74b25G76QfVxUNuAqHc0ZVVfQxb4hqU5V1WLaeNPNPnpjf1tAW3fp/GzW5tMRaqcjpvccDloopURowmP+BsNhQq5icDc5Dcpl3aVRi9+TLRiyfiqXom62sqpQGk0A75CLpSjevXgCksmkSSST3lgDV2i2Jz6xGtHIADIjG/XvNvtsNkuhwDat6GXL97CrHz8k7mustkvuovdv/cRr0dSj4f6WTjlBIpEgl8/ZAqI4rRoHHX00lZWVHtLwSz4auodidz397mM7YulMLkvfvv3s/1vSzjEr7O5asvvYfRRZGCntPIhwKtwKg6xix7iJ9luwRT+wKX2Tfs7eSru2SdwOn+fnuL1d7b1t73oVVXI7ELmqIITC/wM6ubfu/ZhtSwAAAABJRU5ErkJggg=="

# STATION_INFO: 정류소 정보 캐시 딕셔너리 (현재는 빈 딕셔너리로 초기화만 됨).
#               향후 정류소 정보를 로컬에 캐싱할 때 사용하기 위해 예약.
STATION_INFO = {}

# ══════════════════════════════════════════════════════════
# 【2-3】 서울시 버스 공공데이터 API 주소(URL) 상수
#   공공데이터포털(data.go.kr) 서울시 버스 OpenAPI 엔드포인트.
#   fetch_api(url, params)를 통해 호출되며,
#   ServiceKey + 파라미터를 GET으로 전송하면 XML 형태로 응답.
# ══════════════════════════════════════════════════════════
# URL_POS1 (getBusPosByRtid):    노선 ID → 해당 노선 모든 버스 현재 위치 조회.
#                                 출발 감지(첫·두 번째 정류소 통과 판정)에 사용.
# URL_POS2 (getBusPosByRouteSt): 노선의 특정 구간(startOrd~endOrd) 버스 위치.
#                                 종점 도착 판정에 사용.
# URL_SLST (getStaionByRoute):   노선의 전체 정류소 목록 (ID·이름·순번·회차지 등).
#                                 지도 그리기 및 구간 속도 수집에 사용.
# URL_RINF (getRouteInfo):       노선 기본 정보 (운수사·첫차·막차·길이 등).
# URL_SRCH (getBusRouteList):    노선 번호 문자열로 노선 목록 검색.
#                                 노선 검색 대화상자에서 사용.
URL_POS1 = "http://ws.bus.go.kr/api/rest/buspos/getBusPosByRtid"
URL_POS2 = "http://ws.bus.go.kr/api/rest/buspos/getBusPosByRouteSt"
URL_SLST = "http://ws.bus.go.kr/api/rest/busRouteInfo/getStaionByRoute"
URL_RINF = "http://ws.bus.go.kr/api/rest/busRouteInfo/getRouteInfo"
URL_SRCH = "http://ws.bus.go.kr/api/rest/busRouteInfo/getBusRouteList"

# ══════════════════════════════════════════════════════════
# 【2-4】 노선 종류 이름표(LABEL) · 색상표(COLOR)
#   API routeType 코드(숫자 문자열)를 이름·색상으로 매핑.
# ══════════════════════════════════════════════════════════
# 코드 → 이름   → 지도 선 색상(대략)
#  "1" → 공항버스  → 황금색    "#C8A000"
#  "2" → 마을버스  → 연초록    "#6DBF67"
#  "3" → 간선버스  → 파란색    "#1E6FD9"  (서울 주요 간선)
#  "4" → 지선버스  → 짙은 초록 "#005A00"  (간선 보조)
#  "5" → 순환버스  → 황금색    "#E0B800"
#  "6" → 광역버스  → 짙은 빨강 "#5A0E11"  (수도권 광역)
#  "7" → 인천버스  → 청록색    "#20B2AA"
#  "8" → 경기버스  → 청록색    "#20B2AA"
#  "9" → 폐지노선  → 회색      "#888888"
#  "0" → 공용      → 어두운 회색 "#333333"
# DEFAULT_LINE_COLOR: 코드 없을 때 사용하는 기본 선 색상 "#333333"
ROUTE_TYPE_LABEL = {
    "1": "공항", "2": "마을", "3": "간선", "4": "지선",
    "5": "순환", "6": "광역", "7": "인천", "8": "경기",
    "9": "폐지", "0": "공용",
}
ROUTE_TYPE_COLOR = {
    "1": "#C8A000", "2": "#6DBF67", "3": "#1E6FD9", "4": "#005A00",
    "5": "#E0B800", "6": "#5A0E11", "7": "#20B2AA", "8": "#20B2AA",
    "9": "#888888", "0": "#333333",
}
DEFAULT_LINE_COLOR = "#333333"

# ══════════════════════════════════════════════════════════
# 【2-5】 노선 지도 그리기용 크기/위치 상수
#   QGraphicsScene에 정류소·선을 그릴 때 사용하는 기본 크기값.
#   창 크기에 따라 _calc_layout()에서 동적으로 조정된다.
# ══════════════════════════════════════════════════════════
# STOPS_PER_ROW  : 한 행에 배치할 정류소 수 기본값 (15개)
# CELL_W / CELL_H: 정류소 하나가 차지하는 셀의 가로(42)/세로(70) 픽셀
# PAD_X / PAD_Y  : 지도 좌우(55)/상단(110) 여백. PAD_Y는 정보 텍스트 여백 포함.
# CIRCLE_R       : 일반 정류소 원 반지름 (8픽셀)
# CIRCLE_R_SPECIAL: 출발·회차·종점 정류소 원 반지름 (13픽셀, 더 크게)
# TEXT_X_OFFSET  : 정류소 이름 텍스트 X축 미세 조정 (-1픽셀)
# FONT_CAP       : 셀 너비가 이 값(150) 이상이면 폰트 크기 최대(10pt) 고정
STOPS_PER_ROW = 15
CELL_W = 42
CELL_H = 70
PAD_X = 55
PAD_Y = 110
CIRCLE_R = 8
CIRCLE_R_SPECIAL = 13
TEXT_X_OFFSET = -1
FONT_CAP = 150

# ══════════════════════════════════════════════════════════
# 【3-1】 _make_palette(mode) → QPalette
#   라이트/다크 모드에 맞는 Qt 색상 팔레트를 만들어 반환.
#   QPalette는 창 배경·글자·버튼·강조색 등을 한꺼번에 설정하는 객체.
#   앱 전체에 app.setPalette(p)로 적용하면 모든 위젯의 색상이 바뀐다.
#
#   @param mode : "dark" → 다크 테마, 그 외 → 라이트 테마
#   @return     : 색상이 설정된 QPalette 객체
#
#   다크 모드 주요 색상:
#     Window(배경) = #353535, WindowText(글자) = #DCDCDC
#     Base(표/입력창) = #232323, Highlight(선택) = #2A82DA
#   라이트 모드 주요 색상:
#     Window(배경) = #F0F0F0, WindowText(글자) = #000000
#     Base(표/입력창) = #FFFFFF, Highlight(선택) = #0078D7
# ══════════════════════════════════════════════════════════
def _make_palette(mode):
    p = QPalette()
    if mode == "dark":
        p.setColor(QPalette.Window, QColor(53, 53, 53))
        p.setColor(QPalette.WindowText, QColor(220, 220, 220))
        p.setColor(QPalette.Base, QColor(35, 35, 35))
        p.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
        p.setColor(QPalette.Text, QColor(220, 220, 220))
        p.setColor(QPalette.Button, QColor(53, 53, 53))
        p.setColor(QPalette.ButtonText, QColor(220, 220, 220))
        p.setColor(QPalette.Highlight, QColor(42, 130, 218))
        p.setColor(QPalette.HighlightedText, QColor(255, 255, 255))
        p.setColor(QPalette.ToolTipBase, QColor(53, 53, 53))
        p.setColor(QPalette.ToolTipText, QColor(220, 220, 220))
        p.setColor(QPalette.Link, QColor(100, 180, 255))
        p.setColor(QPalette.Disabled, QPalette.ButtonText, QColor(90, 90, 90))
        p.setColor(QPalette.Disabled, QPalette.WindowText, QColor(90, 90, 90))
        p.setColor(QPalette.Disabled, QPalette.Text, QColor(90, 90, 90))
    else:
        p.setColor(QPalette.Window, QColor(240, 240, 240))
        p.setColor(QPalette.WindowText, QColor(0, 0, 0))
        p.setColor(QPalette.Base, QColor(255, 255, 255))
        p.setColor(QPalette.AlternateBase, QColor(245, 245, 245))
        p.setColor(QPalette.Text, QColor(0, 0, 0))
        p.setColor(QPalette.Button, QColor(240, 240, 240))
        p.setColor(QPalette.ButtonText, QColor(0, 0, 0))
        p.setColor(QPalette.Highlight, QColor(0, 120, 215))
        p.setColor(QPalette.HighlightedText, QColor(255, 255, 255))
        p.setColor(QPalette.ToolTipBase, QColor(255, 255, 220))
        p.setColor(QPalette.ToolTipText, QColor(0, 0, 0))
        p.setColor(QPalette.Link, QColor(30, 100, 220))
        p.setColor(QPalette.Disabled, QPalette.ButtonText, QColor(170, 170, 170))
        p.setColor(QPalette.Disabled, QPalette.WindowText, QColor(170, 170, 170))
        p.setColor(QPalette.Disabled, QPalette.Text, QColor(170, 170, 170))
    return p

# ══════════════════════════════════════════════════════════
# 【3-2】 darken_color(hex_color, factor=0.60) → str
#   헥사 색상 코드("#RRGGBB")를 받아 factor 비율만큼 밝기를 낮춰 반환.
#   노선 지도에서 회차 이후 구간 선을 원래 색보다 어둡게 표시할 때 사용.
#
#   처리 과정:
#   1) 헥사 코드 → RGB (0.0~1.0 범위)
#   2) RGB → HSV (색조 H, 채도 S, 명도 V) 변환
#   3) 명도 V에 factor 곱해 낮춤 (0.6이면 40% 어두워짐)
#   4) HSV → RGB → 헥사 코드로 재변환 후 반환
#
#   @param hex_color : 입력 색상 (예: "#1E6FD9")
#   @param factor    : 밝기 비율 (0.0=완전검정, 1.0=원본, 0.6=40% 어둡게)
#   @return          : 어두워진 헥사 색상 (예: "#123F7A")
# ══════════════════════════════════════════════════════════
def darken_color(hex_color, factor=0.60):
    hex_color = hex_color.lstrip("#")
    r, g, b = (int(hex_color[i:i+2], 16) / 255.0 for i in (0, 2, 4))
    h, s, v = colorsys.rgb_to_hsv(r, g, b)
    r2, g2, b2 = colorsys.hsv_to_rgb(h, s, max(0.0, v * factor))
    return "#{:02X}{:02X}{:02X}".format(int(r2 * 255), int(g2 * 255), int(b2 * 255))

# ══════════════════════════════════════════════════════════
# 【3-3】 truncate_name(name, max_len=8) → str
#   정류소 이름이 max_len을 초과하면 앞 7글자만 남기고 ".."을 붙여 반환.
#   지도 위 좁은 공간에 이름을 표시할 때 공간이 부족하기 때문.
#   예: "서울역버스환승센터" (9글자) → "서울역버스환승.."
# ══════════════════════════════════════════════════════════
def truncate_name(name, max_len=8):
    return name if len(name) <= max_len else name[:7] + ".."

# ══════════════════════════════════════════════════════════
# 【3-4】 fmt_bus_no(raw) → str
#   API에서 받은 버스 번호판 문자열을 보기 좋게 포맷.
#   전세버스 형식 "1234사5678" → "서울 1234 사 5678"
#   일반 형식    "1234"       → "서울 1234"
#   정규식 r'^(\d+)(사)(\d+)$'으로 전세버스 형식을 구분.
# ══════════════════════════════════════════════════════════
def fmt_bus_no(raw):
    m = re.match(r'^(\d+)(사)(\d+)$', raw.strip())
    if m:
        return f"서울 {m.group(1)} {m.group(2)} {m.group(3)}"
    return f"서울 {raw}"

# ══════════════════════════════════════════════════════════
# 【3-5】 format_remain_time(seconds) → str
#   남은 시간(초)을 사람이 읽기 좋은 문자열로 변환.
#   3661 → "1시간 1분 1초",  125 → "2분 5초",  45 → "45초"
# ══════════════════════════════════════════════════════════
def format_remain_time(seconds):
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    if h > 0:
        return f"{h}시간 {m}분 {s}초"
    elif m > 0:
        return f"{m}분 {s}초"
    return f"{s}초"

# ══════════════════════════════════════════════════════════
# 【3-6】 format_hhmm(raw) → str
#   API 반환 날짜+시각 문자열에서 "HH:MM" 부분만 추출.
#   "20240101042500" (14자) → "04:25"  (8~10번째=시, 10~12번째=분)
#   "0425"           (4자)  → "04:25"  (앞 2자=시, 뒤 2자=분)
# ══════════════════════════════════════════════════════════
def format_hhmm(raw):
    if not raw or len(raw) < 4:
        return raw
    return f"{raw[8:10]}:{raw[10:12]}" if len(raw) >= 14 else f"{raw[:2]}:{raw[2:4]}"

# ══════════════════════════════════════════════════════════
# 【3-7】 format_datetm(raw) → str
#   14자리·12자리 숫자 형식 날짜+시각 문자열을 "YYYY-MM-DD HH:MM:SS"로 변환.
#   raw가 None이거나 파싱 실패 시 → 현재 시각을 포맷해서 반환.
#   예: "20240101042500" → "2024-01-01 04:25:00"
# ══════════════════════════════════════════════════════════
def format_datetm(raw):
    if not raw:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    raw = str(raw).strip()
    try:
        if len(raw) == 14:
            return f"{raw[0:4]}-{raw[4:6]}-{raw[6:8]} {raw[8:10]}:{raw[10:12]}:{raw[12:14]}"
        if len(raw) == 12:
            return f"{raw[0:4]}-{raw[4:6]}-{raw[6:8]} {raw[8:10]}:{raw[10:12]}:00"
    except:
        pass
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# ══════════════════════════════════════════════════════════
# 【3-8】 load_pixmap_from_b64(b64_str) → QPixmap
#   Base64 인코딩된 이미지 문자열을 QPixmap 이미지 객체로 변환.
#   실패 시 빈(null) QPixmap 반환.
#   처리: base64.b64decode() → 이진 데이터 → QPixmap.loadFromData()
# ══════════════════════════════════════════════════════════
def load_pixmap_from_b64(b64_str):
    try:
        raw = base64.b64decode(b64_str)
        pm = QPixmap()
        pm.loadFromData(raw)
        return pm
    except:
        return QPixmap()

# ══════════════════════════════════════════════════════════
# 【3-9】 make_bus_pixmap() → QPixmap
#   지도에 버스 위치 표시용 25×25픽셀 아이콘을 만들어 반환.
#   ICON_B64 복원 실패 시 파란 사각형(#1E6FD9)으로 대체.
#   Qt.KeepAspectRatio + Qt.SmoothTransformation으로 비율 유지하며 축소.
# ══════════════════════════════════════════════════════════
def make_bus_pixmap():
    pm = load_pixmap_from_b64(ICON_B64)
    if pm.isNull():
        pm = QPixmap(25, 25)
        pm.fill(QColor("#1E6FD9"))
    else:
        pm = pm.scaled(25, 25, Qt.KeepAspectRatio, Qt.SmoothTransformation)
    return pm

# ══════════════════════════════════════════════════════════
# 【3-10~13】 현재 테마 색상 getter 함수들
#   테마가 바뀌어도 항상 현재 팔레트의 올바른 색상을 반환.
# ══════════════════════════════════════════════════════════
# get_app_bg_color()   → QPalette.Window      (창 배경색)
# get_text_color()     → QPalette.WindowText  (기본 글자색)
# get_base_color()     → QPalette.Base        (표·입력창 배경색)
# get_header_bg_color()→ QPalette.Button      (표 헤더 배경색)
def get_app_bg_color():
    return QApplication.palette().color(QPalette.Window)

def get_text_color():
    return QApplication.palette().color(QPalette.WindowText)

def get_base_color():
    return QApplication.palette().color(QPalette.Base)

def get_header_bg_color():
    return QApplication.palette().color(QPalette.Button)

# ══════════════════════════════════════════════════════════
# 【3-14】 detect_os_dark_mode() → bool
#   OS의 다크 모드 설정을 감지. 프로그램 시작 시 테마 자동 선택에 사용.
#   Windows: 레지스트리 "AppsUseLightTheme" 값 확인 (0=다크)
#   macOS  : "defaults read -g AppleInterfaceStyle" 명령 실행 (결과 "Dark"=다크)
#   기타 OS: 항상 False(라이트 모드) 반환
# ══════════════════════════════════════════════════════════
def detect_os_dark_mode():
    try:
        if sys.platform == "win32":
            import winreg
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                r"Software\Microsoft\Windows\CurrentVersion\Themes\Personalize")
            val, _ = winreg.QueryValueEx(key, "AppsUseLightTheme")
            winreg.CloseKey(key)
            return val == 0
        elif sys.platform == "darwin":
            import subprocess
            result = subprocess.run(
                ["defaults", "read", "-g", "AppleInterfaceStyle"],
                capture_output=True, text=True, timeout=3
            )
            return result.stdout.strip().lower() == "dark"
    except:
        pass
    return False

# ══════════════════════════════════════════════════════════
# 【4】 RouteMapPanel 클래스
#   버스 노선의 정류소를 그래픽으로 표시하고
#   운행 중인 버스 아이콘을 실시간 갱신하는 패널 위젯.
#
#   화면 구조:
#   ┌─────────────────────────────────────────────┐
#   │ [간선] 101 | 총 35개 정류소 | 약 35.2km ...  │ ← 노선 정보 텍스트
#   │  ①─②─③─ … ─⑮                             │
#   │  ↓                                          │ ← QGraphicsScene
#   │  ㉚─㉙─ … ─⑯  (뱀 형태 지그재그 배치)       │   (정류소 + 버스 아이콘)
#   ├─────────────────────────────────────────────┤
#   │ 종점도착예정: [1234] 5분20초 | [5678] 8분45초 │ ← _table (도착 예정 표)
#   └─────────────────────────────────────────────┘
#
#   @param route_rnm: 이 패널이 담당하는 노선 번호/이름 (예: "101")
# ══════════════════════════════════════════════════════════
class RouteMapPanel(QWidget):
# ──────────────────────────────────────────────────────────
# 【4-1】 __init__(route_rnm, parent=None)
#   패널 위젯을 초기화. 인스턴스 변수 선언 및 화면 요소 생성.
#
#   주요 인스턴스 변수:
#   _sect_speeds   : {seq번호: 속도(km/h)} 구간별 속도 딕셔너리
#   _stations      : [{station, seq, name, transYn, ...}] 정류소 정보 리스트
#   _current_rtype : 현재 노선 종류 코드 (기본값 "3" = 간선)
#   _current_length: 노선 총 길이 문자열
#   _last_buses    : 최근 그린 버스 목록 (창 크기 변경 시 재그리기용)
#   _bus_seconds   : {b_idx: 남은초} 버스별 종점까지 남은 시간
#   _table_bus_info: [{col, b_idx}] 도착 예정 표의 버스 정보
#
#   주요 화면 요소:
#   _scene     : 정류소·버스를 그리는 QGraphicsScene (무대)
#   _view      : _scene을 화면에 표시하는 QGraphicsView (스크롤 가능)
#   _table     : 버스 도착 예정 시간 표 (2행 × N열, 최대높이 65px)
#   _tip_label : 마우스 오버 시 표시되는 툴팁 QLabel (ToolTip 플래그)
#   _tick_timer: 1000ms 간격으로 _tick_countdown을 호출하는 타이머
#   _resize_timer: 창 크기 변경 후 300ms 뒤 지도 재그리기 (단발성 타이머)
# ──────────────────────────────────────────────────────────
    def __init__(self, route_rnm, parent=None):
        super().__init__(parent)
        self.route_rnm = route_rnm
        self._sect_speeds = {} 
        self._stations = []
        self._current_rtype = "3"
        self._current_length = ""
        self._last_buses = None
        self._bus_seconds = {}
        self._table_bus_info = []
        self._min_map_width = 600

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        self._scene = QGraphicsScene(self)
        self._view = QGraphicsView(self._scene, self)
        self._view.setRenderHint(QPainter.Antialiasing)
        self._view.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self._view.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self._view.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self._view.setMouseTracking(True)
        self._view.viewport().setMouseTracking(True)
        self._update_bg_color()
        layout.addWidget(self._view, stretch=1)

        self._table = QTableWidget(2, 1)
        self._table.setMaximumHeight(65)
        self._table.setMinimumHeight(65)
        self._table.verticalHeader().setVisible(False)
        self._table.horizontalHeader().setVisible(False)
        self._table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._table.setSelectionMode(QAbstractItemView.NoSelection)
        self._table.setShowGrid(True)
        layout.addWidget(self._table)

        t = self._scene.addText("노선 데이터 로딩 중...", QFont(FONT_FAMILY, 12))
        t.setDefaultTextColor(QColor("#777777"))
        t.setPos(200, 100)

        self._tip_label = QLabel(self)
        self._tip_label.setWindowFlags(Qt.ToolTip)
        self._tip_label.setFont(QFont(FONT_FAMILY, 9))
        self._update_tip_style()
        self._tip_label.hide()
        self._view.viewport().installEventFilter(self)
        self._view.setMinimumWidth(self._min_map_width)

        self._tick_timer = QTimer(self)
        self._tick_timer.setInterval(1000)
        self._tick_timer.timeout.connect(self._tick_countdown)
        self._tick_timer.start()

        self._resize_timer = QTimer(self)
        self._resize_timer.setSingleShot(True)
        self._resize_timer.setInterval(300)
        self._resize_timer.timeout.connect(self._on_resize_done)
        
# ──────────────────────────────────────────────────────────
# 【4-2】 set_sect_speeds(speeds)
#   구간별 속도 데이터를 저장. _speed_color()에서 선 색상 결정 시 참조.
#   @param speeds: {seq번호(int): 속도(float km/h)} dict, None이면 빈 dict로 초기화
# ──────────────────────────────────────────────────────────
    def set_sect_speeds(self, speeds):
        self._sect_speeds = speeds if speeds else {}

# ──────────────────────────────────────────────────────────
# 【4-3】 eventFilter(obj, event) → bool
#   지도 뷰포트에서 발생하는 마우스 이벤트를 가로채어 툴팁을 제어.
#   __init__에서 installEventFilter(self)로 등록되어 자동 호출됨.
#
#   MouseMove: 마우스가 올라간 아이템의 data(0) 값(툴팁 문자열)을 읽어
#              _tip_label을 마우스 근처에 표시. 아이템 없으면 숨김.
#   Leave    : 마우스가 지도 영역을 벗어나면 툴팁 숨김.
#   @return  : False (이벤트를 소비하지 않고 원래 처리도 계속 허용)
# ──────────────────────────────────────────────────────────
    def eventFilter(self, obj, event):
        if obj is self._view.viewport():
            if event.type() == event.Type.MouseMove:
                pos = self._view.mapToScene(event.position().toPoint())
                item = self._scene.itemAt(pos, self._view.transform())
                tip = item.data(0) if item else None
                if tip:
                    self._tip_label.setText(tip)
                    self._tip_label.adjustSize()
                    gp = self._view.viewport().mapToGlobal(event.position().toPoint())
                    self._tip_label.move(gp.x() + 12, gp.y() + 12)
                    if not self._tip_label.isVisible():
                        self._tip_label.show()
                else:
                    self._tip_label.hide()
                return False
            elif event.type() == event.Type.Leave:
                self._tip_label.hide()
                return False
        return super().eventFilter(obj, event)

# ──────────────────────────────────────────────────────────
# 【4-4】 resizeEvent(event)
#   창 크기 변경 시 Qt가 자동 호출. 성능을 위해 즉시 재그리기 대신
#   _resize_timer를 300ms로 (재)시작하여 변경이 끝난 후 한 번만 재그림.
# ──────────────────────────────────────────────────────────
    def resizeEvent(self, event):
        super().resizeEvent(event)
        if self._stations:
            self._resize_timer.start()

# ──────────────────────────────────────────────────────────
# 【4-5】 _on_resize_done()
#   _resize_timer 만료(300ms 경과) 후 호출. 정류소 데이터가 있을 때만
#   _last_buses와 함께 _draw()를 다시 호출해 지도 재그리기.
# ──────────────────────────────────────────────────────────
    def _on_resize_done(self):
        if self._stations:
            self._draw(self._last_buses)

# ──────────────────────────────────────────────────────────
# 【4-6】 _calc_layout(n) → (spr, cw, ch, font_size)
#   현재 뷰포트 크기와 정류소 수를 기반으로 레이아웃 값 계산.
#
#   계산 흐름:
#   1) usable_w/h = 뷰포트 크기 - 패딩
#   2) spr(한 행 정류소 수) = usable_w / min_cw, 범위 5~15
#   3) cw(셀 너비) = usable_w / spr
#   4) rows = ceil(n / spr),  ch = usable_h / rows, 최소 CELL_H 보장
#   5) font_size: cw에 비례하여 6pt~10pt 선형 보간
#
#   @param n   : 정류소 총 수
#   @return    : (한행정류소수, 셀너비, 셀높이, 폰트크기) 튜플
# ──────────────────────────────────────────────────────────
    def _calc_layout(self, n):
        view_w = max(self._view.viewport().width(), self._min_map_width)
        view_h = self._view.viewport().height()
        usable_w = view_w - PAD_X * 2 - 20
        usable_h = view_h - PAD_Y - 60
        min_cw = CELL_W
        min_ch = CELL_H
        min_fs = 6
        max_fs = 10

        if n <= 0 or usable_w <= 0 or usable_h <= 0:
            return STOPS_PER_ROW, CELL_W, CELL_H, min_fs

        spr = max(5, int(usable_w // min_cw))
        spr = min(spr, 15)
        if spr > n:
            spr = n

        cw = usable_w / spr

        rows = (n + spr - 1) // spr
        if rows > 0 and usable_h > 0:
            ch = usable_h / rows
            ch = min(ch, cw)
            ch = max(min_ch, ch)
        else:
            ch = min_ch

        if cw <= min_cw:
            font_size = min_fs
        elif cw >= FONT_CAP:
            font_size = max_fs
        else:
            ratio = (cw - min_cw) / (FONT_CAP - min_cw)
            font_size = int(min_fs + ratio * (max_fs - min_fs))

        return spr, cw, ch, font_size

# ──────────────────────────────────────────────────────────
# 【4-7】 _update_bg_color()
#   지도 뷰(_view)의 배경색을 현재 앱 테마의 창 배경색으로 갱신.
#   테마 변경 시 refresh_theme()에서 호출.
# ──────────────────────────────────────────────────────────
    def _update_bg_color(self):
        self._view.setBackgroundBrush(QBrush(get_app_bg_color()))
        
# ──────────────────────────────────────────────────────────
# 【4-8】 _update_tip_style()
#   툴팁 라벨(_tip_label)의 CSS 스타일을 현재 테마에 맞게 갱신.
#   다크: 어두운 배경(#1E1E1E) + 밝은 글자(#DCDCDC)
#   라이트: 밝은 배경(#FDFDFD) + 검정 글자(#000000)
# ──────────────────────────────────────────────────────────
    def _update_tip_style(self):
        if get_app_bg_color().lightness() < 128:
            self._tip_label.setStyleSheet(
                "QLabel { background-color: #1E1E1E; color: #DCDCDC; "
                "border: 1px solid #555; padding: 4px; }")
        else:
            self._tip_label.setStyleSheet(
                "QLabel { background-color: #FDFDFD; color: #000000; "
                "border: 1px solid #AAAAAA; padding: 4px; }")

# ──────────────────────────────────────────────────────────
# 【4-9】 refresh_theme()
#   테마 변경 시 이 패널 전체를 새 테마로 갱신.
#   _apply_theme()에서 모든 RouteMapPanel에 대해 호출됨.
#   처리: 배경색 → 툴팁 스타일 → 지도 재그리기 → 표 색상 재적용
# ──────────────────────────────────────────────────────────
    def refresh_theme(self):
        self._update_bg_color()
        self._update_tip_style()
        if self._stations:
            self._draw(self._last_buses)
        self._rebuild_table_colors()

# ──────────────────────────────────────────────────────────
# 【4-10】 _rebuild_table_colors()
#   도착 예정 표(_table)의 모든 셀 배경색·글자색을 현재 테마로 재적용.
#   0번 열(라벨) 또는 0번 행(헤더)은 header_bg, 나머지는 base_bg 적용.
# ──────────────────────────────────────────────────────────
    def _rebuild_table_colors(self):
        header_bg = get_header_bg_color()
        base_bg = get_base_color()
        txt_color = get_text_color()
        for row in range(self._table.rowCount()):
            for col in range(self._table.columnCount()):
                item = self._table.item(row, col)
                if item:
                    item.setForeground(QBrush(txt_color))
                    if col == 0 or row == 0:
                        item.setBackground(QBrush(header_bg))
                    else:
                        item.setBackground(QBrush(base_bg))

# ──────────────────────────────────────────────────────────
# 【4-11】 load_route(stations, rtype="3", length="")
#   새 노선의 정류소 목록을 받아 저장하고 지도를 처음 그림 (buses=None).
#   _slot_route_loaded()에서 패널 생성 후 바로 호출됨.
#
#   @param stations: [{station, seq, name, arsId, transYn, fullSectDist}, ...] 리스트
#   @param rtype   : 노선 종류 코드
#   @param length  : 노선 총 길이 문자열
# ──────────────────────────────────────────────────────────
    def load_route(self, stations, rtype="3", length=""):
        self._stations = stations
        self._current_rtype = rtype
        self._current_length = length
        self._draw(None)

# ──────────────────────────────────────────────────────────
# 【4-12】 update_buses(buses)
#   현재 운행 중인 버스 목록을 받아 지도를 갱신.
#   _slot_update_map 시그널 슬롯에서 호출됨.
#   정류소 데이터가 없으면 즉시 반환.
#
#   @param buses: [{vehId, plainNo, busType, lastStnId, sectDist, ...}, ...] 리스트
# ──────────────────────────────────────────────────────────
    def update_buses(self, buses):
        if not self._stations:
            return
        self._last_buses = buses
        self._draw(buses)
        
# ──────────────────────────────────────────────────────────
# 【4-13】 _speed_color(seq_idx) → QColor or None
#   구간 순번에 해당하는 속도를 조회하여 색상 반환.
#     0~9 km/h   → 빨강  QColor(255, 0, 0)   (정체/서행)
#    10~19 km/h  → 노랑  QColor(255, 255, 0)
#    20 km/h 이상 → 초록 QColor(0, 255, 0)   (원활)
#    데이터 없음   → None (기본 노선 색 사용)
# ──────────────────────────────────────────────────────────
    def _speed_color(self, seq_idx):
        spd = self._sect_speeds.get(seq_idx, -1)
        if spd < 0:
            return None
        if spd < 10:
            return QColor(255, 0, 0)
        elif spd < 20:
            return QColor(255, 255, 0)
        else:
            return QColor(0, 255, 0)

# ──────────────────────────────────────────────────────────
# 【4-14】 _draw(buses)  ★ 지도 그리기 핵심 함수 ★
#   QGraphicsScene을 clear하고 모든 요소를 처음부터 재그림.
#
#   그리는 순서:
#   ① scene 초기화 & 변수 준비
#   ② 상단 노선 정보 텍스트 (노선 종류·정류소 수·길이·운행 버스 수)
#   ③ 뱀 형태(지그재그) 좌표 계산 (coords 배열)
#      → 짝수 행: 왼→오른쪽,  홀수 행: 오른→왼쪽
#   ④ 정류소 간 연결선 (구간 속도 색 또는 노선 색, 회차 이후는 어두운 색)
#   ⑤ 정류소 원 + 번호/이름 텍스트
#      출발·회차·종점: 큰 원(CIRCLE_R_SPECIAL) + 굵은 레이블
#      일반 정류소  : 작은 원(CIRCLE_R) + 순번 숫자
#   ⑥ 버스 아이콘 (buses가 있을 때)
#      - lastStnId + sectDist 비율로 정확한 위치 보간
#      - 같은 Y대에 겹치는 레이블은 X 오프셋으로 분리
#      - 저상 버스 "저상" 텍스트, 막차 "[막차]" 텍스트 표시
#   ⑦ 행 연결 화살표(↓) 그리기
#   ⑧ sceneRect 설정 + _build_table() 호출
#
#   @param buses: 버스 정보 리스트 (None이면 버스 아이콘 없이 그림)
# ──────────────────────────────────────────────────────────
    def _draw(self, buses):
        self._scene.clear()
        self._bus_seconds.clear()
        stations = self._stations
        n = len(stations)

        theme_text_color = get_text_color()
        is_dark = get_app_bg_color().lightness() < 128

        line_color_hex = ROUTE_TYPE_COLOR.get(self._current_rtype, DEFAULT_LINE_COLOR)
        line_dark_hex = darken_color(line_color_hex, 0.58)
        line_color = QColor(line_color_hex)
        line_dark = QColor(line_dark_hex)

        turn_idx = next((i for i, st in enumerate(stations) if st.get("transYn") == "Y"), None)

        stops_per_row, cell_w, cell_h, dyn_font = self._calc_layout(n)

        typ_str = ROUTE_TYPE_LABEL.get(self._current_rtype, "기타")
        dist_str = self._current_length if self._current_length else "?"
        bus_cnt = len(buses) if buses else 0
        info_str = (f"[{typ_str}] {self.route_rnm} | "
                    f"총 {n}개 정류소 | 노선 총 길이 약 {dist_str}km | 현재 운행중인 버스 {bus_cnt}대")
        total_w = PAD_X * 2 + (stops_per_row - 1) * cell_w + 20

        info_item = self._scene.addText(info_str, QFont(FONT_FAMILY, 11, QFont.Bold))
        info_item.setDefaultTextColor(theme_text_color)
        info_item.setPos((total_w - info_item.boundingRect().width()) / 2, 10)

        if n == 0:
            t = self._scene.addText("정류소 정보가 없습니다.", QFont(FONT_FAMILY, 11))
            t.setDefaultTextColor(theme_text_color)
            t.setPos(total_w / 2 - 80, 110)
            self._scene.setSceneRect(0, 0, total_w, 200)
            self._build_table([])
            return

        rows = (n + stops_per_row - 1) // stops_per_row
        coords = []
        for i in range(n):
            ri = i // stops_per_row
            ci = i % stops_per_row
            x = PAD_X + ci * cell_w if ri % 2 == 0 else PAD_X + (stops_per_row - 1 - ci) * cell_w
            coords.append((x, PAD_Y + ri * cell_h))

        for i in range(n - 1):
            x1, y1 = coords[i]
            x2, y2 = coords[i + 1]
            spd_color = self._speed_color(i + 2)
            if spd_color is not None:
                sc = spd_color
            else:
                sc = line_dark if (turn_idx is not None and i >= turn_idx) else line_color
            pen = QPen(sc, 3)
            pen.setCapStyle(Qt.RoundCap)
            if (i // stops_per_row) == ((i + 1) // stops_per_row):
                self._scene.addLine(x1, y1, x2, y2, pen)
            else:
                self._scene.addLine(x1, y1, x1, y2, pen)

        for i, (st, (x, y)) in enumerate(zip(stations, coords)):
            is_first = (i == 0)
            is_last = (i == n - 1)
            is_turn = (st.get("transYn") == "Y")
            after_turn = (turn_idx is not None and i > turn_idx)
            nd = truncate_name(st.get("name", "?"))

            if is_turn:
                cr, cf, co = CIRCLE_R_SPECIAL, QColor("white"), line_color
                li, tc, fs, fb = "회차", line_color, dyn_font + 1, True
            elif is_first:
                cr, cf, co = CIRCLE_R_SPECIAL, QColor("white"), line_color
                li, tc, fs, fb = "출발", line_color, dyn_font + 1, True
            elif is_last:
                cr, cf, co = CIRCLE_R_SPECIAL, QColor("white"), line_color
                li, tc, fs, fb = "종점", line_color, dyn_font + 1, True
            else:
                sc2 = line_dark if after_turn else line_color
                if is_dark:
                    cr, cf, co = CIRCLE_R, QColor(60, 60, 60), sc2
                    li, tc, fs, fb = str(i + 1), QColor(200, 200, 200), dyn_font, False
                else:
                    cr, cf, co = CIRCLE_R, QColor("white"), sc2
                    li, tc, fs, fb = str(i + 1), theme_text_color, dyn_font, False

            self._scene.addEllipse(x - cr, y - cr, cr * 2, cr * 2, QPen(co, 2), QBrush(cf))
            fi = QFont(FONT_FAMILY, fs, QFont.Bold if fb else QFont.Normal)
            it = self._scene.addText(li, fi)
            it.setDefaultTextColor(tc)
            br = it.boundingRect()
            it.setPos(x - br.width() / 2, y - br.height() / 2)

            if is_first or is_last or is_turn:
                nc = tc
            else:
                nc = QColor(140, 140, 140) if is_dark else QColor(130, 130, 130)
            nt = self._scene.addText(nd, QFont(FONT_FAMILY, fs, QFont.Bold if fb else QFont.Normal))
            nt.setDefaultTextColor(nc)
            nt.setRotation(-45)
            nt.setPos(x + TEXT_X_OFFSET, y - cr - 2)

        table_buses = []
        if buses:
            bus_pm = make_bus_pixmap()
            bus_draw_list = []
            for b_idx, bus in enumerate(buses):
                lsi = bus.get("lastStnId", "")
                rpn = bus.get("plainNo", "").replace("서울", "").strip()
                ilb = (bus.get("islastyn") == "1")
                try:
                    sm = float(bus.get("sectDist", "0")) * 1000.0
                except:
                    sm = 0.0
                fi2 = next((i for i, st in enumerate(stations) if st.get("station", "") == lsi), None)
                if fi2 is None:
                    continue
                ti = fi2 + 1 if fi2 + 1 < n else fi2
                try:
                    sdm = float(stations[ti].get("fullSectDist", "0"))
                except:
                    sdm = 0.0
                ratio = min(sm / sdm, 1.0) if sdm > 0 and ti != fi2 else 0.0
                x1, y1 = coords[fi2]
                x2, y2 = coords[ti]
                if (fi2 // stops_per_row) == (ti // stops_per_row):
                    bx, by = x1 + (x2 - x1) * ratio, y1
                else:
                    bx, by = x1, y1 + (y2 - y1) * ratio
                try:
                    secs = int(bus.get("lastStTm", 0))
                except:
                    secs = 0
                bus_draw_list.append({
                    "b_idx": b_idx, "bus": bus, "bx": bx, "by": by,
                    "rpn": rpn, "ilb": ilb, "secs": secs,
                    "fi2": fi2, "ratio": ratio
                })

            Y_THRESH = 12 + dyn_font
            SEP = 15 + (dyn_font - 6) * 3
            groups = defaultdict(list)
            for bd in bus_draw_list:
                gk = round(bd["by"] / Y_THRESH) * Y_THRESH
                groups[gk].append(bd)
            label_offsets = {}
            for gk in sorted(groups.keys()):
                grp = sorted(groups[gk], key=lambda b: b["bx"])
                prev_label_x = None
                for bd in grp:
                    offset_x = 0
                    if prev_label_x is not None:
                        dist = bd["bx"] - prev_label_x
                        if dist < SEP:
                            offset_x = SEP - dist
                    label_offsets[id(bd)] = offset_x
                    prev_label_x = bd["bx"] + offset_x

            for bd in bus_draw_list:
                bus = bd["bus"]
                bx = bd["bx"]
                by = bd["by"]
                rpn = bd["rpn"]
                ilb = bd["ilb"]
                secs = bd["secs"]
                b_idx = bd["b_idx"]
                fi2 = bd["fi2"]
                ratio = bd["ratio"]
                offset_x = label_offsets.get(id(bd), 0)

                cc = bus.get("congestion", "0")
                _CONG = {"3": "여유", "4": "보통", "5": "혼잡"}
                cs = str(cc)
                if self._current_rtype == "6":
                    if cs in ("99", "0"):
                        cl = ""
                    else:
                        try:
                            cl = f"잔여{int(cs)}석"
                        except:
                            cl = ""
                else:
                    cl = _CONG.get(cs, "")
                is_low = bus.get("busType") == "1"
                parts = fmt_bus_no(rpn)
                if cl:
                    parts += f"  {cl}"
                if is_low:
                    parts += "  ( 저상 )"
                if ilb:
                    parts += "  [ 막차 ]"
                tip = f"{parts}\n종점도착까지 남은 시간 : {format_remain_time(secs)}"

                pi = self._scene.addPixmap(bus_pm)
                pi.setPos(bx - 12, by - 12)
                pi.setZValue(5)
                pi.setData(0, tip)

                hit = self._scene.addRect(bx - 15, by - 15, 30, 30,
                                        QPen(Qt.NoPen), QBrush(QColor(0, 0, 0, 0)))
                hit.setData(0, tip)
                hit.setZValue(8)

                bt = self._scene.addText(rpn, QFont(FONT_FAMILY, dyn_font + 3, QFont.Bold))
                bt.setDefaultTextColor(theme_text_color)
                bt.setRotation(-45)
                bt.setPos(bx - 2 + offset_x, by - 18)
                bt.setZValue(6)
                bt.setData(0, tip)

                if ilb:
                    lt = self._scene.addText("[막차]", QFont(FONT_FAMILY, dyn_font + 2, QFont.Bold))
                    lt.setDefaultTextColor(theme_text_color)
                    lt.setPos(bx - 18 + offset_x, by + 7)
                    lt.setZValue(6)

                if is_low:
                    lw = self._scene.addText("저상", QFont(FONT_FAMILY, 6, QFont.Bold))
                    lw.setDefaultTextColor(QColor(0, 0, 0))
                    lw.setPos(bx - 12, by - 14)
                    lw.setZValue(7)

                table_buses.append({
                    "raw_plain_no": rpn, "is_last_bus": ilb,
                    "seconds": secs, "from_idx": fi2, "ratio": ratio, "b_idx": b_idx
                })
                self._bus_seconds[b_idx] = secs

        for ri in range(rows - 1):
            ry = PAD_Y + ri * cell_h
            ny = PAD_Y + (ri + 1) * cell_h
            ax = (PAD_X + (stops_per_row - 1) * cell_w - 3) if ri % 2 == 0 else (PAD_X - 22)
            lir = min((ri + 1) * stops_per_row - 1, n - 1)
            spd_color = self._speed_color(lir + 2)
            if spd_color is not None:
                ac = spd_color
            else:
                ac = line_dark if (turn_idx is not None and lir >= turn_idx) else line_color
            at = self._scene.addText("↓", QFont(FONT_FAMILY, 11, QFont.Bold))
            at.setDefaultTextColor(ac)
            at.setPos(ax, (ry + ny) / 2 - 10)
            at.setZValue(-1)

        self._scene.setSceneRect(0, 0, total_w, PAD_Y + (rows - 1) * cell_h + 60)
        self._view.resetTransform()
        self._view.ensureVisible(0, 0, 10, 10)
        self._build_table(table_buses)

# ──────────────────────────────────────────────────────────
# 【4-15】 _build_table(table_buses)
#   지도 하단의 "종점 도착 예정 버스" 표(2행 × N열)를 구성.
#
#   표 구조:
#   ┌──────────────────────┬─────────┬─────────┬──────────┐
#   │ 종점 도착 예정 버스번호│  1234   │  5678   │ 9012(막차)│
#   │ 종점 도착까지 남은 시간│ 5분 20초│ 8분 45초│  2분 10초 │
#   └──────────────────────┴─────────┴─────────┴──────────┘
#
#   정렬: 종점에 가장 가까운 버스(from_idx 높은 순)가 왼쪽
#   막차 버스는 항상 맨 오른쪽
#   표 너비에 따라 보여줄 버스 열 수 동적 결정
#
#   @param table_buses: [{raw_plain_no, is_last_bus, seconds, from_idx, ratio, b_idx}] 리스트
# ──────────────────────────────────────────────────────────
    def _build_table(self, table_buses):
        self._table_bus_info = []
        header_bg = get_header_bg_color()
        base_bg = get_base_color()
        txt_color = get_text_color()

        if not table_buses:
            self._table.clear()
            self._table.setColumnCount(1)
            self._table.setRowCount(1)
            it = QTableWidgetItem("")
            it.setBackground(QBrush(base_bg))
            it.setForeground(QBrush(txt_color))
            self._table.setItem(0, 0, it)
            return

        sb = sorted(table_buses, key=lambda x: (x['from_idx'], x['ratio']), reverse=True)
        lb = next((b for b in sb if b['is_last_bus']), None)
        normal = [b for b in sb if not b['is_last_bus']]

        label_col_w = 160
        bus_col_w = 100
        table_w = self._table.width()
        if table_w < 200:
            table_w = 700
        available_w = table_w - label_col_w
        if available_w < bus_col_w:
            available_w = bus_col_w
        max_bus_cols = max(1, available_w // bus_col_w)

        if lb:
            cols = normal[:max_bus_cols - 1] + [lb]
        else:
            cols = normal[:max_bus_cols]

        if not cols:
            self._table.clear()
            self._table.setColumnCount(1)
            self._table.setRowCount(1)
            return

        cc = 1 + len(cols)
        self._table.setColumnCount(cc)
        self._table.setRowCount(2)
        hf = QFont(FONT_FAMILY, 8, QFont.Bold)
        cf = QFont(FONT_FAMILY, 8)

        def _mi(txt, font, is_header=False):
            it = QTableWidgetItem(txt)
            it.setFont(font)
            it.setTextAlignment(Qt.AlignCenter)
            it.setBackground(QBrush(header_bg if is_header else base_bg))
            it.setForeground(QBrush(txt_color))
            return it

        self._table.setItem(0, 0, _mi("종점 도착 예정 버스 번호", hf, True))
        self._table.setItem(1, 0, _mi("종점 도착까지 남은 시간", hf, True))

        for ci, bus in enumerate(cols):
            title = f"{bus['raw_plain_no']} (막차)" if bus['is_last_bus'] else bus['raw_plain_no']
            self._table.setItem(0, ci + 1, _mi(title, hf, True))
            self._table.setItem(1, ci + 1, _mi(format_remain_time(bus['seconds']), cf, False))
            self._table_bus_info.append({"col": ci + 1, "b_idx": bus['b_idx']})

        self._table.horizontalHeader().setDefaultSectionSize(bus_col_w)
        self._table.setColumnWidth(0, label_col_w)

# ──────────────────────────────────────────────────────────
# 【4-16】 _tick_countdown()
#   _tick_timer에 의해 1초마다 자동 호출.
#   _bus_seconds의 각 버스 남은 시간을 1씩 감소시키고
#   표(_table) "남은 시간" 셀을 즉시 갱신.
#   → API 갱신 주기 사이에도 카운트다운이 부드럽게 작동.
# ──────────────────────────────────────────────────────────
    def _tick_countdown(self):
        txt_color = get_text_color()
        for k in list(self._bus_seconds.keys()):
            if self._bus_seconds[k] > 0:
                self._bus_seconds[k] -= 1
        for info in self._table_bus_info:
            sec = self._bus_seconds.get(info['b_idx'], 0)
            item = self._table.item(1, info['col'])
            if item:
                item.setText(format_remain_time(sec))
                item.setForeground(QBrush(txt_color))

# ──────────────────────────────────────────────────────────
# 【4-17】 pause_tick()  : _tick_timer 정지 (기록 중지 시 호출)
# 【4-18】 resume_tick() : _tick_timer 재시작 (기록 시작 시 호출)
# ──────────────────────────────────────────────────────────
    def pause_tick(self):
        self._tick_timer.stop()

    def resume_tick(self):
        self._tick_timer.start()


# ══════════════════════════════════════════════════════════
# 【5】 RecordTable 클래스
#   버스 운행 기록을 표 형태로 보여주는 위젯.
#   메인 창 우측에 "운행 출발 시각 기록"과 "운행 종료 시각 기록"
#   두 개가 세로로 배치됨.
#   컬럼: [시각(145px)] [노선(80px)] [차량번호(95px)] [상태(나머지)]
# ══════════════════════════════════════════════════════════
class RecordTable(QWidget):
# ──────────────────────────────────────────────────────────
# 【5-1】 __init__(title, parent=None)
#   표 위에 제목 라벨을 배치하고, 컬럼 너비·정렬·스타일을 설정.
#   - 편집 불가(NoEditTriggers): 실수 수정 방지
#   - 행 선택(SelectRows): 클릭 시 한 행 전체 선택
#   - 교대 배경색(AlternatingRowColors): 홀짝 행 배경이 달라 읽기 편함
#   - 세로 헤더 숨김: 줄 번호 표시 안 함
# ──────────────────────────────────────────────────────────
    def __init__(self, title, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(3, 3, 3, 3)
        lbl = QLabel(f"  {title}  ")
        lbl.setFont(QFont(FONT_FAMILY, 10, QFont.Bold))
        layout.addWidget(lbl)
        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["시각", "노선", "차량번호", "상태"])
        self.table.verticalHeader().setDefaultSectionSize(15)
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.Fixed)
        hdr.setSectionResizeMode(1, QHeaderView.Fixed)
        hdr.setSectionResizeMode(2, QHeaderView.Fixed)
        hdr.setSectionResizeMode(3, QHeaderView.Stretch)
        self.table.setColumnWidth(0, 145)
        self.table.setColumnWidth(1, 80)
        self.table.setColumnWidth(2, 95)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        layout.addWidget(self.table)

# ──────────────────────────────────────────────────────────
# 【5-2】 add_row(values)
#   표 맨 아래에 새 행을 삽입하고 scrollToBottom()으로 자동 스크롤.
#   시각·노선·차량번호(col 0~2)는 AlignCenter, 상태(col 3)는 기본 정렬.
#   @param values: [시각, 노선, 차량번호, 상태] 순서 리스트
# ──────────────────────────────────────────────────────────
    def add_row(self, values):
        row = self.table.rowCount()
        self.table.insertRow(row)
        for col, val in enumerate(values):
            it = QTableWidgetItem(str(val))
            if col < 3:
                it.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, col, it)
        self.table.scrollToBottom()

# ══════════════════════════════════════════════════════════
# 【6】 SeoulBusRecorder 클래스  ★ 프로그램의 핵심 ★
#   메인 창(QMainWindow)이자 프로그램 전체를 총괄하는 클래스.
#   메뉴, 노선 지도 패널, 운행 기록 표, 로그 창을 모두 포함하며
#   API 호출·운행 판정·엑셀 저장을 관리.
#
#   시그널(Signal) 목록 - 스레드 간 안전한 UI 업데이트용:
#   sig_log          (str)         → 로그 메시지 전달
#   sig_record       (int, tuple)  → 운행 기록 표 추가
#   sig_update_map   (str, list)   → 지도 버스 위치 갱신
#   sig_clear_map    (str)         → 지도 버스 제거
#   sig_search_done  (list)        → 노선 검색 결과 전달
#   sig_route_loaded (dict)        → 노선 정보 로드 완료
#   sig_key_verified (str,str,str,str) → API 키 검증 완료
#
#   ※ 왜 시그널을 사용하나?
#   API 호출은 백그라운드 스레드에서 실행되지만,
#   Qt UI 업데이트는 반드시 메인(GUI) 스레드에서만 해야 한다.
#   시그널-슬롯이 이 스레드 간 안전한 통신을 자동으로 처리해준다.
# ══════════════════════════════════════════════════════════
class SeoulBusRecorder(QMainWindow):
    sig_log = Signal(str)
    sig_record = Signal(int, tuple)
    sig_update_map = Signal(str, list)
    sig_clear_map = Signal(str)
    sig_search_done = Signal(list)
    sig_route_loaded = Signal(dict)
    sig_key_verified = Signal(str, str, str, str)

    _SECRET = "l'existence précède l'essence"

# ──────────────────────────────────────────────────────────
# 【6-1】 __init__()
#   메인 창 초기화. 처리 순서:
#   1) 창 제목·크기 설정, 아이콘 적용
#   2) current_dir 결정 (실행파일/스크립트/macOS .app 각각 처리)
#   3) 설정 파일 경로 결정 (Seoul_Bus_Config.ini)
#   4) 인스턴스 변수 초기화:
#      api_key_main/back     : API 인증키 (빈 문자열로 시작)
#      is_monitoring         : 모니터링 중 여부 (False)
#      routes                : 모니터링할 노선 데이터 리스트
#      refresh_interval      : API 갱신 주기 초 (기본 25)
#      recorded_data         : [(시각,구분,정류소,노선,차번), ...] 기록 리스트
#      _saved_record_count   : 마지막 저장 시점 기록 수 (중복 저장 방지)
#      last_arrival_logs     : {(방향,rid,vn): timestamp} 최근 기록 시각
#                              → 같은 이벤트를 40분(2400초) 내 중복 기록 방지
#      departed_vehicles     : {(rid,vn): timestamp} 출발 감지된 버스 목록
#                              → 6시간(21600초) 후 자동 삭제
#      pos_suspend_until     : {rid: datetime} 운행 종료 후 API 호출 중지 시각
#      _save_lock            : 엑셀 저장 동시 접근 방지 스레드 잠금
#      _refresh_lock         : 갱신 중복 실행 방지 스레드 잠금
#      api_stats_today/yesterday: 오늘/어제 API 호출 횟수 통계
#      route_map_panels      : {노선명: RouteMapPanel} 딕셔너리
#   5) 시그널-슬롯 연결
#   6) _load_config() → _setup_ui() → 초기 로그 출력
# ──────────────────────────────────────────────────────────
    def __init__(self):
        super().__init__()
        self.setWindowTitle("서울시내버스 노선 운행기록 수집 프로그램 v1.15")
        self.setMinimumSize(800, 500)
        self.resize(1350, 1000)
        try:
            pm = load_pixmap_from_b64(ICON_B64)
            if not pm.isNull():
                self.setWindowIcon(QIcon(pm))
        except:
            pass

        if getattr(sys, 'frozen', False):
            _exe = os.path.abspath(sys.executable)
            if sys.platform == "darwin" and ".app/Contents/MacOS" in _exe:
                self.current_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(_exe))))
            else:
                self.current_dir = os.path.dirname(_exe)
        else:
            self.current_dir = os.path.dirname(os.path.abspath(__file__))

        self._CFG_FILE = os.path.join(self.current_dir, "Seoul_Bus_Config.ini")

        self.api_key_main = ""
        self.api_key_back = ""

        self.is_monitoring = False
        self.routes = []
        self.routes_ready = True
        self.refresh_interval = 25
        self.recorded_data = []
        self._saved_record_count = 0
        self.last_arrival_logs = {}
        self.departed_vehicles = {}
        self.pos_suspend_until = {}
        self.pos_resume_logged = set()
        self.temp_pos1_data = {}
        self.temp_pos2_data = {}
        self._key_limit_notified = False
        self._last_date = datetime.now().date()
        self._completed_dates_saved = set()
        self._save_lock = threading.Lock()
        self._refresh_lock = threading.Lock()
        self.auto_save_path = None
        self.can_auto_save = False
        self.api_stats_today = {"POS1": 0, "POS2": 0, "SLST": 0, "RINF": 0, "SRCH": 0, "기타": 0}
        self.api_stats_yesterday = {"POS1": 0, "POS2": 0, "SLST": 0, "RINF": 0, "SRCH": 0, "기타": 0}
        self.route_map_panels = {}
        self.route_map_current = None

        self.sig_log.connect(self._slot_log)
        self.sig_record.connect(self._slot_record)
        self.sig_update_map.connect(self._slot_update_map)
        self.sig_clear_map.connect(self._slot_clear_map)
        self.sig_route_loaded.connect(self._slot_route_loaded)

        self._load_config()
        self._setup_ui()
        self.log("노선 검색 버튼으로 노선을 추가하세요.")

# ──────────────────────────────────────────────────────────
# 【6-2】 _make_info_html(link_color=None) → str
#   메뉴바 우측 모서리에 표시되는 제작자·데이터출처 정보 HTML 문자열 생성.
#   link_color: 현재 팔레트의 링크 색상. 테마 변경 시 함께 업데이트됨.
# ──────────────────────────────────────────────────────────
    def _make_info_html(self, link_color=None):
        if link_color is None:
            link_color = QApplication.palette().color(QPalette.Link).name()
        return (
            f'● 만든이 : 박 국 환 ( '
            f'<a href="mailto:ggoyong2@naver.com" style="color:{link_color};">ggoyong2@naver.com</a>'
            f' )   &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;● 데이터 출처 : 공공데이터포털 Open API ( '
            f'<a href="https://www.data.go.kr" style="color:{link_color};">https://www.data.go.kr</a> )')

# ──────────────────────────────────────────────────────────
# 【6-3】 _setup_ui()
#   메인 창의 메뉴바·레이아웃·위젯 전체를 구성.
#
#   메뉴 구성:
#   [메뉴]      : 인증키 입력 / 갱신주기 입력 / 기록 시작 / 프로그램 종료
#   [노선 검색] : 노선 검색 열기
#   [테마]      : 라이트모드 / 다크모드
#   [API 현황]  : API 호출 현황 보기
#   [프로그램 정보]: 프로그램 정보 보기
#   (우측 모서리): 제작자 정보 HTML 라벨
#
#   레이아웃 구조:
#   QVBoxLayout(메인)
#   └─ outer_splitter (세로 분할자)
#      ├─ mid_splitter (가로 분할자)
#      │  ├─ map_stack      ← RouteMapPanel들이 쌓이는 스택 위젯
#      │  └─ right_widget
#      │     ├─ table_depart  ← 출발 기록 표 (RecordTable)
#      │     └─ table_arrive  ← 종료 기록 표 (RecordTable)
#      └─ log_container
#         ├─ log_header ("로그" 라벨 + "접기" 버튼)
#         └─ log_text  (QPlainTextEdit, 최대 3000줄)
# ──────────────────────────────────────────────────────────
    def _setup_ui(self):
        menubar = self.menuBar()
        menubar.setStyleSheet("QMenu::item:disabled { color: #AAAAAA; }")

        menu_main = menubar.addMenu("메뉴")
        menu_main.addAction("인증키 입력").triggered.connect(self._show_key_input)
        menu_main.addAction("갱신주기 입력").triggered.connect(self._ask_interval)
        self.act_toggle = menu_main.addAction("기록 시작")
        self.act_toggle.setEnabled(False)
        self.act_toggle.triggered.connect(self._on_toggle)
        menu_main.addSeparator()
        menu_main.addAction("프로그램 종료").triggered.connect(self.close)

        menu_search = menubar.addMenu("노선 검색")
        self.act_search = menu_search.addAction("노선 검색 열기")
        self.act_search.triggered.connect(self._show_route_search)

        menu_theme = menubar.addMenu("테마")
        menu_theme.addAction("라이트모드").triggered.connect(lambda: self._apply_theme("light"))
        menu_theme.addAction("다크모드").triggered.connect(lambda: self._apply_theme("dark"))

        menu_api = menubar.addMenu("API 현황")
        menu_api.addAction("API 현황 보기").triggered.connect(self._show_api_status)

        menu_info = menubar.addMenu("프로그램 정보")
        menu_info.addAction("프로그램 정보 보기").triggered.connect(self._show_program_info)

        # 메뉴바 오른쪽에 정보 라벨 삽입
        self._info_lbl = QLabel(self._make_info_html())
        self._info_lbl.setFont(QFont(FONT_FAMILY, 8))
        self._info_lbl.setOpenExternalLinks(True)
        self._info_lbl.setTextFormat(Qt.RichText)
        self._info_lbl.setContentsMargins(20, 0, 10, 0)
        menubar.setCornerWidget(self._info_lbl, Qt.TopRightCorner)

        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(6, 6, 6, 6)

        outer_splitter = QSplitter(Qt.Vertical)
        main_layout.addWidget(outer_splitter)
        mid_splitter = QSplitter(Qt.Horizontal)
        outer_splitter.addWidget(mid_splitter)

        self.map_stack = QStackedWidget()
        mid_splitter.addWidget(self.map_stack)

        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(2)
        self.table_depart = RecordTable("운행 출발 시각 기록")
        self.table_arrive = RecordTable("운행 종료 시각 기록")
        right_layout.addWidget(self.table_depart)
        right_layout.addWidget(self.table_arrive)
        mid_splitter.addWidget(right_widget)

        mid_splitter.setSizes([740, 560])
        mid_splitter.setStretchFactor(0, 1)
        mid_splitter.setStretchFactor(1, 1)

        log_container = QWidget()
        log_container_layout = QVBoxLayout(log_container)
        log_container_layout.setContentsMargins(0, 0, 0, 0)
        log_container_layout.setSpacing(0)

        log_header = QHBoxLayout()
        log_header.setContentsMargins(6, 0, 6, 0)
        log_title_btn = QLabel("  로그  ")
        log_title_btn.setFont(QFont(FONT_FAMILY, 9, QFont.Bold))
        log_header.addWidget(log_title_btn)
        log_header.addStretch()
        self._log_toggle_btn = QPushButton("접기")
        self._log_toggle_btn.setFixedSize(45, 16)
        self._log_toggle_btn.setFont(QFont(FONT_FAMILY, 7))
        self._log_toggle_btn.clicked.connect(self._toggle_log_panel)
        log_header.addWidget(self._log_toggle_btn)
        log_container_layout.addLayout(log_header)

        self.log_text = QPlainTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont(FONT_MONO, 8))
        self.log_text.setMaximumBlockCount(3000)
        self._log_text_widget = self.log_text
        log_container_layout.addWidget(self.log_text)

        outer_splitter.addWidget(log_container)
        outer_splitter.setSizes([700, 154])
        self._outer_splitter = outer_splitter
        self._log_container = log_container
        self._log_expanded = True
        self._log_last_height = 200

        line_h = self.log_text.fontMetrics().lineSpacing()
        self._log_min_height = line_h * 5 + 30
        self.log_text.setMinimumHeight(line_h * 5 + 10)
        log_container.setMinimumHeight(self._log_min_height)
        outer_splitter.setCollapsible(outer_splitter.indexOf(log_container), False)

# ──────────────────────────────────────────────────────────
# 【6-4】 _toggle_log_panel()
#   로그 창 접기/펼치기 토글.
#   접기  : log_text를 숨기고, 로그 컨테이너를 24px(헤더만)으로 줄임.
#   펼치기: log_text를 보여주고, 이전 저장 높이로 복원.
#   outer_splitter.setSizes([위 크기, 아래 크기])로 영역 재배분.
# ──────────────────────────────────────────────────────────
    def _toggle_log_panel(self):
        if self._log_expanded:
            sizes = self._outer_splitter.sizes()
            self._log_last_height = sizes[1]
            self._log_text_widget.hide()
            self._log_container.setMinimumHeight(0)
            self.log_text.setMinimumHeight(0)
            self._outer_splitter.setSizes([sizes[0] + sizes[1] - 24, 24])
            self._log_toggle_btn.setText("펼치기")
            self._log_expanded = False
        else:
            self._log_text_widget.show()
            self._log_container.setMinimumHeight(self._log_min_height)
            self.log_text.setMinimumHeight(self._log_min_height - 60)
            sizes = self._outer_splitter.sizes()
            restore_h = max(self._log_last_height, self._log_min_height)
            self._outer_splitter.setSizes([sizes[0] + sizes[1] - restore_h, restore_h])
            self._log_toggle_btn.setText("접기")
            self._log_expanded = True

# ──────────────────────────────────────────────────────────
# 【6-5】 _apply_theme(mode)
#   라이트/다크 테마를 앱 전체에 적용.
#   ① QApplication에 새 팔레트 적용
#   ② allWidgets()로 모든 위젯에 팔레트 강제 적용
#   ③ 메뉴바 우측 정보 라벨 링크 색상 갱신
#   ④ 모든 RouteMapPanel.refresh_theme() 호출
# ──────────────────────────────────────────────────────────
    def _apply_theme(self, mode):
        app = QApplication.instance()
        app.setStyle(QStyleFactory.create("Fusion"))
        p = _make_palette(mode)
        app.setPalette(p)
        for widget in app.allWidgets():
            widget.setPalette(p)
        if hasattr(self, '_info_lbl'):
            self._info_lbl.setText(self._make_info_html(p.color(QPalette.Link).name()))
        for panel in self.route_map_panels.values():
            panel.refresh_theme()

# ──────────────────────────────────────────────────────────
# 【6-6】 _slot_log(msg) [슬롯]
#   sig_log 수신 → log_text에 "[HH:MM:SS] 메시지" 형식으로 추가.
#
# 【6-7】 _slot_record(idx, entry) [슬롯]
#   sig_record 수신 → idx=0이면 table_depart, 1이면 table_arrive에 추가.
#
# 【6-8】 _slot_update_map(rnm, buses) [슬롯]
#   sig_update_map 수신 → 해당 노선 RouteMapPanel의 update_buses() 호출.
#
# 【6-9】 _slot_clear_map(rnm) [슬롯]
#   sig_clear_map 수신 → 해당 노선 지도에 빈 리스트 전달하여 버스 제거.
#
# 【6-10】 log(msg)
#   sig_log를 emit하는 헬퍼. 백그라운드 스레드에서 UI를 안전하게 업데이트.
# ──────────────────────────────────────────────────────────
    @Slot(str)
    def _slot_log(self, msg):
        self.log_text.appendPlainText(datetime.now().strftime("[%H:%M:%S] ") + msg)

    @Slot(int, tuple)
    def _slot_record(self, idx, entry):
        (self.table_depart if idx == 0 else self.table_arrive).add_row(entry)

    @Slot(str, list)
    def _slot_update_map(self, rnm, buses):
        if rnm in self.route_map_panels:
            self.route_map_panels[rnm].update_buses(buses)

    @Slot(str)
    def _slot_clear_map(self, rnm):
        if rnm in self.route_map_panels:
            self.route_map_panels[rnm].update_buses([])

    def log(self, msg):
        self.sig_log.emit(msg)

# ──────────────────────────────────────────────────────────
# 【6-11】 _make_fernet() → Fernet or None
#   _SECRET 문장을 pbkdf2_hmac(SHA-256, 100000회 반복)으로
#   32바이트 키로 변환 후 urlsafe_b64encode하여 Fernet 암호화 객체 생성.
#   cryptography 라이브러리 없으면 None 반환.
#
# 【6-12】 _enc_key(raw) → str
#   평문 API 키를 Fernet으로 암호화. 실패 시 원본 반환.
#
# 【6-13】 _dec_key(enc) → str
#   암호화된 API 키를 복호화. 실패 시 원본 반환.
# ──────────────────────────────────────────────────────────
    def _make_fernet(self):
        if not _CRYPTO_OK:
            return None
        dk = _hl.pbkdf2_hmac(
            "sha256", self._SECRET.encode("utf-8"),
            b"SeoulBusSalt2025", iterations=100_000, dklen=32)
        return _Fernet(base64.urlsafe_b64encode(dk))

    def _enc_key(self, raw):
        if not raw or not _CRYPTO_OK:
            return raw
        try:
            return self._make_fernet().encrypt(raw.encode("utf-8")).decode("ascii")
        except:
            return raw

    def _dec_key(self, enc):
        if not enc or not _CRYPTO_OK:
            return enc
        try:
            return self._make_fernet().decrypt(enc.encode("ascii")).decode("utf-8")
        except:
            return enc

# ──────────────────────────────────────────────────────────
# 【6-14】 _load_config()
#   Seoul_Bus_Config.ini 파일에서 암호화된 API 키를 읽어 복호화 후 저장.
#   INI 구조: [keys] main_key = (암호화값) / back_key = (암호화값)
#   파일 없으면 조용히 리턴.
#
# 【6-15】 _save_config()
#   현재 API 키를 암호화하여 설정 파일에 저장.
#   기존 파일의 favorites 섹션 등 다른 설정은 보존.
# ──────────────────────────────────────────────────────────
    def _load_config(self):
        cfg = configparser.ConfigParser()
        if not os.path.exists(self._CFG_FILE):
            return
        try:
            cfg.read(self._CFG_FILE, encoding="utf-8")
            enc_m = cfg.get("keys", "main_key", fallback="")
            enc_b = cfg.get("keys", "back_key", fallback="")
            if enc_m:
                self.api_key_main = self._dec_key(enc_m)
            if enc_b:
                self.api_key_back = self._dec_key(enc_b)
        except:
            pass

    def _save_config(self):
        cfg = configparser.ConfigParser()
        if os.path.exists(self._CFG_FILE):
            try:
                cfg.read(self._CFG_FILE, encoding="utf-8")
            except:
                pass
        if "keys" not in cfg:
            cfg["keys"] = {}
        cfg["keys"]["main_key"] = self._enc_key(self.api_key_main)
        cfg["keys"]["back_key"] = self._enc_key(self.api_key_back)
        try:
            with open(self._CFG_FILE, "w", encoding="utf-8") as f:
                cfg.write(f)
        except:
            pass

# ──────────────────────────────────────────────────────────
# 【6-16】 _show_key_input()
#   API 인증키를 입력·검증하는 대화상자.
#
#   구성: 메인키(64자) 입력창 + 보조키(64자 또는 공란) 입력창
#         + 상태 라벨 + [입력][취소] 버튼
#
#   검증 과정:
#   ① 입력 중 실시간으로 글자 수 확인 → 64자이면 [입력] 버튼 활성화
#   ② [입력] 클릭 → 백그라운드 스레드에서 실제 API 호출(_test_key)
#   ③ sig_key_verified 시그널로 결과를 메인 스레드에 전달
#   ④ 결과 처리:
#      "REJECTED" → 등록되지 않은 키 메시지 (빨간 글자)
#      "OK"/"LIMIT" → 성공 → 키 저장 후 대화상자 닫기
#      "LIMIT" → 한도 초과 경고 후 저장 (여전히 유효한 키)
#      기타 오류 → 오류 메시지 표시
# ──────────────────────────────────────────────────────────
    def _show_key_input(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("인증키 입력")
        dlg.setFixedSize(560, 260)
        lay = QVBoxLayout(dlg)
        lay.addWidget(QLabel("메인 인증키 (64자리, 필수)"))
        edt_main = QLineEdit(self.api_key_main)
        edt_main.setFont(QFont(FONT_MONO, 8))
        lay.addWidget(edt_main)
        lay.addWidget(QLabel("보조 인증키 (64자리 또는 공란)"))
        edt_back = QLineEdit(self.api_key_back)
        edt_back.setFont(QFont(FONT_MONO, 8))
        lay.addWidget(edt_back)
        lbl_status = QLabel("")
        lbl_status.setStyleSheet("color: gray;")
        lay.addWidget(lbl_status)
        btn_bar = QHBoxLayout()
        btn_ok = QPushButton("입력")
        btn_ok.setEnabled(False)
        btn_cancel = QPushButton("취소")
        btn_bar.addStretch()
        btn_bar.addWidget(btn_ok)
        btn_bar.addWidget(btn_cancel)
        lay.addLayout(btn_bar)
        btn_cancel.clicked.connect(dlg.reject)

        def _check_len():
            km = len(edt_main.text().strip())
            kb = len(edt_back.text().strip())
            btn_ok.setEnabled(km == 64 and (kb == 0 or kb == 64))
        edt_main.textChanged.connect(lambda: _check_len())
        edt_back.textChanged.connect(lambda: _check_len())
        _check_len()

        def _do_ok():
            km = edt_main.text().strip()
            kb = edt_back.text().strip()
            lbl_status.setText("인증키 검증 중...")
            lbl_status.setStyleSheet("color: gray;")
            btn_ok.setEnabled(False)

            def _test_key(k):
                if not k:
                    return None
                try:
                    resp = requests.get(URL_SRCH, params={"ServiceKey": k, "strSrch": "13"}, timeout=8)
                    self.api_stats_today["SRCH"] += 1
                    raw_up = resp.text.upper()
                    if any(p in raw_up for p in ("TOO MANY", "LIMITED NUMBER", "LIMITED_NUMBER", "RATE LIMIT")):
                        return "LIMIT"
                    if any(p in raw_up for p in ("NOT REGISTERED", "UNREGISTERED", "SERVICE KEY IS NOT")):
                        return "REJECTED"
                    root = ET.fromstring(resp.text)
                    code = (root.findtext(".//headerCd") or "").strip()
                    if code in ("0", "3", "4"):
                        return "OK"
                    return f"오류코드: {code}"
                except Exception as e:
                    return f"오류: {e}"

            _results = [None, None]

            def _key_run(): 
                _results[0] = _test_key(km)
                _results[1] = _test_key(kb)
                self.sig_key_verified.emit(
                    str(_results[0] or ""),
                    str(_results[1] or ""),
                    km, kb)

            def _on_verified(rm_str, rb_str, key_m, key_b):
                if not dlg.isVisible():
                    return
                rm = None if rm_str == "" else rm_str
                rb = None if rb_str == "" else rb_str
                if rm == "REJECTED":
                    lbl_status.setText("메인키: 등록되지 않은 인증키")
                    lbl_status.setStyleSheet("color: red;")
                    btn_ok.setEnabled(True)
                elif rm not in ("OK", "LIMIT", None):
                    lbl_status.setText(f"메인키: {rm}")
                    lbl_status.setStyleSheet("color: red;")
                    btn_ok.setEnabled(True)
                elif rb == "REJECTED":
                    lbl_status.setText("보조키: 등록되지 않은 인증키")
                    lbl_status.setStyleSheet("color: red;")
                    btn_ok.setEnabled(True)
                elif rb not in ("OK", "LIMIT", None) and key_b:
                    lbl_status.setText(f"보조키: {rb}")
                    lbl_status.setStyleSheet("color: red;")
                    btn_ok.setEnabled(True)
                else:
                    self.api_key_main = key_m
                    self.api_key_back = key_b
                    self._save_config()
                    if self.routes:
                        self.act_toggle.setEnabled(True)
                    self.log(f"인증키 등록 완료 (메인: 설정됨, 보조: {'설정됨' if key_b else '없음'})")
                    if rm == "LIMIT" or rb == "LIMIT":
                        self.log("⚠ 일부 키가 한도초과 상태입니다 (유효)")
                    dlg.accept()

            self.sig_key_verified.connect(_on_verified)
            threading.Thread(target=_key_run, daemon=True).start()

        btn_ok.clicked.connect(_do_ok)
        dlg.exec()
        try:
            self.sig_key_verified.disconnect(_on_verified)
        except (TypeError, RuntimeError):
            pass
        
# ──────────────────────────────────────────────────────────
# 【6-17】 _show_route_search()
#   노선 검색 및 즐겨찾기 관리 대화상자.
#
#   구성:
#   - 검색어 입력창 + [검색] 버튼
#   - 검색 결과 트리위젯 (노선번호·유형·기점·종점·첫차·막차·간격·길이)
#   - ★ 즐겨찾기 목록 트리위젯 (동일 구조, 최대 150px)
#   - [★즐겨찾기 추가][✕삭제][노선 선택하기][닫기] 버튼
#
#   주요 동작:
#   - 검색: 백그라운드 스레드 → URL_SRCH API 호출 → sig_search_done → 결과 표시
#   - 선택: [노선 선택하기] 또는 더블클릭/엔터
#           → _load_route_from_search() 스레드 실행
#           → sig_route_loaded → _slot_route_loaded
#   - 즐겨찾기: 추가/삭제 후 _save_favorites() 로 설정 파일에 저장
#   - 창 크기 변경 시 컬럼 너비 비례 재배분 (지연 타이머 사용)
# ──────────────────────────────────────────────────────────
    def _show_route_search(self):
        if hasattr(self, '_search_dlg') and self._search_dlg and self._search_dlg.isVisible():
            self._search_dlg.raise_()
            self._search_dlg.activateWindow()
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("노선 검색")
        dlg.resize(875, 580)
        dlg.setMinimumWidth(875)
        dlg.setModal(True)
        self._search_dlg = dlg
        lay = QVBoxLayout(dlg)

        top = QHBoxLayout()
        edt = QLineEdit()
        edt.setPlaceholderText("노선번호 입력 (예: 101, 7016, N37, 알파벳은 대문자로 검색)")
        edt.setFont(QFont(FONT_FAMILY, 11))
        btn_search = QPushButton("검색")
        btn_search.setEnabled(False)
        top.addWidget(edt)
        top.addWidget(btn_search)
        lay.addLayout(top)

        hint = QLabel("")
        hint.setFont(QFont(FONT_FAMILY, 9))
        hint.setAlignment(Qt.AlignCenter)
        lay.addWidget(hint)

        col_labels = ["노선번호", "노선유형", "기점", "종점",
                      "첫차출발시각", "막차출발시각", "간격(분)", "노선길이(km)"]

        base_col_widths = [103, 63, 180, 180, 95, 95, 50, 72]
        base_total = sum(base_col_widths)

        def _apply_proportional_widths(tw, available_width):
            """트리위젯의 컬럼을 available_width에 비례하여 배분"""
            if available_width < base_total:
                available_width = base_total
            for i, bw in enumerate(base_col_widths):
                new_w = int(available_width * bw / base_total)
                tw.setColumnWidth(i, new_w)

        tree = QTreeWidget()
        tree.setHeaderLabels(col_labels)
        tree.setFont(QFont(FONT_FAMILY, 9))
        tree.setRootIsDecorated(False)
        tree.setAlternatingRowColors(True)
        hdr = tree.header()
        hdr.setSectionResizeMode(QHeaderView.Interactive)
        hdr.setDefaultAlignment(Qt.AlignCenter)
        for i, w in enumerate(base_col_widths):
            tree.setColumnWidth(i, w)
        lay.addWidget(tree)

        fav_tree = QTreeWidget()
        fav_tree.setHeaderLabels(col_labels)
        fav_tree.setFont(QFont(FONT_FAMILY, 9))
        fav_tree.setRootIsDecorated(False)
        fav_tree.setAlternatingRowColors(True)
        fav_hdr = fav_tree.header()
        fav_hdr.setSectionResizeMode(QHeaderView.Interactive)
        fav_hdr.setDefaultAlignment(Qt.AlignCenter)
        for i, w in enumerate(base_col_widths):
            fav_tree.setColumnWidth(i, w)
        fav_tree.setMaximumHeight(150)

        fav_label = QLabel("  ★ 즐겨찾기 노선  ")
        fav_label.setFont(QFont(FONT_FAMILY, 9, QFont.Bold))
        lay.addWidget(fav_label)
        lay.addWidget(fav_tree)

        btn_bar = QHBoxLayout()
        btn_fav_add = QPushButton("★ 즐겨찾기 추가")
        btn_fav_del = QPushButton("✕ 즐겨찾기 삭제")
        btn_load = QPushButton("노선 선택하기")
        btn_load.setEnabled(False)
        btn_close = QPushButton("닫기")
        btn_bar.addWidget(btn_fav_add)
        btn_bar.addWidget(btn_fav_del)
        btn_bar.addStretch()
        btn_bar.addWidget(btn_load)
        btn_bar.addWidget(btn_close)
        lay.addLayout(btn_bar)
        btn_close.clicked.connect(dlg.close)

        favorites = self._load_favorites()
        _active_tree = [None]

        _resize_timer = QTimer(dlg)
        _resize_timer.setSingleShot(True)
        _resize_timer.setInterval(100)

        def _on_dlg_resize():
            available = tree.viewport().width()
            if available > base_total:
                _apply_proportional_widths(tree, available)
                _apply_proportional_widths(fav_tree, available)

        _resize_timer.timeout.connect(_on_dlg_resize)

        _orig_resize = dlg.resizeEvent
        def _dlg_resize_event(event):
            if _orig_resize:
                _orig_resize(event)
            _resize_timer.start()
        dlg.resizeEvent = _dlg_resize_event

        def _fmt_time(raw):
            digits = ''.join(c for c in str(raw).strip() if c.isdigit())
            if len(digits) >= 12:
                return f"{int(digits[8:10])}시 {int(digits[10:12])}분"
            if len(digits) >= 4:
                return f"{int(digits[:2])}시 {int(digits[2:4])}분"
            return raw

        def _fmt_length(raw):
            raw = str(raw).strip()
            if not raw:
                return raw
            try:
                val = float(raw)
                if val >= 100:
                    return f"{val / 1000:.1f}"
                else:
                    return f"{val:.1f}"
            except ValueError:
                return raw

        def _make_item(r):
            texts = [
                r.get("name", ""), r.get("type", ""),
                r.get("st", ""), r.get("ed", ""),
                _fmt_time(r.get("first", "")),
                _fmt_time(r.get("last", "")),
                r.get("term", ""),
                _fmt_length(r.get("length", ""))]
            item = QTreeWidgetItem(texts)
            for i in range(len(texts)):
                item.setTextAlignment(i, Qt.AlignCenter)
            return item

        def _populate_fav():
            fav_tree.clear()
            for fav in favorites:
                item = _make_item(fav)
                item.setData(0, Qt.UserRole, fav.get("route_id", ""))
                item.setData(0, Qt.UserRole + 1, fav)
                fav_tree.addTopLevelItem(item)

        _populate_fav()

        def _on_search_text_changed(text):
            btn_search.setEnabled(len(text.strip()) > 0)

        edt.textChanged.connect(_on_search_text_changed)

        def _do_search():
            q = edt.text().strip()
            if not q:
                return
            if not self.api_key_main:
                hint.setText("인증키를 먼저 입력하세요.")
                return
            hint.setText("검색중...")
            btn_search.setEnabled(False)
            tree.clear()

            def _search_run():
                rows = []
                try:
                    p = {"ServiceKey": unquote(self.api_key_main), "strSrch": q}
                    resp = requests.get(URL_SRCH, params=p, timeout=10)
                    self.api_stats_today["SRCH"] += 1
                    root = ET.fromstring(resp.content)
                    for it in root.iter("itemList"):
                        rows.append({
                            "route_id": (it.findtext("busRouteId") or "").strip(),
                            "name": (it.findtext("busRouteNm") or "").strip(),
                            "type": ROUTE_TYPE_LABEL.get((it.findtext("routeType") or "").strip(), "기타"),
                            "route_type_raw": (it.findtext("routeType") or "").strip(),
                            "st": (it.findtext("stStationNm") or "").strip(),
                            "ed": (it.findtext("edStationNm") or "").strip(),
                            "first": (it.findtext("firstBusTm") or "").strip(),
                            "last": (it.findtext("lastBusTm") or "").strip(),
                            "term": (it.findtext("term") or "").strip(),
                            "length": (it.findtext("length") or "").strip(),
                        })
                except:
                    pass
                self.sig_search_done.emit(rows)

            threading.Thread(target=_search_run, daemon=True).start()

        def _on_search_done(rows):
            btn_search.setEnabled(len(edt.text().strip()) > 0)
            if not dlg.isVisible():
                return
            tree.clear()
            if not rows:
                hint.setText("검색 결과가 없습니다.")
                return
            for r in rows:
                item = _make_item(r)
                item.setData(0, Qt.UserRole, r["route_id"])
                item.setData(0, Qt.UserRole + 1, r)
                tree.addTopLevelItem(item)
            hint.setText(f"검색 결과 {len(rows)}건")

        self.sig_search_done.connect(_on_search_done)

        def _update_load_btn():
            cur = _get_selected()
            btn_load.setEnabled(cur is not None)

        def _on_tree_sel():
            _active_tree[0] = tree
            fav_tree.clearSelection()
            _update_load_btn()

        def _on_fav_sel():
            _active_tree[0] = fav_tree
            tree.clearSelection()
            _update_load_btn()

        tree.currentItemChanged.connect(_on_tree_sel)
        fav_tree.currentItemChanged.connect(_on_fav_sel)

        def _get_selected():
            t = _active_tree[0]
            if t and t.currentItem():
                return t.currentItem()
            return None

        def _add_fav():
            cur = tree.currentItem()
            if not cur:
                return
            rid = cur.data(0, Qt.UserRole)
            for f in favorites:
                if f.get("route_id") == rid:
                    hint.setText("이미 즐겨찾기에 있습니다.")
                    return
            rdata = cur.data(0, Qt.UserRole + 1)
            if rdata:
                favorites.append(rdata)
                self._save_favorites(favorites)
                _populate_fav()
                hint.setText(f"'{cur.text(0)}' 즐겨찾기 추가됨")

        def _del_fav():
            cur = fav_tree.currentItem()
            if not cur:
                return
            rid = cur.data(0, Qt.UserRole)
            before = len(favorites)
            favorites[:] = [f for f in favorites if f.get("route_id") != rid]
            if len(favorites) < before:
                self._save_favorites(favorites)
                _populate_fav()
                hint.setText("즐겨찾기에서 삭제됨")

        btn_fav_add.clicked.connect(_add_fav)
        btn_fav_del.clicked.connect(_del_fav)

        def _do_load():
            cur = _get_selected()
            if not cur:
                return
            rid = cur.data(0, Qt.UserRole)
            rdata = cur.data(0, Qt.UserRole + 1)
            rname = cur.text(0)
            rtype = rdata.get("route_type_raw", "3") if rdata else "3"
            hint.setText(f"'{rname}' 노선 정보 불러오는 중...")
            btn_load.setEnabled(False)
            def _load_run():
                self._load_route_from_search(rid, rname, rtype, dlg)
            threading.Thread(target=_load_run, daemon=True).start()

        btn_load.clicked.connect(_do_load)
        btn_search.clicked.connect(_do_search)
        edt.returnPressed.connect(_do_search)

        tree.itemDoubleClicked.connect(lambda item, col: _do_load())
        fav_tree.itemDoubleClicked.connect(lambda item, col: _do_load())

        tree.itemActivated.connect(lambda item, col: _do_load())
        fav_tree.itemActivated.connect(lambda item, col: _do_load())

        dlg.exec()
        try:
            self.sig_search_done.disconnect(_on_search_done)
        except:
            pass
        self._search_dlg = None

# ──────────────────────────────────────────────────────────
# 【6-18】 _load_route_from_search(route_id, route_name, route_type, dlg)
#   백그라운드 스레드에서 실행. 선택 노선의 상세 정보를 API로 수집.
#
#   API 호출:
#   ① URL_SLST → 정류소 목록 (seq 순 정렬)
#   ② URL_RINF → 회차 순번·운수사·첫차·막차·노선 길이 추출
#
#   route_data 주요 키:
#   stops, turn_index, first/second/last _st_id/nm/ars,
#   corp_nm, first/last_bus_tm, rlength, rtype
#
#   완성된 route_data를 sig_route_loaded.emit()으로 메인 스레드에 전달.
# ──────────────────────────────────────────────────────────
    def _load_route_from_search(self, route_id, route_name, route_type, dlg):
        root_slst = self.fetch_api(URL_SLST, {"busRouteId": route_id})
        if root_slst is None or isinstance(root_slst, tuple):
            self.log(f"[오류] '{route_name}' 정류소 목록 API 실패")
            self.sig_route_loaded.emit({})
            return

        root_rinf = self.fetch_api(URL_RINF, {"busRouteId": route_id})
        turn_seq = "0"
        corp_nm = "서울시내버스"
        first_bus_tm = None
        last_bus_tm = None
        rlength = ""
        if root_rinf is not None and not isinstance(root_rinf, tuple):
            for it in root_rinf.iter("itemList"):
                ts = (it.findtext("turnSeq") or "0").strip()
                if ts and ts != "0":
                    turn_seq = ts
                cn = (it.findtext("corpNm") or "").strip()
                if cn:
                    corp_nm = cn
                ft = (it.findtext("firstBusTm") or "").strip()
                lt = (it.findtext("lastBusTm") or "").strip()
                if ft:
                    first_bus_tm = format_hhmm(ft)
                if lt:
                    last_bus_tm = format_hhmm(lt)
                rl = (it.findtext("length") or "").strip()
                if rl:
                    rlength = rl
                rt = (it.findtext("routeType") or "").strip()
                if rt:
                    route_type = rt
                break

        stops = []
        for it in root_slst.iter("itemList"):
            sid = (it.findtext("station") or "").strip()
            seq_raw = (it.findtext("seq") or "0").strip()
            nm = (it.findtext("stationNm") or "").strip()
            arsid = (it.findtext("arsId") or "").strip()
            trans = (it.findtext("transYn") or "N").strip().upper()
            full_sect = (it.findtext("fullSectDist") or "0").strip()
            stops.append({
                "station": sid, "seq": seq_raw, "name": nm,
                "arsId": arsid, "transYn": trans, "fullSectDist": full_sect,
            })
        stops.sort(key=lambda s: int(s["seq"]) if s["seq"].isdigit() else 0)

        if not stops:
            self.log(f"[오류] '{route_name}' 정류소 없음")
            self.sig_route_loaded.emit({})
            return

        turn_idx = 0
        try:
            ti = int(turn_seq)
            if 0 < ti <= len(stops):
                turn_idx = ti - 1
        except:
            turn_idx = len(stops) // 2

        color = ROUTE_TYPE_COLOR.get(route_type, DEFAULT_LINE_COLOR)
        rkey = route_name
        n = len(stops)

        route_data = {
            "key": rkey, "route_id": route_id, "route_name": route_name,
            "route_type": route_type, "route_color": color,
            "stops": stops, "turn_index": turn_idx,
            "rnm": route_name, "rid": route_id, "st_cnt": n,
            "first_st_id": stops[0]["station"] if stops else "",
            "second_st_id": stops[1]["station"] if n >= 2 else "",
            "last_st_id": stops[-1]["station"] if stops else "",
            "first_nm": stops[0]["name"] if stops else "?",
            "second_nm": stops[1]["name"] if n >= 2 else "?",
            "last_nm": stops[-1]["name"] if stops else "?",
            "first_ars": stops[0]["arsId"] if stops else "?",
            "second_ars": stops[1]["arsId"] if n >= 2 else "?",
            "last_ars": stops[-1]["arsId"] if stops else "?",
            "corp_nm": corp_nm, "first_bus_tm": first_bus_tm,
            "last_bus_tm": last_bus_tm, "rtype": route_type,
            "rlength": rlength, "stations": stops, "dlg": dlg,
        }
        self.sig_route_loaded.emit(route_data)

# ──────────────────────────────────────────────────────────
# 【6-19】 _slot_route_loaded(data) [슬롯]
#   sig_route_loaded 수신 → 새 노선을 메인 창에 등록.
#   ① routes 리스트 교체
#   ② 해당 노선 RouteMapPanel 없으면 생성 후 map_stack에 추가
#   ③ panel.load_route() → map_stack.setCurrentWidget()
#   ④ API 키+노선 모두 있으면 [기록 시작] 버튼 활성화
#   ⑤ 노선 정보 로그 출력
#   ⑥ 검색 대화상자 닫기 (dlg.accept)
# ──────────────────────────────────────────────────────────
    @Slot(dict)
    def _slot_route_loaded(self, data):
        if not data:
            return
        rnm = data.get("rnm") or data.get("route_name", "")
        route_name = data.get("route_name", rnm)
        stops = data["stops"]
        turn_idx = data["turn_index"]
        dlg = data.get("dlg")
        n = len(stops)

        self.routes = [data]

        if rnm not in self.route_map_panels:
            panel = RouteMapPanel(rnm)
            self.map_stack.addWidget(panel)
            self.route_map_panels[rnm] = panel

        panel = self.route_map_panels[rnm]
        panel.load_route(stops, data.get("rtype", "3"), data.get("rlength", ""))
        self.map_stack.setCurrentWidget(panel)

        if self.api_key_main and self.routes:
            self.act_toggle.setEnabled(True)

        self.log(f"[{route_name}] 정류소 {n}개 · "
                 f"첫정류소: {data.get('first_nm','?')} / "
                 f"두번째: {data.get('second_nm','?')} / 종점: {data.get('last_nm','?')}")
        self.log(f"[{route_name}] 노선 등록 완료 · "
                 f"운수사: {data.get('corp_nm','')} / "
                 f"첫차: {data.get('first_bus_tm','')} / 막차: {data.get('last_bus_tm','')}")

        if dlg and dlg.isVisible():
            try:
                dlg.accept()
            except:
                pass

# ──────────────────────────────────────────────────────────
# 【6-20】 fetch_api(url, params) → XML root | tuple | None  ★ API 공통 호출 ★
#   메인 키 → 보조 키 순으로 API 호출 시도.
#
#   반환값:
#   XML root Element  : 정상 응답 (headerCd == "0")
#   ('NO_BUS', 첫차시각) : 결과 없음 (데이터 없음 메시지)
#   None              : 모든 키 실패
#
#   처리 흐름 (각 키에 대해):
#   ① requests.get(url, params={ServiceKey:key, ...}, timeout=10)
#   ② 한도 초과 패턴 감지 → lh(한도초과 횟수) 증가 후 다음 키
#      패턴: "TOO MANY", "LIMITED NUMBER", "RATE LIMIT", "<headerCd>22</headerCd>"
#   ③ 미등록 키 감지 → 다음 키
#   ④ HTTP 상태코드 != 200 → 다음 키
#   ⑤ XML 파싱 → headerCd "0" → 성공 반환
#   ⑥ "결과가 없습니다"/"NODATA" → ('NO_BUS', ...) 반환
#   ⑦ 모든 키 한도 초과 시 경고 로그 (1회 출력 후 _key_limit_notified=True)
#   ⑧ API 호출 통계(api_stats_today) 갱신
# ──────────────────────────────────────────────────────────
    def fetch_api(self, url, params):
        lh = 0
        kc = sum(1 for k in [self.api_key_main, self.api_key_back] if k)
        for key in [self.api_key_main, self.api_key_back]:
            if not key:
                continue
            try:
                p = dict(params)
                p['ServiceKey'] = unquote(key)
                resp = requests.get(url, params=p, timeout=10)
                LP = ("TOO MANY", "LIMITED NUMBER", "LIMITED_NUMBER", "RATE LIMIT", "<headerCd>22</headerCd>")
                if any(pt in resp.text.upper() for pt in LP):
                    lh += 1
                    continue
                if "SERVICE KEY IS NOT REGISTERED" in resp.text or "UNREGISTERED_KEY" in resp.text:
                    continue
                if resp.status_code != 200:
                    continue
                root = ET.fromstring(resp.text)
                hc = root.findtext(".//headerCd") or ""
                em = root.findtext(".//headerMsg") or ""
                if URL_POS1 in url:
                    self.api_stats_today["POS1"] += 1
                elif URL_POS2 in url:
                    self.api_stats_today["POS2"] += 1
                elif URL_SLST in url:
                    self.api_stats_today["SLST"] += 1
                elif URL_RINF in url:
                    self.api_stats_today["RINF"] += 1
                elif URL_SRCH in url:
                    self.api_stats_today["SRCH"] += 1
                else:
                    self.api_stats_today["기타"] += 1
                if hc == "0":
                    return root
                nd = "결과가 없습니다" in em or "NODATA" in em
                if nd:
                    if URL_POS2 in url:
                        return ('NO_BUS', self._fetch_first_time(params.get('busRouteId', '')))
                    else:
                        return ('NO_BUS', None)
            except:
                continue
        if lh > 0 and lh >= kc and kc > 0 and not self._key_limit_notified:
            self._key_limit_notified = True
            self.log("⚠ 모든 인증키의 호출 한도가 초과되었습니다.")
        return None

# ──────────────────────────────────────────────────────────
# 【6-21】 _fetch_first_time(rid) → "HH:MM" or None
#   URL_RINF API로 특정 노선의 첫차 출발 시각 조회.
#   운행 종료 감지 후 다음 첫차까지 API 호출을 일시 중지할 때 필요.
# ──────────────────────────────────────────────────────────
    def _fetch_first_time(self, rid):
        if not rid:
            return None
        try:
            r = self.fetch_api(URL_RINF, {'busRouteId': rid})
            if r is None or isinstance(r, tuple):
                return None
            raw = r.findtext(".//firstBusTm") or ""
            if raw:
                return format_hhmm(raw)
        except:
            pass
        return None

# ──────────────────────────────────────────────────────────
# 【6-22】 _ask_interval()  ← 첫 번째 정의 (줄 2050의 두 번째 정의로 덮어씌워짐)
#   갱신 주기(초, 최소 10초) 입력 대화상자. 메뉴 [갱신주기 입력]에서 호출.
# ──────────────────────────────────────────────────────────
    def _ask_interval(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("갱신주기 설정")
        dlg.setFixedSize(300, 140)
        lo = QVBoxLayout(dlg)
        lo.addWidget(QLabel("갱신주기(초)를 입력하세요 (10초 이상):"))
        ed = QLineEdit(str(self.refresh_interval))
        ed.setAlignment(Qt.AlignCenter)
        lo.addWidget(ed)
        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        lo.addWidget(bb)
        def ok():
            try:
                v = int(ed.text())
            except:
                QMessageBox.warning(dlg, "알림", "숫자를 입력해주세요.")
                return
            if v < 10:
                QMessageBox.warning(dlg, "알림", "갱신주기는 10초 이상이어야 합니다.")
                return
            self.refresh_interval = v
            self.log(f"갱신주기가 {v}초로 변경되었습니다.")
            dlg.accept()
        bb.accepted.connect(ok)
        bb.rejected.connect(dlg.reject)
        ed.returnPressed.connect(ok)
        dlg.exec()

# ──────────────────────────────────────────────────────────
# 【6-23】 _show_api_status()  ← 첫 번째 정의 (줄 2078의 두 번째 정의로 덮어씌워짐)
#   API 엔드포인트별 오늘·어제 호출 횟수를 표로 보여주는 대화상자.
#   공공데이터 API는 하루 호출 횟수 제한이 있으므로 사용량 모니터링에 활용.
# ──────────────────────────────────────────────────────────
    def _show_api_status(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("API 현황")
        lo = QVBoxLayout(dlg)
        t = QTableWidget(6, 4)
        t.setHorizontalHeaderLabels(["구분", "API URL", "오늘", "어제"])
        hdr = t.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.Fixed)
        hdr.setSectionResizeMode(1, QHeaderView.Stretch)
        hdr.setSectionResizeMode(2, QHeaderView.Fixed)
        hdr.setSectionResizeMode(3, QHeaderView.Fixed)
        t.setColumnWidth(0, 145)
        t.setColumnWidth(2, 75)
        t.setColumnWidth(3, 75)
        t.setEditTriggers(QAbstractItemView.NoEditTriggers)
        t.verticalHeader().setVisible(False)
        data = [
            ("운행 출발 판정 API", "POS1", URL_POS1),
            ("종점 도착 판정 API", "POS2", URL_POS2),
            ("노선 정류소 목록 API", "SLST", URL_SLST),
            ("버스 노선 정보 API", "RINF", URL_RINF),
            ("버스 노선 검색 API", "SRCH", URL_SRCH),
            ("기타", "기타", "기타"),
        ]
        for row, (lb, ky, url) in enumerate(data):
            t.setItem(row, 0, QTableWidgetItem(lb))
            t.setItem(row, 1, QTableWidgetItem(url))
            t.setItem(row, 2, QTableWidgetItem(f"{self.api_stats_today.get(ky, 0):,}회"))
            t.setItem(row, 3, QTableWidgetItem(f"{self.api_stats_yesterday.get(ky, 0):,}회"))
        lo.addWidget(t)
        row_h = t.verticalHeader().defaultSectionSize()
        header_h = t.horizontalHeader().height()
        total_h = header_h + row_h * 6 + 10
        t.setFixedHeight(total_h)
        dlg.resize(750, total_h + 50)
        dlg.exec()

# ──────────────────────────────────────────────────────────
# 【6-24】 _show_program_info()  ← 첫 번째 정의 (줄 2116의 두 번째 정의로 덮어씌워짐)
#   프로그램 버전·사용 API·제작자·라이선스 정보 대화상자.
#   CC(크리에이티브 커먼즈)와 공공누리 마크 이미지 포함.
# ──────────────────────────────────────────────────────────
    def _show_program_info(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("프로그램 정보")
        dlg.setMinimumSize(400, 400)
        lo = QVBoxLayout(dlg)
        lc = QApplication.palette().color(QPalette.Link).name()
        ih = (f'<p style="font-size:10pt; font-weight:bold;">서울시내버스 노선 운행기록 수집 프로그램 v1.15</p>'
              f'<p>이 프로그램은 Python + PySide6로 작성하고,<br>'
              f'공공누리 제1유형으로 개방한 공공데이터 API 서비스를 이용하였으며,<br>'
              f'API 서비스는 아래의 페이지에서 무료로 이용할 수 있습니다.</p>'
              f'<p>공공데이터포털<br><a href="https://www.data.go.kr" style="color:{lc};">https://www.data.go.kr</a></p>'
              f'<p>서울특별시_버스위치정보조회 서비스<br>'
              f'<a href="https://www.data.go.kr/data/15000332/openapi.do" style="color:{lc};">https://www.data.go.kr/data/15000332/openapi.do</a></p>'
              f'<p>서울특별시_노선정보조회 서비스<br>'
              f'<a href="https://www.data.go.kr/data/15000193/openapi.do" style="color:{lc};">https://www.data.go.kr/data/15000193/openapi.do</a></p>'
              f'<p>작성자 : 박 국 환 ( <a href="mailto:ggoyong2@naver.com" style="color:{lc};">ggoyong2@naver.com</a> )</p>')
        lb = QLabel(ih)
        lb.setFont(QFont(FONT_FAMILY, 9))
        lb.setWordWrap(True)
        lb.setOpenExternalLinks(True)
        lb.setTextFormat(Qt.RichText)
        lo.addWidget(lb)
        il = QHBoxLayout()
        for b64 in [CC_IMG_B64, GG_IMG_B64]:
            pm = load_pixmap_from_b64(b64)
            if not pm.isNull():
                il2 = QLabel()
                il2.setPixmap(pm)
                il.addWidget(il2)
        il.addStretch()
        lo.addLayout(il)
        btn = QPushButton("닫기")
        btn.clicked.connect(dlg.accept)
        lo.addWidget(btn, alignment=Qt.AlignCenter)
        dlg.exec()

# ──────────────────────────────────────────────────────────
# 【6-25】 _on_toggle()
#   [기록 시작]/[기록 중지] 메뉴 클릭 시 호출. 현재 상태에 따라 분기.
#
# 【6-26】 _start_monitoring()
#   모니터링 시작:
#   ① 노선 등록 확인 → 자동 저장 파일 생성 (운행기록_YYYYMMDD_HHMMSS.xlsx)
#   ② is_monitoring=True, 메뉴 "기록 중지"로 변경, [노선 검색] 비활성화
#   ③ resume_tick() 호출, _main_loop 데몬 스레드 시작
#
# 【6-27】 _stop_monitoring()
#   모니터링 중지:
#   ① 확인 메시지 → is_monitoring=False, 메뉴 원복
#   ② pause_tick() 호출, 미저장 기록 있으면 즉시 저장
# ──────────────────────────────────────────────────────────
    def _on_toggle(self):
        if self.is_monitoring:
            self._stop_monitoring()
        else:
            self._start_monitoring()

    def _start_monitoring(self):
        if not self.routes:
            QMessageBox.warning(self, "알림", "불러온 노선이 없습니다. 먼저 노선을 검색하세요.")
            return
        fn = f"운행기록_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.auto_save_path = os.path.join(self.current_dir, fn)
        try:
            pd.DataFrame(columns=["데이터시각", "운행시작/종료", "정류소이름(번호)", "노선", "차량번호"]).to_excel(self.auto_save_path, index=False)
            self.can_auto_save = True
            self.log(f"자동 저장 파일: {fn}")
        except Exception as e:
            QMessageBox.warning(self, "오류", f"엑셀 파일 생성 실패: {e}")
            return
        self.is_monitoring = True
        self.act_toggle.setText("기록 중지")
        self.act_search.setEnabled(False)
        for p in self.route_map_panels.values():
            p.resume_tick()
        threading.Thread(target=self._main_loop, daemon=True).start()
        self.log(f"▶ 자동 기록을 시작합니다. (주기: {self.refresh_interval}초)")

    def _stop_monitoring(self):
        if QMessageBox.question(self, "중지 확인", "정말 기록을 중지하시겠습니까?") != QMessageBox.Yes:
            return
        self.is_monitoring = False
        self.act_toggle.setText("기록 시작")
        self.act_search.setEnabled(True)
        for p in self.route_map_panels.values():
            p.pause_tick()
        if self.recorded_data and self.auto_save_path and self.can_auto_save and len(self.recorded_data) > self._saved_record_count:
            self._perform_auto_save()
        self.log("■ 자동 기록을 중지합니다.")

# ──────────────────────────────────────────────────────────
# 【6-28】 _main_loop()
#   백그라운드 데몬 스레드에서 실행되는 갱신 루프.
#   is_monitoring이 True인 동안 반복:
#   ① _refresh_data() 호출
#   ② 다음 실행 시각(nc) = 현재 + refresh_interval
#   ③ 남은 시간을 0.1초 단위로 나눠 sleep (중지 명령에 빠르게 반응하기 위해)
#   ④ API 호출이 갱신 주기를 초과하면 nc를 현재 시각으로 리셋 (지연 누적 방지)
# ──────────────────────────────────────────────────────────
    def _main_loop(self):
        nc = time.time()
        while self.is_monitoring:
            self._refresh_data()
            nc += self.refresh_interval
            st = nc - time.time()
            if st < 0:
                nc = time.time()
                st = 0
            for _ in range(int(st * 10)):
                if not self.is_monitoring:
                    break
                time.sleep(0.1)

# ──────────────────────────────────────────────────────────
# 【6-29】 _refresh_data()
#   한 번의 갱신 실행. _refresh_lock으로 동시 실행 방지.
#   (이미 갱신 중이면 blocking=False로 잠금 실패 → 즉시 리턴)
#   ① _process_routes() → ② 미저장 기록 있으면 _perform_auto_save()
# ──────────────────────────────────────────────────────────
    def _refresh_data(self):
        if not self._refresh_lock.acquire(blocking=False):
            return
        try:
            self._process_routes()
            if self.recorded_data and self.auto_save_path and self.can_auto_save and len(self.recorded_data) > self._saved_record_count:
                self._perform_auto_save()
        finally:
            self._refresh_lock.release()

# ──────────────────────────────────────────────────────────
# 【6-30】 _process_routes()  ★ 운행 기록 판정 핵심 로직 ★
#
#   ① 날짜 변경 감지: 자정 이후이면 통계 초기화·pos_suspend_until 초기화
#   ② 6시간 이상 지난 departed_vehicles 항목 자동 삭제
#   ③ 각 노선에 대해:
#      [출발 판정]
#      URL_POS1 호출 → 버스 없음(운행 종료 시간대)이면:
#        - 지도 버스 제거 (sig_clear_map)
#        - 첫차 시각까지 API 호출 중지 (pos_suspend_until 설정)
#      버스 있으면:
#        - lastStnId가 첫·두 번째 정류소이면 출발 기록 (_record)
#          (40분=2400초 내 중복 방지, departed_vehicles에 등록)
#        - URL_SLST로 구간 속도 수집 → 지도에 _sect_speeds 갱신
#        - sig_update_map으로 지도 버스 위치 갱신
#      [도착 판정]
#      URL_POS2 호출 → lastStnId가 종점이고 departed_vehicles에 있는 버스면:
#        - 도착 기록 (_record)
#        - departed_vehicles에서 해당 버스 삭제
# ──────────────────────────────────────────────────────────
    def _process_routes(self):
        _now = time.time()
        now_dt = datetime.now()
        today = now_dt.date()
        if today != self._last_date:
            self._last_date = today
            self.pos_suspend_until.clear()
            self.api_stats_yesterday = self.api_stats_today.copy()
            for k in self.api_stats_today:
                self.api_stats_today[k] = 0
            self._key_limit_notified = False
            self.log("📅 날짜가 바뀌었습니다. 통계를 초기화합니다.")
        for k in [k for k, ts in self.departed_vehicles.items() if _now - ts >= 21600]:
            del self.departed_vehicles[k]
        self.temp_pos1_data = {}
        self.temp_pos2_data = {}
        for route in self.routes:
            rnm = route['rnm']
            rid = route['rid']
            sc = route['st_cnt']
            if rid in self.pos_suspend_until:
                if now_dt < self.pos_suspend_until[rid]:
                    continue
                else:
                    del self.pos_suspend_until[rid]
                    self.pos_resume_logged.discard(rid)
            if rid in self.temp_pos1_data:
                rp1 = self.temp_pos1_data[rid]
            else:
                rp1 = self.fetch_api(URL_POS1, {'busRouteId': rid})
                if rp1 is not None and not isinstance(rp1, tuple):
                    self.temp_pos1_data[rid] = rp1
            if rp1 is None:
                continue
            p1b = not isinstance(rp1, tuple) and len(rp1.findall(".//itemList")) > 0
            if not p1b:
                ins = False
                ft = route.get('first_bus_tm')
                lt = route.get('last_bus_tm')
                if ft and lt:
                    try:
                        fh, fm = map(int, ft.split(":"))
                        lh, lm = map(int, lt.split(":"))
                        ps = now_dt.replace(hour=fh, minute=fm, second=0, microsecond=0) - timedelta(minutes=5)
                        pe = now_dt.replace(hour=lh, minute=lm, second=0, microsecond=0) + timedelta(minutes=30)
                        if pe < ps:
                            pe += timedelta(days=1)
                        ins = ps <= now_dt <= pe
                    except:
                        pass
                self.sig_clear_map.emit(rnm)
                if not ins:
                    fs = ft or self._fetch_first_time(rid)
                    if fs:
                        try:
                            fhm = datetime.strptime(fs, "%H:%M")
                            base = now_dt.replace(hour=fhm.hour, minute=fhm.minute, second=0, microsecond=0)
                            if base <= now_dt:
                                base += timedelta(days=1)
                            resume = base - timedelta(minutes=5)
                            self.pos_suspend_until[rid] = resume if resume > now_dt else now_dt + timedelta(minutes=1)
                            if resume > now_dt:
                                self.log(f"💤 {rnm} 운행 종료 → 첫차 {fs} 5분 전까지 POS 정지")
                        except:
                            self.pos_suspend_until[rid] = now_dt + timedelta(minutes=30)
                    else:
                        self.pos_suspend_until[rid] = now_dt + timedelta(minutes=30)
            if p1b:
                for bus in rp1.findall(".//itemList"):
                    ls = bus.findtext("lastStnId") or ""
                    vn = bus.findtext("plainNo") or ""
                    try:
                        stm = int(bus.findtext("lastStTm") or "0")
                    except:
                        stm = 0
                    if stm <= 0:
                        continue
                    if (rid, vn) not in self.departed_vehicles:
                        self.departed_vehicles[(rid, vn)] = _now
                    ifs = route['first_st_id'] and ls == route['first_st_id']
                    iss = route['second_st_id'] and ls == route['second_st_id']
                    if not (ifs or iss):
                        continue
                    k0 = (0, rid, vn)
                    if k0 not in self.last_arrival_logs or _now - self.last_arrival_logs[k0] >= 2400:
                        ft2 = format_datetm(bus.findtext("dataTm"))
                        if ifs:
                            dn, da = route['first_nm'], route['first_ars']
                            st = f"[{dn}({da}) 출발]"
                        else:
                            dn, da = route['second_nm'], route['second_ars']
                            st = f"[{dn}({da}) 출발 - 2번째 정류소 감지]"
                        self._record(0, ft2, rnm, vn, dn, da, st)
                        self.last_arrival_logs[k0] = _now
                        self.departed_vehicles[(rid, vn)] = _now

                sect_speeds = {}
                root_slst_spd = self.fetch_api(URL_SLST, {'busRouteId': rid})
                if root_slst_spd is not None and not isinstance(root_slst_spd, tuple):
                    for item in root_slst_spd.findall(".//itemList"):
                        seq_str = (item.findtext("seq") or "0").strip()
                        spd_str = (item.findtext("sectSpd") or "").strip()
                        try:
                            seq_val = int(seq_str)
                            spd_val = float(spd_str) if spd_str else -1
                            sect_speeds[seq_val] = spd_val
                        except ValueError:
                            pass

                if rnm in self.route_map_panels:
                    self.route_map_panels[rnm]._sect_speeds = sect_speeds

                bfm = []
                for item in rp1.findall(".//itemList"):
                    bfm.append({
                        "vehId": item.findtext("vehId") or "",
                        "plainNo": item.findtext("plainNo") or "",
                        "busType": item.findtext("busType") or "",
                        "lastStnId": item.findtext("lastStnId") or "",
                        "sectDist": item.findtext("sectDist") or "0",
                        "islastyn": item.findtext("islastyn") or "",
                        "lastStTm": item.findtext("lastStTm") or "0",
                        "congestion": item.findtext("congetion") or "0"
                    })
                self.sig_update_map.emit(rnm, bfm)

            if rid in self.temp_pos2_data:
                rp2 = self.temp_pos2_data[rid]
            else:
                r2 = self.fetch_api(URL_POS2, {'busRouteId': rid, 'startOrd': '1', 'endOrd': str(sc)})
                if isinstance(r2, tuple) and r2[0] == 'NO_BUS':
                    continue
                rp2 = r2
                if rp2 is not None:
                    self.temp_pos2_data[rid] = rp2
            if rp2 is None:
                continue
            for bus in rp2.findall(".//itemList"):
                ls = bus.findtext("lastStnId") or ""
                vn = bus.findtext("plainNo") or ""
                if not (route['last_st_id'] and ls == route['last_st_id']):
                    continue
                if (rid, vn) not in self.departed_vehicles:
                    continue
                k1 = (1, rid, vn)
                if k1 not in self.last_arrival_logs or _now - self.last_arrival_logs[k1] >= 2400:
                    ft2 = format_datetm(bus.findtext("dataTm"))
                    st = f"[{route['last_nm']}({route['last_ars']}) 도착]"
                    self._record(1, ft2, rnm, vn, route['last_nm'], route['last_ars'], st)
                    self.last_arrival_logs[k1] = _now
                    del self.departed_vehicles[(rid, vn)]

# ──────────────────────────────────────────────────────────
# 【6-31】 _record(idx, ft, rnm, vn, sn, sa, status)
#   운행 이벤트(출발/도착)를 기록하는 함수.
#   ① recorded_data에 (시각, "운행시작"/"운행종료", "정류소(ARS)", 노선, 차번) 추가
#   ② _perform_auto_save() 즉시 저장
#   ③ log() 로그 출력
#   ④ sig_record 시그널 발행 → UI 표에 반영
# ──────────────────────────────────────────────────────────
    def _record(self, idx, ft, rnm, vn, sn, sa, status):
        op = "운행시작" if idx == 0 else "운행종료"
        self.recorded_data.append((ft, op, f"{sn} ({sa})", rnm, vn))
        self._perform_auto_save()
        self.log(f"★ {rnm} {vn} → {status}")
        self.sig_record.emit(idx, (ft, rnm, vn, status))

# ──────────────────────────────────────────────────────────
# 【6-32】 _perform_auto_save()
#   _save_lock으로 동시 저장 방지 후 _core_excel_save() 호출.
#   저장 조건 모두 충족해야 실행: recorded_data 비어있지 않음 +
#   auto_save_path 설정됨 + can_auto_save=True
# ──────────────────────────────────────────────────────────
    def _perform_auto_save(self):
        if not self.recorded_data or not self.auto_save_path or not self.can_auto_save:
            return
        if not self._save_lock.acquire(blocking=False):
            return
        try:
            self._core_excel_save(self.auto_save_path, True)
        finally:
            self._save_lock.release()

# ──────────────────────────────────────────────────────────
# 【6-33】 _core_excel_save(tp, sc=False)
#   운행 기록을 날짜별 시트로 분리하여 엑셀 저장.
#
#   날짜 분리 기준: 새벽 3시 이전(0~2시)은 전날 운행으로 간주
#   (심야 버스 등 자정 이후 운행을 같은 영업일로 처리)
#
#   처리 순서:
#   ① DataFrame 생성 → gbd()로 각 데이터의 영업일(BizDate) 계산
#   ② 현재 날짜(cb) 결정 (새벽 3시 이전이면 전날)
#   ③ 완결 날짜(cd) = 오늘·Unknown 제외한 모든 날짜
#   ④ ExcelWriter로 날짜별 시트에 저장 + _axs()로 스타일 적용
#   ⑤ sc=True이면 완결 날짜의 별도 완료 파일 생성
#      → "운행기록_YYYYMMDD_완료.xlsx" (_completed_dates_saved로 중복 방지)
#   ⑥ _saved_record_count 갱신
# ──────────────────────────────────────────────────────────
    def _core_excel_save(self, tp, sc=False):
        try:
            cols = ["데이터시각", "운행시작/종료", "정류소이름(번호)", "노선", "차량번호"]
            df = pd.DataFrame(self.recorded_data, columns=cols)
            def gbd(ds):
                try:
                    d = datetime.strptime(ds, "%Y-%m-%d %H:%M:%S")
                    if d.hour < 3:
                        d -= timedelta(days=1)
                    return d.strftime("%Y-%m-%d")
                except:
                    return "Unknown"
            df['BizDate'] = df['데이터시각'].apply(gbd)
            now = datetime.now()
            cb = (now - timedelta(days=1)).strftime("%Y-%m-%d") if now.hour < 3 else now.strftime("%Y-%m-%d")
            cd = set(df['BizDate'].unique()) - {cb, "Unknown"}
            with pd.ExcelWriter(tp, engine='openpyxl') as w:
                for bd, g in df.groupby('BizDate'):
                    sd = g.drop(columns=['BizDate'])
                    sd.to_excel(w, sheet_name=bd, index=False)
                    self._axs(w.sheets[bd], sd)
            if sc:
                for bd in sorted(cd - self._completed_dates_saved):
                    dd = df[df['BizDate'] == bd].drop(columns=['BizDate'])
                    if dd.empty:
                        continue
                    sd2 = bd.replace("-", "")
                    cp = os.path.join(self.current_dir, f"운행기록_{sd2}_완료.xlsx")
                    with pd.ExcelWriter(cp, engine='openpyxl') as cw:
                        dd.to_excel(cw, sheet_name=bd, index=False)
                        self._axs(cw.sheets[bd], dd)
                    self._completed_dates_saved.add(bd)
                    self.log(f"📁 완결 파일 저장: 운행기록_{sd2}_완료.xlsx")
            self._saved_record_count = len(self.recorded_data)
        except PermissionError:
            self.log("⚠ 엑셀 파일이 열려 있어 저장을 건너뜁니다.")
        except Exception as e:
            self.log(f"❌ 저장 오류: {e}")

# ──────────────────────────────────────────────────────────
# 【6-34】 _axs(ws, df)
#   엑셀 워크시트에 스타일 적용.
#   헤더 행(1행): 연파랑 배경(#DDEBF7) + 굵은 글자(11pt) + 가운데 정렬
#   컬럼 너비: max(데이터 최대 글자 수, 헤더 글자 수) + 컬럼별 가산값
#     데이터시각(+1), 운행시작/종료(+7), 정류소이름(번호)(+15), 노선(+8), 기타(+5)
# ──────────────────────────────────────────────────────────
    def _axs(self, ws, df):
        hf = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        hft = XlFont(bold=True, size=11)
        ha = Alignment(horizontal='center', vertical='center')
        cw = {
            "데이터시각": lambda b: b + 1,
            "운행시작/종료": lambda b: b + 7,
            "정류소이름(번호)": lambda b: b + 15,
            "노선": lambda b: b + 8
        }
        for ci, col in enumerate(df.columns):
            c = ws.cell(row=1, column=ci + 1)
            c.fill = hf
            c.font = hft
            c.alignment = ha
            ml = df[col].astype(str).map(len).max()
            bw = max(ml, len(str(col)))
            ws.column_dimensions[c.column_letter].width = cw.get(col, lambda b: b + 5)(bw)
            

# ──────────────────────────────────────────────────────────
# 【6-35】 _ask_interval()  ← 두 번째(유효) 정의
#   Python에서 같은 이름 메서드를 두 번 정의하면 나중 것이 사용됨.
#   이 버전이 줄 1658의 첫 번째 정의를 덮어씀.
#   기능: 갱신 주기(10초 이상)를 QLineEdit으로 입력 받는 대화상자.
# ──────────────────────────────────────────────────────────
    def _ask_interval(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("갱신주기 설정")
        dlg.setFixedSize(300, 140)
        lo = QVBoxLayout(dlg)
        lo.addWidget(QLabel("갱신주기(초)를 입력하세요 (10초 이상):"))
        ed = QLineEdit(str(self.refresh_interval))
        ed.setAlignment(Qt.AlignCenter)
        lo.addWidget(ed)
        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        lo.addWidget(bb)
        def ok():
            try:
                v = int(ed.text())
            except:
                QMessageBox.warning(dlg, "알림", "숫자를 입력해주세요.")
                return
            if v < 10:
                QMessageBox.warning(dlg, "알림", "갱신주기는 10초 이상이어야 합니다.")
                return
            self.refresh_interval = v
            self.log(f"갱신주기가 {v}초로 변경되었습니다.")
            dlg.accept()
        bb.accepted.connect(ok)
        bb.rejected.connect(dlg.reject)
        ed.returnPressed.connect(ok)
        dlg.exec()

# ──────────────────────────────────────────────────────────
# 【6-36】 _show_api_status()  ← 두 번째(유효) 정의
#   줄 1686의 첫 번째 정의를 덮어씀.
#   이 버전은 setMinimumSectionSize(60)이 추가되어
#   컬럼이 너무 좁아지지 않도록 보호함.
# ──────────────────────────────────────────────────────────
    def _show_api_status(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("API 현황")
        lo = QVBoxLayout(dlg)
        t = QTableWidget(6, 4)
        t.setHorizontalHeaderLabels(["구분", "API URL", "오늘", "어제"])
        hdr = t.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.Fixed)
        hdr.setSectionResizeMode(1, QHeaderView.Stretch)
        hdr.setSectionResizeMode(2, QHeaderView.Fixed)
        hdr.setSectionResizeMode(3, QHeaderView.Fixed)
        t.setColumnWidth(0, 145)
        t.setColumnWidth(2, 75)
        t.setColumnWidth(3, 75)
        t.setEditTriggers(QAbstractItemView.NoEditTriggers)
        t.verticalHeader().setVisible(False)
        data = [
            ("운행 출발 판정 API", "POS1", URL_POS1),
            ("종점 도착 판정 API", "POS2", URL_POS2),
            ("노선 정류소 목록 API", "SLST", URL_SLST),
            ("버스 노선 정보 API", "RINF", URL_RINF),
            ("버스 노선 검색 API", "SRCH", URL_SRCH),
            ("기타", "기타", "기타"),
        ]
        for row, (lb, ky, url) in enumerate(data):
            t.setItem(row, 0, QTableWidgetItem(lb))
            t.setItem(row, 1, QTableWidgetItem(url))
            t.setItem(row, 2, QTableWidgetItem(f"{self.api_stats_today.get(ky, 0):,}회"))
            t.setItem(row, 3, QTableWidgetItem(f"{self.api_stats_yesterday.get(ky, 0):,}회"))
        t.horizontalHeader().setMinimumSectionSize(60)
        lo.addWidget(t)
        row_h = t.verticalHeader().defaultSectionSize()
        header_h = t.horizontalHeader().height()
        total_h = header_h + row_h * 6 + 10
        t.setFixedHeight(total_h)
        dlg.resize(750, total_h + 50)
        dlg.exec()

# ──────────────────────────────────────────────────────────
# 【6-37】 _show_program_info()  ← 두 번째(유효) 정의
#   줄 1723의 첫 번째 정의를 덮어씀. 기능 동일.
# ──────────────────────────────────────────────────────────
    def _show_program_info(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("프로그램 정보")
        dlg.setMinimumSize(400, 400)
        lo = QVBoxLayout(dlg)
        lc = QApplication.palette().color(QPalette.Link).name()
        ih = (f'<p style="font-size:10pt; font-weight:bold;">서울시내버스 노선 운행기록 수집 프로그램 v1.15</p>'
              f'<p>이 프로그램은 Python + PySide6로 작성하고,<br>'
              f'공공누리 제1유형으로 개방한 공공데이터 API 서비스를 이용하였으며,<br>'
              f'API 서비스는 아래의 페이지에서 무료로 이용할 수 있습니다.</p>'
              f'<p>공공데이터포털<br><a href="https://www.data.go.kr" style="color:{lc};">https://www.data.go.kr</a></p>'
              f'<p>서울특별시_버스위치정보조회 서비스<br>'
              f'<a href="https://www.data.go.kr/data/15000332/openapi.do" style="color:{lc};">https://www.data.go.kr/data/15000332/openapi.do</a></p>'
              f'<p>서울특별시_노선정보조회 서비스<br>'
              f'<a href="https://www.data.go.kr/data/15000193/openapi.do" style="color:{lc};">https://www.data.go.kr/data/15000193/openapi.do</a></p>'
              f'<p>작성자 : 박 국 환 ( <a href="mailto:ggoyong2@naver.com" style="color:{lc};">ggoyong2@naver.com</a> )</p>')
        lb = QLabel(ih)
        lb.setFont(QFont(FONT_FAMILY, 9))
        lb.setWordWrap(True)
        lb.setOpenExternalLinks(True)
        lb.setTextFormat(Qt.RichText)
        lo.addWidget(lb)
        il = QHBoxLayout()
        for b64 in [CC_IMG_B64, GG_IMG_B64]:
            pm = load_pixmap_from_b64(b64)
            if not pm.isNull():
                il2 = QLabel()
                il2.setPixmap(pm)
                il.addWidget(il2)
        il.addStretch()
        lo.addLayout(il)
        btn = QPushButton("닫기")
        btn.clicked.connect(dlg.accept)
        lo.addWidget(btn, alignment=Qt.AlignCenter)
        dlg.exec()

# ──────────────────────────────────────────────────────────
# 【6-38】 closeEvent(event)
#   창 닫기(X 버튼 또는 종료 메뉴) 시 Qt가 자동 호출.
#   ① 종료 확인 메시지 → 아니오이면 event.ignore()로 취소
#   ② is_monitoring=False 설정 (갱신 루프 중지)
#   ③ 미저장 기록 있으면 _perform_auto_save()로 마지막 저장
#   ④ event.accept()로 창 닫기 허용
# ──────────────────────────────────────────────────────────
    def closeEvent(self, event):
        if QMessageBox.question(self, "종료 확인", "프로그램을 종료하시겠습니까?") != QMessageBox.Yes:
            event.ignore()
            return
        self.is_monitoring = False
        if self.recorded_data and self.auto_save_path and self.can_auto_save and len(self.recorded_data) > self._saved_record_count:
            self._perform_auto_save()
        event.accept()

# ──────────────────────────────────────────────────────────
# 【6-39】 _load_favorites() → list
#   설정 파일(Seoul_Bus_Config.ini)의 [favorites] 섹션에서
#   즐겨찾기 노선 목록을 JSON으로 읽어 리스트로 반환.
#   파싱 실패 시 빈 리스트 반환.
#
# 【6-40】 _save_favorites(favs)
#   즐겨찾기 리스트를 JSON으로 직렬화하여 설정 파일에 저장.
#   동시에 현재 API 키도 갱신 (favorites 저장 시 keys도 함께 기록).
# ──────────────────────────────────────────────────────────
    def _load_favorites(self):
        cfg = configparser.ConfigParser()
        if os.path.exists(self._CFG_FILE):
            cfg.read(self._CFG_FILE, encoding="utf-8")
        raw = cfg.get("favorites", "data", fallback="[]")
        try:
            return json.loads(raw)
        except:
            return []

    def _save_favorites(self, favs):
        cfg = configparser.ConfigParser()
        if os.path.exists(self._CFG_FILE):
            cfg.read(self._CFG_FILE, encoding="utf-8")
        if "keys" not in cfg:
            cfg["keys"] = {}
        cfg["keys"]["main_key"] = self._enc_key(self.api_key_main)
        cfg["keys"]["back_key"] = self._enc_key(self.api_key_back)
        if "favorites" not in cfg:
            cfg["favorites"] = {}
        cfg["favorites"]["data"] = json.dumps(favs, ensure_ascii=False)
        try:
            with open(self._CFG_FILE, "w", encoding="utf-8") as f:
                cfg.write(f)
        except:
            pass

# ──────────────────────────────────────────────────────────
# 【6-41】 _cleanup_search_signal(handler)
#   sig_search_done 시그널에서 특정 핸들러(슬롯)의 연결을 안전하게 해제.
#   대화상자가 닫힌 후 더 이상 필요 없는 슬롯이 계속 연결되어
#   의도치 않게 호출되는 것을 방지.
# ──────────────────────────────────────────────────────────
    def _cleanup_search_signal(self, handler):
        try:
            self.sig_search_done.disconnect(handler)
        except:
            pass

# ══════════════════════════════════════════════════════════
# 【7】 프로그램 진입점
#   파이썬 파일을 직접 실행했을 때만 이 블록이 실행됨.
#   다른 파일에서 import할 때는 실행되지 않음.
# ══════════════════════════════════════════════════════════
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    mode = "dark" if detect_os_dark_mode() else "light"
    app.setPalette(_make_palette(mode))
    window = SeoulBusRecorder()
    window.show()
    sys.exit(app.exec())